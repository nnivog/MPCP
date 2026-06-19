"""
Sipradi SC-MPCP Control System v3.0
Flask + SQLite | python app.py | http://localhost:5050
Patches: cascade_links API, perf/quick entry, FY auto-detect,
         duplicate guard, SQL analytics, global error handlers
"""

from flask import Flask, jsonify, request, send_file, session, redirect, render_template_string
import sqlite3, csv, io, os, json, datetime, random, string, hashlib, secrets

# ── HF Spaces Persistent Storage Config ──────────────────────────────────────
import os as _os
import shutil as _shutil

_DATA_DIR = "/data" if _os.path.exists("/data") else "."

# All DB paths — edit here if you add more databases
_DB_MAP = {
    "master.db":  _os.path.join(_DATA_DIR, "master.db"),
    "scm.db":     _os.path.join(_DATA_DIR, "scm.db"),
}

# Seed on first boot: copy bundled .db files into /data if not present
for _src, _dst in _DB_MAP.items():
    if not _os.path.exists(_dst) and _os.path.exists(_src):
        _shutil.copy(_src, _dst)
        print(f"[BOOT] Seeded {_src} → {_dst}")

def _db(name: str) -> str:
    """Return the correct DB path for local or HF environment."""
    return _DB_MAP.get(name, _os.path.join(_DATA_DIR, name))
# ─────────────────────────────────────────────────────────────────────────────


try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
# Use stable key — generate once and save to data/secret.key
_key_path = os.path.join(os.path.dirname(__file__), 'data', 'secret.key')
os.makedirs(os.path.dirname(_key_path), exist_ok=True)
if os.path.exists(_key_path):
    with open(_key_path,'r') as _f: app.secret_key = _f.read().strip()
else:
    app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
    with open(_key_path,'w') as _f: _f.write(app.secret_key)

# ── DATA DIR & DB PATHS ────────────────────────────────────────────────────
DATA_DIR   = os.environ.get('MPCP_DATA_DIR', os.path.join(os.path.dirname(__file__), 'data'))
os.makedirs(DATA_DIR, exist_ok=True)
MASTER_DB  = os.path.join(DATA_DIR, 'master.db')
DB         = os.path.join(DATA_DIR, 'scm.db')  # points to data dir

def get_dept_db_path(dept_code):
    return os.path.join(DATA_DIR, f'{dept_code}.db')

def get_master_conn():
    c = sqlite3.connect(MASTER_DB)
    c.row_factory = sqlite3.Row
    return c

# ── PASSWORD UTILS ─────────────────────────────────────────────────────────

def log_audit(action, target_type='', target_id='', detail=''):
    try:
        import datetime as _dt
        mdb = get_master_conn()
        u = session.get('mpcp_user') or {}
        mdb.execute(
            "INSERT INTO audit_log(ts,actor_id,actor_name,action,target_type,target_id,detail,ip) VALUES(?,?,?,?,?,?,?,?)",
            (_dt.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
             u.get('id',''), u.get('full_name','system'),
             action, target_type, str(target_id), detail,
             request.remote_addr if request else '')
        )
        mdb.commit()
        mdb.close()
    except Exception:
        pass


def hash_password(pw):
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 260000)
    return salt + ':' + h.hex()

def verify_password(pw, stored):
    try:
        salt, h = stored.split(':',1)
        check = hashlib.pbkdf2_hmac('sha256', pw.encode(), salt.encode(), 260000)
        return check.hex() == h
    except: return False

# ── MASTER DB INIT ─────────────────────────────────────────────────────────
def seed_masters(db):
    import random as _r, string as _s
    def _uid(): return ''.join(_r.choices(_s.ascii_lowercase+_s.digits, k=8))
    defaults = [
        ('frequency','Daily','Daily',1),
        ('frequency','Weekly','Weekly',2),
        ('frequency','Fortnightly','Fortnightly',3),
        ('frequency','Monthly','Monthly',4),
        ('frequency','Quarterly','Quarterly',5),
        ('unit','%','Percentage',1),
        ('unit','Days','Days',2),
        ('unit','Hours','Hours',3),
        ('unit','Nos','Numbers',4),
        ('unit','Amt','Amount',5),
        ('unit','Units','Units',6),
        ('unit','Visits','Visits',7),
        ('emp_level','1','Head of Department',1),
        ('emp_level','2','Team Lead / Manager',2),
        ('emp_level','3','Operations Staff',3),
        ('cp_source','','General',1),
        ('cp_source','System','System',2),
        ('cp_source','Manual','Manual',3),
        ('cp_source','Check/System','Check/System',4),
        ('location_type','Branch','Branch',1),
        ('location_type','Office','Office',2),
        ('location_type','Warehouse','Warehouse',3),
        ('location_type','Workshop','Workshop',4),
    ]
    existing = db.execute("SELECT COUNT(*) FROM masters").fetchone()[0]
    if existing > 0: return
    for cat,val,lbl,srt in defaults:
        try:
            db.execute("INSERT OR IGNORE INTO masters(id,category,value,label,sort_order) VALUES(?,?,?,?,?)",
                      (_uid(),cat,val,lbl,srt))
        except: pass
    db.commit()

def init_master_db():
    db = get_master_conn()
    db.executescript("""
CREATE TABLE IF NOT EXISTS departments(
  id TEXT PRIMARY KEY,
  code TEXT UNIQUE NOT NULL,
  name TEXT NOT NULL,
  active INTEGER DEFAULT 1,
  created_at TEXT DEFAULT '');

CREATE TABLE IF NOT EXISTS users(
  id TEXT PRIMARY KEY,
  username TEXT UNIQUE NOT NULL,
  password_hash TEXT NOT NULL,
  full_name TEXT NOT NULL,
  role TEXT NOT NULL DEFAULT 'user',
  dept_code TEXT DEFAULT NULL,
  emp_code TEXT DEFAULT '',
  active INTEGER DEFAULT 1,
  created_at TEXT DEFAULT '');
CREATE TABLE IF NOT EXISTS masters(
  id TEXT PRIMARY KEY,
  category TEXT NOT NULL,
  value TEXT NOT NULL,
  label TEXT DEFAULT '',
  sort_order INTEGER DEFAULT 0,
  active INTEGER DEFAULT 1,
  UNIQUE(category, value)
);
CREATE TABLE IF NOT EXISTS audit_log(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  ts TEXT NOT NULL,
  actor_id TEXT,
  actor_name TEXT,
  action TEXT NOT NULL,
  target_type TEXT,
  target_id TEXT,
  detail TEXT,
  ip TEXT);
""")
    # Create default master admin if no users exist
    existing = db.execute("SELECT COUNT(*) c FROM users").fetchone()['c']
    if existing == 0:
        import datetime as _dt
        pid = ''.join(random.choices(string.ascii_lowercase+string.digits, k=8))
        db.execute("INSERT INTO users VALUES(?,?,?,?,?,?,?,?,?)",
            (pid,'admin', hash_password('admin123'),
             'Master Admin','master_admin',None,'',1,
             _dt.datetime.now().isoformat()))
        db.commit()
        print("  Created default admin user: admin / admin123")
    db.commit()
    db.close()

init_master_db()

# ── AUTH MIDDLEWARE ───────────────────────────────────────────────────────
PUBLIC_PATHS = {'/login', '/logout', '/static'}

@app.before_request
def auth_guard():
    if request.path.startswith('/static'): return
    if request.path in ('/login','/logout','/change_password'): return
    if not session.get('mpcp_user'):
        if request.path.startswith('/api/'):
            return jsonify({'error':'Not authenticated','redirect':'/login'}),401
        return redirect('/login')

# ── GLOBAL ERROR HANDLERS
@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Route not found', 'path': request.path}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': str(e)}), 500

# ── NEPALI CALENDAR ────────────────────────────────────────────────────────
BS_MONTHS = ["Shrawan","Bhadra","Ashwin","Kartik","Mangsir","Poush",
             "Magh","Falgun","Chaitra","Baisakh","Jestha","Ashadh"]

BS_Q = {"Shrawan":"Q1","Bhadra":"Q1","Ashwin":"Q1","Kartik":"Q2","Mangsir":"Q2","Poush":"Q2",
        "Magh":"Q3","Falgun":"Q3","Chaitra":"Q3","Baisakh":"Q4","Jestha":"Q4","Ashadh":"Q4"}

AD_TO_BS = {
    "Jul":"Shrawan","Aug":"Bhadra","Sep":"Ashwin","Oct":"Kartik","Nov":"Mangsir","Dec":"Poush",
    "Jan":"Magh","Feb":"Falgun","Mar":"Chaitra","Apr":"Baisakh","May":"Jestha","Jun":"Ashadh",
    "July":"Shrawan","August":"Bhadra","September":"Ashwin","October":"Kartik",
    "November":"Mangsir","December":"Poush","January":"Magh","February":"Falgun",
    "March":"Chaitra","April":"Baisakh","June":"Ashadh"}

# Nepali FY runs Shrawan→Ashadh (~Jul 16 → Jul 15 next year)
FY_MAP = [
    (datetime.date(2022, 7, 17), datetime.date(2023, 7, 16), "2079-80"),
    (datetime.date(2023, 7, 17), datetime.date(2024, 7, 15), "2080-81"),
    (datetime.date(2024, 7, 16), datetime.date(2025, 7, 15), "2081-82"),
    (datetime.date(2025, 7, 16), datetime.date(2026, 7, 15), "2082-83"),
    (datetime.date(2026, 7, 16), datetime.date(2027, 7, 15), "2083-84"),
]

def ad_date_to_fy_and_month(dt):
    d = dt.date() if isinstance(dt, datetime.datetime) else dt
    for start, end, fy in FY_MAP:
        if start <= d <= end:
            return fy, AD_TO_BS.get(dt.strftime('%B'), 'Shrawan')
    return '2081-82', 'Shrawan'

def norm_month(m):
    m = str(m or '').strip()
    if m in BS_MONTHS: return m
    if m in AD_TO_BS: return AD_TO_BS[m]
    for bs in BS_MONTHS:
        if m.lower() == bs.lower(): return bs
    for ad, bs in AD_TO_BS.items():
        if m.lower() == ad.lower(): return bs
    return m or "Shrawan"

def bs_q(month): return BS_Q.get(month, "Q1")
def uid(): return ''.join(random.choices(string.ascii_lowercase+string.digits, k=8))

def get_db(dept_override=None):
    user = session.get('mpcp_user')
    if not user:
        conn = sqlite3.connect(DB); conn.row_factory = sqlite3.Row; return conn
    role = user.get('role','user')
    if role == 'master_admin':
        dept = dept_override or request.args.get('dept') or user.get('active_dept')
        path = get_dept_db_path(dept) if dept else DB
    else:
        path = get_dept_db_path(user['dept_code'])
    if not os.path.exists(path): _init_dept_db(path)
    conn = sqlite3.connect(path); conn.row_factory = sqlite3.Row; return conn

def R(rows): return [dict(r) for r in rows]

def calc_status(actual, target, unit):
    try: a, t = float(actual), float(target)
    except: return "C"
    lb = any(u in str(unit).lower() for u in ["day","hour","hr"])
    return "C" if (a <= t*1.05 if lb else a >= t*0.95) else "NC"

def validate_required(data, *fields):
    return [f for f in fields if not data.get(f)]

def json_error(msg, code=400):
    return jsonify({'error': msg}), code

# ── SCHEMA ─────────────────────────────────────────────────────────────────
SCHEMA = """
CREATE TABLE IF NOT EXISTS employees(
  id TEXT PRIMARY KEY, emp_code TEXT UNIQUE,
  name TEXT NOT NULL, role TEXT DEFAULT '', level INTEGER DEFAULT 3,
  dept TEXT DEFAULT 'Ops', manager_id TEXT, email TEXT DEFAULT '');

CREATE TABLE IF NOT EXISTS mps(
  id TEXT PRIMARY KEY, ref TEXT NOT NULL, title TEXT NOT NULL,
  target TEXT DEFAULT '', freq TEXT DEFAULT 'Monthly',
  kpi_c INTEGER DEFAULT 0, kpi_nc INTEGER DEFAULT 0, kpi_total INTEGER DEFAULT 0);

CREATE TABLE IF NOT EXISTS mp_owners(mp_id TEXT, emp_id TEXT, PRIMARY KEY(mp_id,emp_id));

CREATE TABLE IF NOT EXISTS cps(
  id TEXT PRIMARY KEY, ref TEXT NOT NULL, title TEXT NOT NULL,
  target TEXT DEFAULT '', freq TEXT DEFAULT 'Daily',
  source TEXT DEFAULT '', mp_id TEXT DEFAULT '');

CREATE TABLE IF NOT EXISTS cp_owners(cp_id TEXT, emp_id TEXT, PRIMARY KEY(cp_id,emp_id));

CREATE TABLE IF NOT EXISTS roles(
  id TEXT PRIMARY KEY, code TEXT NOT NULL, name TEXT NOT NULL,
  description TEXT DEFAULT '', color TEXT DEFAULT '#1d4ed8');

CREATE TABLE IF NOT EXISTS role_mps(role_id TEXT, mp_id TEXT, PRIMARY KEY(role_id,mp_id));
CREATE TABLE IF NOT EXISTS role_cps(role_id TEXT, cp_id TEXT, PRIMARY KEY(role_id,cp_id));
CREATE TABLE IF NOT EXISTS emp_roles(emp_id TEXT, role_id TEXT, PRIMARY KEY(emp_id,role_id));
CREATE TABLE IF NOT EXISTS emp_mps(emp_id TEXT, mp_id TEXT, PRIMARY KEY(emp_id,mp_id));
CREATE TABLE IF NOT EXISTS emp_cps(emp_id TEXT, cp_id TEXT, PRIMARY KEY(emp_id,cp_id));

CREATE TABLE IF NOT EXISTS perf(
  id TEXT PRIMARY KEY, fy TEXT NOT NULL, bs_month TEXT NOT NULL,
  quarter TEXT DEFAULT 'Q1', emp_id TEXT DEFAULT '', emp_code TEXT DEFAULT '',
  mp_ref TEXT DEFAULT '', cp_ref TEXT DEFAULT '', metric TEXT DEFAULT '',
  total INTEGER DEFAULT 0, compliant INTEGER DEFAULT 0,
  non_compliant INTEGER DEFAULT 0,
  pct_compliant REAL DEFAULT 0, pct_nc REAL DEFAULT 0,
  target_val REAL DEFAULT 0, actual_val REAL DEFAULT 0,
  unit TEXT DEFAULT '%', status TEXT DEFAULT 'C', notes TEXT DEFAULT '');


CREATE TABLE IF NOT EXISTS emp_sectors(
  emp_id TEXT NOT NULL,
  sector_id TEXT NOT NULL,
  is_primary INTEGER DEFAULT 0,
  PRIMARY KEY(emp_id,sector_id));

CREATE TABLE IF NOT EXISTS emp_locations(
  emp_id TEXT NOT NULL,
  loc_id TEXT NOT NULL,
  is_primary INTEGER DEFAULT 0,
  PRIMARY KEY(emp_id,loc_id));
CREATE TABLE IF NOT EXISTS perf_cache(
  fy TEXT PRIMARY KEY, label TEXT NOT NULL,
  record_count INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT, locked INTEGER DEFAULT 0);

CREATE TABLE IF NOT EXISTS cascade_links(
  id TEXT PRIMARY KEY,
  superior_emp_id    TEXT NOT NULL,
  superior_cp_id     TEXT NOT NULL,
  subordinate_emp_id TEXT NOT NULL,
  subordinate_mp_id  TEXT DEFAULT '',
  auto_created       INTEGER DEFAULT 0,
  created_at         TEXT DEFAULT '');
"""

def init_db():
    with get_db() as db:
        db.executescript(SCHEMA)
        # migrate: add photo column if missing
        cols = [r[1] for r in db.execute("PRAGMA table_info(employees)")]
        if 'photo' not in cols:
            db.execute("ALTER TABLE employees ADD COLUMN photo TEXT DEFAULT ''")
            db.commit()
        if not db.execute("SELECT 1 FROM employees LIMIT 1").fetchone():
            _seed(db)

        # ── Schema migrations: add columns missing in older DB versions ──────────
        def _migrate(table, migrations):
            existing = {row[1] for row in db.execute(f"PRAGMA table_info({table})").fetchall()}
            for col, sql in migrations:
                if col not in existing:
                    try: db.execute(sql)
                    except Exception: pass

        _migrate("perf", [
            ("quarter",       "ALTER TABLE perf ADD COLUMN quarter TEXT DEFAULT 'Q1'"),
            ("emp_id",        "ALTER TABLE perf ADD COLUMN emp_id TEXT DEFAULT ''"),
            ("emp_code",      "ALTER TABLE perf ADD COLUMN emp_code TEXT DEFAULT ''"),
            ("mp_ref",        "ALTER TABLE perf ADD COLUMN mp_ref TEXT DEFAULT ''"),
            ("metric",        "ALTER TABLE perf ADD COLUMN metric TEXT DEFAULT ''"),
            ("non_compliant", "ALTER TABLE perf ADD COLUMN non_compliant INTEGER DEFAULT 0"),
            ("pct_nc",        "ALTER TABLE perf ADD COLUMN pct_nc REAL DEFAULT 0"),
            ("target_val",    "ALTER TABLE perf ADD COLUMN target_val REAL DEFAULT 0"),
            ("actual_val",    "ALTER TABLE perf ADD COLUMN actual_val REAL DEFAULT 0"),
            ("unit",          "ALTER TABLE perf ADD COLUMN unit TEXT DEFAULT '%'"),
            ("status",        "ALTER TABLE perf ADD COLUMN status TEXT DEFAULT 'C'"),
            ("notes",         "ALTER TABLE perf ADD COLUMN notes TEXT DEFAULT ''"),
        ])
        _migrate("perf_cache", [
            ("locked", "ALTER TABLE perf_cache ADD COLUMN locked INTEGER DEFAULT 0"),
        ])
        _migrate("employees", [
            ("email", "ALTER TABLE employees ADD COLUMN email TEXT DEFAULT ''"),
            ("photo", "ALTER TABLE employees ADD COLUMN photo TEXT DEFAULT ''"),
        ])
        _migrate("cps", [
            ("source", "ALTER TABLE cps ADD COLUMN source TEXT DEFAULT ''"),
            ("mp_id",  "ALTER TABLE cps ADD COLUMN mp_id TEXT DEFAULT ''"),
            ("status", "ALTER TABLE cps ADD COLUMN status TEXT DEFAULT 'active'"),
        ])
        _migrate("mps", [
            ("status", "ALTER TABLE mps ADD COLUMN status TEXT DEFAULT 'active'"),
        ])
        _migrate("roles", [
            ("status", "ALTER TABLE roles ADD COLUMN status TEXT DEFAULT 'active'"),
        ])

        # cascade_links: old DB may use parent_/child_ naming instead of superior_/subordinate_
        cl_cols = {row[1] for row in db.execute("PRAGMA table_info(cascade_links)").fetchall()}
        if cl_cols and 'superior_emp_id' not in cl_cols:
            # Old schema — rebuild with new column names preserving data
            try:
                db.execute("""CREATE TABLE IF NOT EXISTS cascade_links_new(
                  id TEXT PRIMARY KEY,
                  superior_emp_id    TEXT NOT NULL DEFAULT '',
                  superior_cp_id     TEXT NOT NULL DEFAULT '',
                  subordinate_emp_id TEXT NOT NULL DEFAULT '',
                  subordinate_mp_id  TEXT DEFAULT '',
                  auto_created       INTEGER DEFAULT 0,
                  created_at         TEXT DEFAULT '')""")
                # Map old columns to new — try both naming conventions
                old_sup  = 'parent_emp_id'  if 'parent_emp_id'  in cl_cols else ('sup_emp_id'  if 'sup_emp_id'  in cl_cols else 'superior_emp_id')
                old_cp   = 'parent_cp_id'   if 'parent_cp_id'   in cl_cols else ('sup_cp_id'   if 'sup_cp_id'   in cl_cols else 'superior_cp_id')
                old_sub  = 'child_emp_id'   if 'child_emp_id'   in cl_cols else ('sub_emp_id'  if 'sub_emp_id'  in cl_cols else 'subordinate_emp_id')
                old_mp   = 'child_mp_id'    if 'child_mp_id'    in cl_cols else ('sub_mp_id'   if 'sub_mp_id'   in cl_cols else 'subordinate_mp_id')
                old_auto = 'auto_created'   if 'auto_created'   in cl_cols else '0'
                old_ts   = 'created_at'     if 'created_at'     in cl_cols else "datetime('now')"
                db.execute(f"""INSERT OR IGNORE INTO cascade_links_new
                    (id, superior_emp_id, superior_cp_id, subordinate_emp_id,
                     subordinate_mp_id, auto_created, created_at)
                    SELECT id, {old_sup}, {old_cp}, {old_sub}, {old_mp}, {old_auto}, {old_ts}
                    FROM cascade_links""")
                db.execute("DROP TABLE cascade_links")
                db.execute("ALTER TABLE cascade_links_new RENAME TO cascade_links")
            except Exception as e:
                pass  # If migration fails, table stays as-is; new rows will fail gracefully
        elif not cl_cols:
            pass  # Table doesn't exist yet — SCHEMA will create it
        else:
            # Table exists with correct schema — add any missing columns
            _migrate("cascade_links", [
                ("subordinate_mp_id", "ALTER TABLE cascade_links ADD COLUMN subordinate_mp_id TEXT DEFAULT ''"),
                ("auto_created",      "ALTER TABLE cascade_links ADD COLUMN auto_created INTEGER DEFAULT 0"),
                ("created_at",        "ALTER TABLE cascade_links ADD COLUMN created_at TEXT DEFAULT ''"),
            ])

        db.commit()

        # ── Ensure current FY exists in cache ─────────────────────────────────────
        import datetime as _dt
        _today = _dt.date.today()
        for _start, _end, _fy in FY_MAP:
            if _start <= _today <= _end:
                _now = _dt.datetime.now().isoformat()
                db.execute("INSERT OR IGNORE INTO perf_cache VALUES(?,?,?,?,?,?)",
                           (_fy, f"FY {_fy}", 0, _now, _now, 0))
                db.commit()
                break

def _upd_cache(db, fy):
    cnt = db.execute("SELECT COUNT(*) FROM perf WHERE fy=?", (fy,)).fetchone()[0]
    now = datetime.datetime.now().isoformat()
    db.execute("INSERT OR IGNORE INTO perf_cache VALUES(?,?,?,?,?,?)",
               (fy, f"FY {fy}", cnt, now, now, 0))
    db.execute("UPDATE perf_cache SET record_count=?,updated_at=? WHERE fy=?", (cnt, now, fy))

def _seed(db):
    emps = [
        ("e1","EMP-001","Govinda Upadhyay","Head of Department",1,"HOD",None,"govinda@sipradi.com"),
        ("e2","EMP-002","Sandeep Gupta Kanu","Vehicle Import & Logistics",2,"Vehicle","e1","sandeep@sipradi.com"),
        ("e3","EMP-003","Sagar Koirala","Custom Clearance Support",2,"Vehicle","e1","sagar@sipradi.com"),
        ("e4","EMP-004","Rajeshwori Maharjan","Registration & WOW Manager",2,"Registration","e1","rajeshwori@sipradi.com"),
        ("e5","EMP-005","Laxmi","Vehicle Registration Exec.",3,"Registration","e4","laxmi@sipradi.com"),
        ("e6","EMP-006","Ishwor Shrestha","Registration Process",3,"Registration","e4","ishwor@sipradi.com"),
        ("e7","EMP-007","Hari Prasad Pandey","Customs Agent Coord.",3,"Vehicle","e2","hari@sipradi.com"),
        ("e8","EMP-008","Aashish Patel","Physical Stock Verification",3,"Stock","e1","aashish@sipradi.com"),
        ("e9","EMP-009","Bipin KC","Vehicle Delivery",3,"Vehicle","e2","bipin@sipradi.com"),
        ("e10","EMP-010","Mahesh Adhikari","Vehicle Movement",3,"Vehicle","e2","mahesh@sipradi.com"),
        ("e11","EMP-011","Rose Basnet","Goods Receipt & GRN",2,"Warehouse","e1","rose@sipradi.com"),
        ("e12","EMP-012","Sanjay Rauniyar","Warehouse Receipt",3,"Warehouse","e11","sanjay@sipradi.com"),
        ("e13","EMP-013","Rabindra","Goods Clearance",3,"Warehouse","e11","rabindra@sipradi.com"),
        ("e14","EMP-014","Ranjit Chhetri","Goods Dispatch",3,"Warehouse","e11","ranjit@sipradi.com"),
        ("e15","EMP-015","Purshottam Kr. Patel","Border Clearance",3,"Warehouse","e11","purshottam@sipradi.com"),
        ("e16","EMP-016","Babim / Manoj","Operations Support",3,"Ops","e1","babim@sipradi.com"),
    ]
    db.executemany("INSERT OR IGNORE INTO employees VALUES(?,?,?,?,?,?,?,?)", emps)

    mps = [
        ("mp1","HODL-1","Fully compliant and timely vehicle import","1 Day","Monthly",3183,116,3299),
        ("mp2","HODL-2","100% NIAC Claim resolution within 120 Days","120 Days","Fortnightly",0,0,0),
        ("mp3","HODL-3","Error and Penalty free registration process","15 Days","Fortnightly",1420,219,1639),
        ("mp4","HODL-4","100% ownership requests timebound","3 Days","Weekly",1531,108,1639),
        ("mp5","HODL-5","Vehicle readiness and driver arrangements within 24hrs","24 Hours","Fortnightly",3396,70,3466),
        ("mp6","HODL-6","Vehicle delivery hygiene and cost control","24 Hours","Quarterly",5580,300,5880),
        ("mp7","HODL-7","100% WOW update within 15 days CV / 10 days PV","10-15 Days","Monthly",0,0,0),
        ("mp8","HODL-8","Monthly physical verification of new vehicles","Monthly","Monthly",174,0,174),
        ("mp9","HODL-9","Goods availability within 5 days of border arrival","5 Days","Fortnightly",127,9,136),
        ("mp10","HODL-10","Performance visibility and cost optimization 20% reduction","20% Reduction","Monthly",1194,445,1639),
        ("mp11","HODL-11","Employee motivation and process capability","Per Calendar","Quarterly",0,0,0),
        ("mp12","HODL-12","Strengthen monitoring and review mechanism","Fortnightly","Fortnightly",0,0,0),
    ]
    db.executemany("INSERT OR IGNORE INTO mps VALUES(?,?,?,?,?,?,?,?)", mps)

    mpo = [("mp1","e2"),("mp1","e3"),("mp2","e2"),("mp2","e3"),("mp3","e2"),("mp3","e4"),
           ("mp4","e4"),("mp4","e2"),("mp5","e2"),("mp5","e3"),("mp5","e10"),("mp6","e2"),
           ("mp6","e3"),("mp6","e10"),("mp7","e4"),("mp8","e4"),("mp8","e8"),("mp9","e2"),
           ("mp9","e11"),("mp9","e14"),("mp10","e4"),("mp10","e2"),("mp10","e11"),("mp12","e4")]
    db.executemany("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", mpo)

    cps = [
        ("cp1","LM-VEH-1","Tracking of New Vehicles till arrival at border","100%","Monthly","Manual","mp1"),
        ("cp2","LM-VEH-2-B","On time custom clearance within 3 days of CC memo","3 Days","Daily","System","mp1"),
        ("cp3","LM-VEH-2-S","Custom clearance monitoring","3 Days","Daily","System","mp1"),
        ("cp4","LM-VEH-3-B","100% vehicles with valid GST before expiry","90 Days","Weekly","System","mp1"),
        ("cp5","LM-VEH-3-S","GST validity monitoring","90 Days","Weekly","System","mp1"),
        ("cp6","LM-VEH-4-B","100% POE submission within 90 days of CC","90 Days","Weekly","Manual","mp1"),
        ("cp7","LM-VEH-4-S","POE submission monitoring","90 Days","Weekly","","mp1"),
        ("cp8","LM-VEH-5-B","Registration of Vehicles within 15 Days of PP","15 Days","Weekly","System","mp3"),
        ("cp9","LM-VEH-5-A","Registration tracking","15 Days","Weekly","","mp3"),
        ("cp10","LM-VEH-6-B","Ownership Transfer within 3 Days","3 Days","Daily","System","mp4"),
        ("cp11","LM-VEH-6-A","Ownership transfer support","3 Days","Daily","","mp4"),
        ("cp12","LM-VEH-7-B","Final Delivery of Vehicles within 24 Hours","24 Hours","Daily","System","mp5"),
        ("cp13","LM-VEH-7-S","Final Delivery coordination","24 Hours","Daily","","mp5"),
        ("cp14","LM-VEH-7-A","Final Delivery support","24 Hours","Daily","","mp5"),
        ("cp15","LM-VEH-8-B","Vehicle Dispatch to Locations within 24 hrs","24 Hours","Daily","","mp6"),
        ("cp16","LM-VEH-8-S","Dispatch coordination","24 Hours","Daily","System","mp6"),
        ("cp17","LM-VEH-8-A","Dispatch support","24 Hours","Daily","","mp6"),
        ("cp18","LM-VEH-11","WOW update within 10/15 Days of Delivery PV/CV","10/15 Days","Weekly","System","mp7"),
        ("cp19","LM-VEH-12","Monthly Physical Verification of new vehicles","Monthly","Monthly","Manual","mp8"),
        ("cp20","LM-WH-1","Custom Clearance of Goods within 2 working days","2 Days","Daily","Manual","mp9"),
        ("cp21","LM-WH-2","GRN and Verification of Goods within 2 Days of CC","2 Days","Daily","System","mp9"),
        ("cp22","LM-WH-3-B","Goods delivered as per SLA","SLA","Daily","System","mp9"),
        ("cp23","LM-WH-3-A","Goods delivery SLA monitoring","SLA","Daily","System","mp9"),
        ("cp24","LM-WH-4-B","Handling discrepancies 100% closure","100%","Daily","","mp10"),
        ("cp25","LM-WH-4-A","Discrepancy closure","100%","Daily","","mp10"),
        ("cp26","LM-VEH-13","Sahamati Just in Time for tax reduction","20% Reduction","Daily","","mp10"),
        ("cp27","LM-VH-14","Conduct Periodic Review Meeting","Fortnightly","Fortnightly","","mp12"),
    ]
    db.executemany("INSERT OR IGNORE INTO cps VALUES(?,?,?,?,?,?,?)", cps)

    cpo = [("cp1","e2"),("cp2","e2"),("cp3","e3"),("cp4","e2"),("cp5","e3"),("cp6","e2"),("cp7","e3"),
           ("cp8","e2"),("cp9","e4"),("cp10","e4"),("cp11","e2"),("cp12","e2"),("cp13","e3"),("cp14","e10"),
           ("cp15","e2"),("cp16","e3"),("cp17","e10"),("cp18","e4"),("cp19","e4"),("cp20","e2"),
           ("cp21","e2"),("cp22","e2"),("cp23","e11"),("cp24","e2"),("cp25","e11"),("cp26","e4"),("cp27","e4")]
    db.executemany("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", cpo)

    roles = [
        ("r1","ROLE-VIM","Vehicle Import Manager","Full vehicle import, GST, NIAC, CC and POE management","#1d4ed8"),
        ("r2","ROLE-VDM","Vehicle Delivery Manager","Manages final delivery, dispatch and driver coordination","#0891b2"),
        ("r3","ROLE-REG","Registration Manager","Vehicle registration, ownership transfer and WOW updates","#6d28d9"),
        ("r4","ROLE-WH","Warehouse & Goods Manager","Goods clearance, GRN, dispatch and discrepancy closure","#047857"),
        ("r5","ROLE-STCK","Stock Verification Officer","Physical vehicle stock verification and reporting","#b45309"),
    ]
    db.executemany("INSERT OR IGNORE INTO roles VALUES(?,?,?,?,?)", roles)

    role_mps = [("r1","mp1"),("r1","mp2"),("r2","mp5"),("r2","mp6"),("r3","mp3"),("r3","mp4"),
                ("r3","mp7"),("r4","mp9"),("r4","mp10"),("r5","mp8")]
    db.executemany("INSERT OR IGNORE INTO role_mps VALUES(?,?)", role_mps)

    role_cps = [("r1","cp1"),("r1","cp2"),("r1","cp3"),("r1","cp4"),("r1","cp5"),("r1","cp6"),("r1","cp7"),
                ("r2","cp12"),("r2","cp13"),("r2","cp14"),("r2","cp15"),("r2","cp16"),("r2","cp17"),
                ("r3","cp8"),("r3","cp9"),("r3","cp10"),("r3","cp11"),("r3","cp18"),
                ("r4","cp20"),("r4","cp21"),("r4","cp22"),("r4","cp23"),("r4","cp24"),("r4","cp25"),
                ("r5","cp19")]
    db.executemany("INSERT OR IGNORE INTO role_cps VALUES(?,?)", role_cps)

    emp_roles = [("e2","r1"),("e2","r2"),("e3","r1"),("e4","r3"),("e11","r4"),("e8","r5")]
    db.executemany("INSERT OR IGNORE INTO emp_roles VALUES(?,?)", emp_roles)

    def P(pid,fy,bsm,eid,ec,mpr,cpr,metric,tot,comp,tgt,act,unit,notes=""):
        nc=tot-comp; pct_c=round(comp/tot*100,2) if tot else 0; pct_nc=round(nc/tot*100,2) if tot else 0
        st=calc_status(act,tgt,unit)
        return (pid,fy,bsm,bs_q(bsm),eid,ec,mpr,cpr,metric,tot,comp,nc,pct_c,pct_nc,tgt,act,unit,st,notes)

    perf = [
        P("p01","2080-81","Shrawan","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",250,240,100,96,"%","All tracked"),
        P("p02","2080-81","Shrawan","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",280,265,3,2.8,"Days","On track"),
        P("p03","2080-81","Shrawan","e3","EMP-003","HODL-1","LM-VEH-2-S","CC Monitoring",180,170,3,3.2,"Days","1 delay"),
        P("p04","2080-81","Shrawan","e4","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",320,305,15,14,"Days","Within SLA"),
        P("p05","2080-81","Shrawan","e4","EMP-004","HODL-4","LM-VEH-6-B","Ownership Transfer",180,172,3,2.5,"Days","Good"),
        P("p06","2080-81","Shrawan","e11","EMP-011","HODL-9","LM-WH-3-A","Goods SLA Delivery",95,95,2,2,"Days","On SLA"),
        P("p07","2080-81","Bhadra","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",310,304,100,98,"%",""),
        P("p08","2080-81","Bhadra","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",310,295,3,2.5,"Days",""),
        P("p09","2080-81","Bhadra","e4","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",295,279,15,16,"Days","Slight delay"),
        P("p10","2080-81","Bhadra","e11","EMP-011","HODL-9","LM-WH-3-A","Goods SLA Delivery",88,72,2,2.8,"Days","Delayed"),
        P("p11","2080-81","Ashwin","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",420,420,100,100,"%","100%"),
        P("p12","2080-81","Ashwin","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",420,409,3,2.3,"Days","Best month"),
        P("p13","2080-81","Kartik","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",380,357,100,94,"%","4 missed"),
        P("p14","2080-81","Kartik","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",380,342,3,3.5,"Days","Over SLA"),
        P("p15","2080-81","Kartik","e4","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",340,329,15,13,"Days","Good"),
        P("p16","2080-81","Mangsir","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",350,340,100,97,"%",""),
        P("p17","2080-81","Mangsir","e4","EMP-004","HODL-4","LM-VEH-6-B","Ownership Transfer",200,195,3,3,"Days","On target"),
        P("p18","2080-81","Poush","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",280,258,100,92,"%","Year end"),
        P("p19","2080-81","Poush","e11","EMP-011","HODL-9","LM-WH-3-A","Goods SLA Delivery",72,58,2,3,"Days","Holiday"),
        P("p20","2081-82","Shrawan","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",410,398,100,97,"%",""),
        P("p21","2081-82","Shrawan","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",410,390,3,2.6,"Days",""),
        P("p22","2081-82","Shrawan","e3","EMP-003","HODL-1","LM-VEH-2-S","CC Monitoring",260,248,3,2.9,"Days",""),
        P("p23","2081-82","Shrawan","e4","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",380,370,15,13,"Days","Improved"),
        P("p24","2081-82","Shrawan","e4","EMP-004","HODL-4","LM-VEH-6-B","Ownership Transfer",210,207,3,2.2,"Days","Best ever"),
        P("p25","2081-82","Shrawan","e11","EMP-011","HODL-9","LM-WH-3-A","Goods SLA Delivery",105,103,2,1.8,"Days","Better"),
        P("p26","2081-82","Bhadra","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",450,446,100,99,"%",""),
        P("p27","2081-82","Bhadra","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",450,436,3,2.4,"Days",""),
        P("p28","2081-82","Bhadra","e4","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",400,395,15,12,"Days",""),
        P("p29","2081-82","Ashwin","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",480,470,100,98,"%",""),
        P("p30","2081-82","Ashwin","e2","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",480,458,3,2.7,"Days",""),
        P("p31","2081-82","Ashwin","e4","EMP-004","HODL-4","LM-VEH-6-B","Ownership Transfer",220,218,3,2,"Days",""),
        P("p32","2081-82","Kartik","e2","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",390,374,100,96,"%",""),
        P("p33","2081-82","Kartik","e11","EMP-011","HODL-9","LM-WH-3-A","Goods SLA Delivery",98,94,2,2.1,"Days","Slight over"),
    ]
    # Auto-populate loc from emp_locations before insert
    enriched = []
    for row in perf:
        row = list(row)
        if len(row) > 19 and not row[19]:
            emp_code = row[5] if len(row) > 5 else ''
            loc_name = db.execute(
                """SELECT l.name FROM employees e
                   JOIN emp_locations el ON el.emp_id=e.id
                   JOIN locations l ON l.id=el.loc_id
                   WHERE e.emp_code=? ORDER BY el.is_primary DESC LIMIT 1""",
                (emp_code,)
            ).fetchone()
            row[19] = loc_name[0] if loc_name else ''
        enriched.append(tuple(row))
    perf = enriched
    log_audit("PERF_IMPORT","perf","",str(len(perf))+" records imported")
    db.executemany("INSERT OR IGNORE INTO perf VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", perf)
    for fy in ["2080-81","2081-82"]: _upd_cache(db, fy)

# ── EMPLOYEES ──────────────────────────────────────────────────────────────
def enrich_emp(e, db):
    r = dict(e)
    r['role_ids']   = [x['role_id'] for x in db.execute("SELECT role_id FROM emp_roles WHERE emp_id=?", (r['id'],))]
    r['mp_ids']     = [x['mp_id']   for x in db.execute("SELECT mp_id FROM emp_mps WHERE emp_id=?",   (r['id'],))]
    r['cp_ids']     = [x['cp_id']   for x in db.execute("SELECT cp_id FROM emp_cps WHERE emp_id=?",   (r['id'],))]
    # sector_ids — try emp_sectors junction table first, fallback to emp_roles/dept
    try:
        r['sector_ids'] = [x[0] for x in db.execute("SELECT sector_id FROM emp_sectors WHERE emp_id=? ORDER BY is_primary DESC", (r['id'],)).fetchall()]
    except Exception:
        r['sector_ids'] = []
    # loc_ids — try emp_locations junction table first, fallback to loc_emps
    try:
        r['loc_ids'] = [x[0] for x in db.execute("SELECT loc_id FROM emp_locations WHERE emp_id=? ORDER BY is_primary DESC", (r['id'],)).fetchall()]
    except Exception:
        try:
            r['loc_ids'] = [x[0] for x in db.execute("SELECT loc_id FROM loc_emps WHERE emp_id=?", (r['id'],)).fetchall()]
        except Exception:
            r['loc_ids'] = []
    r.pop('photo', None)  # served via /api/employees/photos to keep list payload small
    return r

@app.route('/api/employees', methods=['GET','POST'])
def employees_api():
    db = get_db()
    if request.method == 'GET':
        return jsonify([enrich_emp(e, db) for e in db.execute("SELECT * FROM employees ORDER BY level,name")])
    d = request.json or {}
    missing = validate_required(d, 'name')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    eid = d.get('id') or uid()
    code = d.get('emp_code','').strip()
    if not code:
        last = db.execute("SELECT emp_code FROM employees WHERE emp_code LIKE 'EMP-%' ORDER BY emp_code DESC LIMIT 1").fetchone()
        try: num = int(last['emp_code'].split('-')[1])+1 if last else 1
        except: num = 1
        code = f"EMP-{num:03d}"
    db.execute("INSERT OR REPLACE INTO employees VALUES(?,?,?,?,?,?,?,?,?)",
               (eid,code,d['name'],d.get('role',''),d.get('level',3),d.get('dept','Ops'),d.get('manager_id') or None,d.get('email',''),d.get('photo','')))
    # Save sector assignments
    try:
        db.execute("DELETE FROM emp_sectors WHERE emp_id=?", (eid,))
        for i, sid in enumerate(d.get('sector_ids', [])):
            db.execute("INSERT OR IGNORE INTO emp_sectors(emp_id,sector_id,is_primary) VALUES(?,?,?)", (eid, sid, 1 if i==0 else 0))
    except Exception: pass
    # Save location assignments
    try:
        db.execute("DELETE FROM emp_locations WHERE emp_id=?", (eid,))
        for i, lid2 in enumerate(d.get('loc_ids', [])):
            db.execute("INSERT OR IGNORE INTO emp_locations(emp_id,loc_id,is_primary) VALUES(?,?,?)", (eid, lid2, 1 if i==0 else 0))
    except Exception: pass
    db.commit()
    return jsonify({'id': eid, 'emp_code': code})

@app.route('/api/employees/photos', methods=['GET'])
def employees_photos():
    db = get_db()
    rows = db.execute("SELECT id, photo FROM employees WHERE photo != '' AND photo IS NOT NULL").fetchall()
    return jsonify({r['id']: r['photo'] for r in rows})

@app.route('/api/employees/<eid>/photo', methods=['POST','DELETE'])
def employee_photo(eid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("UPDATE employees SET photo='' WHERE id=?", (eid,))
        db.commit()
        return jsonify({'ok': True})
    photo = (request.json or {}).get('photo','')
    if len(photo) > 700000:
        return jsonify({'error':'Image too large'}), 400
    db.execute("UPDATE employees SET photo=? WHERE id=?", (photo, eid))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/employees/<eid>', methods=['PUT','DELETE'])
def employee_api(eid):
    db = get_db()
    if request.method == 'DELETE':
        _emp_name = db.execute('SELECT name,emp_code FROM employees WHERE id=?',(eid,)).fetchone()
        _emp_label = (_emp_name['name']+' ('+_emp_name['emp_code']+')') if _emp_name else eid
        for t,c in [('employees','id'),('mp_owners','emp_id'),('cp_owners','emp_id'),
                    ('emp_roles','emp_id'),('emp_mps','emp_id'),('emp_cps','emp_id')]:
            db.execute(f"DELETE FROM {t} WHERE {c}=?", (eid,))
        log_audit('EMP_DELETE','employee',eid,'Deleted: '+_emp_label)
        db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    # Read old record for audit diff
    _old_emp = db.execute('SELECT * FROM employees WHERE id=?',(eid,)).fetchone()
    _old_emp = dict(_old_emp) if _old_emp else {}
    db.execute("UPDATE employees SET emp_code=?,name=?,role=?,level=?,dept=?,manager_id=?,email=?,photo=? WHERE id=?",
               (d.get('emp_code'),d['name'],d.get('role',''),d.get('level',3),d.get('dept','Ops'),d.get('manager_id') or None,d.get('email',''),d.get('photo',''),eid))
    # Update sector assignments
    db.execute("DELETE FROM emp_sectors WHERE emp_id=?", (eid,))
    for i, sid in enumerate(d.get('sector_ids', [])):
        db.execute("INSERT OR IGNORE INTO emp_sectors(emp_id,sector_id,is_primary) VALUES(?,?,?)", (eid, sid, 1 if i==0 else 0))
    # Update location assignments
    db.execute("DELETE FROM emp_locations WHERE emp_id=?", (eid,))
    for i, lid2 in enumerate(d.get('loc_ids', [])):
        db.execute("INSERT OR IGNORE INTO emp_locations(emp_id,loc_id,is_primary) VALUES(?,?,?)", (eid, lid2, 1 if i==0 else 0))
    db.commit()
    # Build diff for audit
    _diffs=[]
    for _k,_lbl in [('name','Name'),('role','Role'),('level','Level'),('dept','Dept'),('emp_code','Emp Code'),('email','Email')]:
        _ov=str(_old_emp.get(_k,'')); _nv=str(d.get(_k,''))
        if _ov!=_nv: _diffs.append(_lbl+': '+_ov+' → '+_nv)
    _detail='; '.join(_diffs) if _diffs else 'No field changes'
    log_audit('EMP_UPDATE','employee',eid,_detail)
    return jsonify({'ok': True})

@app.route('/api/emp_links/<eid>', methods=['GET','POST'])
def emp_links(eid):
    db = get_db()
    if request.method == 'GET':
        return jsonify({
            'role_ids': [r['role_id'] for r in db.execute("SELECT role_id FROM emp_roles WHERE emp_id=?", (eid,))],
            'mp_ids':   [r['mp_id']   for r in db.execute("SELECT mp_id FROM emp_mps WHERE emp_id=?",   (eid,))],
            'cp_ids':   [r['cp_id']   for r in db.execute("SELECT cp_id FROM emp_cps WHERE emp_id=?",   (eid,))]})
    d = request.json or {}
    for t in ['emp_roles','emp_mps','emp_cps']:
        db.execute(f"DELETE FROM {t} WHERE emp_id=?", (eid,))
    for rid in d.get('role_ids',[]):
        db.execute("INSERT OR IGNORE INTO emp_roles VALUES(?,?)", (eid, rid))
        for mid in [r['mp_id'] for r in db.execute("SELECT mp_id FROM role_mps WHERE role_id=?", (rid,))]:
            db.execute("INSERT OR IGNORE INTO emp_mps VALUES(?,?)", (eid, mid))
            db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (mid, eid))
        for cid in [r['cp_id'] for r in db.execute("SELECT cp_id FROM role_cps WHERE role_id=?", (rid,))]:
            db.execute("INSERT OR IGNORE INTO emp_cps VALUES(?,?)", (eid, cid))
            db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, eid))
    for mid in d.get('mp_ids',[]): db.execute("INSERT OR IGNORE INTO emp_mps VALUES(?,?)", (eid, mid)); db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (mid, eid))
    for cid in d.get('cp_ids',[]): db.execute("INSERT OR IGNORE INTO emp_cps VALUES(?,?)", (eid, cid)); db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, eid))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/employees/template')
def employees_template():
    if not HAS_OPENPYXL: return json_error('openpyxl not installed')
    import io
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Employees'
    headers = ['emp_code','name','role','level','department','manager_code','email']
    ws.append(headers)
    ws.append(['EMP000001','John Doe','Manager - Logistics',2,'CVBU','','john@example.com'])
    ws.append(['EMP000002','Jane Smith','Officer - Logistics',3,'CVBU','EMP000001','jane@example.com'])
    # Bold header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='ED1C24')
    # Column widths
    for col, w in zip('ABCDEFG', [15,25,35,8,15,15,30]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition':'attachment;filename=employee_import_template.xlsx'})

@app.route('/api/employees/import', methods=['POST'])
def import_employees():
    f = request.files.get('file')
    if not f: return json_error('No file')
    db = get_db()
    ext = f.filename.lower().split('.')[-1]
    imported = 0; errors = []; rows = []
    if ext in ('xlsx','xls'):
        if not HAS_OPENPYXL: return json_error('pip install openpyxl')
        wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
        hdrs = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        def col(n,al=[]): return next((hdrs.index(a) for a in [n]+al if a in hdrs), None)
        ci=col('emp_code',['code']); cn=col('name',['full name'])
        cr=col('role',['designation']); cl=col('level')
        cd=col('department',['dept']); cm=col('manager_code',['manager code'])
        ce=col('email',['email address'])
        for i,row in enumerate(ws.iter_rows(min_row=3,values_only=True),3):
            try:
                name = str(row[cn] or '').strip() if cn is not None else ''
                if not name or name.startswith('←'): continue
                rows.append({'emp_code':str(row[ci] or '').strip() if ci is not None else '',
                             'name':name,'role':str(row[cr] or '') if cr is not None else '',
                             'level':int(row[cl] or 3) if cl is not None else 3,
                             'dept':str(row[cd] or 'Ops') if cd is not None else 'Ops',
                             'manager_code':str(row[cm] or '').strip() if cm is not None else '',
                             'email':str(row[ce] or '') if ce is not None else ''})
            except Exception as e: errors.append(f"Row {i}: {e}")
    else:
        text = f.read().decode('utf-8-sig'); reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            name = (row.get('Name','') or row.get('name','')).strip()
            if not name: continue
            rows.append({'emp_code':(row.get('Emp_Code','') or row.get('emp_code','')).strip(),
                         'name':name,'role':row.get('Role',''),'level':int(row.get('Level',3) or 3),
                         'dept':row.get('Department','Ops') or 'Ops',
                         'manager_code':(row.get('Manager_Code','') or row.get('manager_code','')).strip(),
                         'email':row.get('Email','') or ''})
    id_map = {}
    for r in rows:
        code = r['emp_code']
        if not code:
            last = db.execute("SELECT emp_code FROM employees WHERE emp_code LIKE 'EMP-%' ORDER BY emp_code DESC LIMIT 1").fetchone()
            try: num = int(last['emp_code'].split('-')[1])+1 if last else 1
            except: num = 1
            code = f"EMP-{num:03d}"
        eid = uid()
        db.execute("INSERT OR REPLACE INTO employees VALUES(?,?,?,?,?,?,?,?,?)",
                   (eid,code,r['name'],r['role'],r['level'],r['dept'],None,r['email'],''))
        id_map[code] = eid; imported += 1
    for r in rows:
        mc = r['manager_code']
        if mc and mc in id_map:
            code = r['emp_code'] or list(id_map.keys())[-1]
            eid  = id_map.get(code); mid = id_map.get(mc)
            if eid and mid: db.execute("UPDATE employees SET manager_id=? WHERE id=?", (mid, eid))
    db.commit()
    return jsonify({'imported': imported, 'errors': errors[:10]})

# ── MPs ────────────────────────────────────────────────────────────────────
def enrich_mp(m, db):
    r = dict(m)
    r['owner_ids'] = [x['emp_id'] for x in db.execute("SELECT emp_id FROM mp_owners WHERE mp_id=?", (r['id'],))]
    r['pct'] = round(r['kpi_c']/r['kpi_total']*100,1) if r['kpi_total'] else None
    return r

@app.route('/api/mps', methods=['GET','POST'])
def mps_api():
    db = get_db()
    if request.method == 'GET':
        return jsonify([enrich_mp(m, db) for m in db.execute("SELECT * FROM mps ORDER BY ref")])
    d = request.json or {}
    missing = validate_required(d, 'ref', 'title')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    mid = d.get('id') or uid()
    log_audit("MP_SAVE","mp",mid,"Saved MP "+str(mid))
    db.execute("INSERT OR REPLACE INTO mps(id,ref,title,target,freq,kpi_c,kpi_nc,kpi_total) VALUES(?,?,?,?,?,?,?,?)",
               (mid,d['ref'],d['title'],d.get('target',''),d.get('freq','Monthly'),d.get('kpi_c',0),d.get('kpi_nc',0),d.get('kpi_total',0)))
    db.execute("DELETE FROM mp_owners WHERE mp_id=?", (mid,))
    for eid in d.get('owner_ids',[]): db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (mid, eid))
    db.commit()
    return jsonify({'id': mid})

@app.route('/api/mps/<mid>', methods=['PUT','DELETE'])
def mp_api(mid):
    db = get_db()
    if request.method == 'DELETE':
        _mp = db.execute('SELECT ref,title FROM mps WHERE id=?',(mid,)).fetchone()
        _mp_label = (_mp['ref']+' — '+_mp['title']) if _mp else mid
        log_audit('MP_DELETE','mp',mid,'Deleted MP: '+_mp_label)
        db.execute("DELETE FROM mps WHERE id=?", (mid,))
        db.execute("DELETE FROM mp_owners WHERE mp_id=?", (mid,))
        db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    _old_mp = db.execute('SELECT * FROM mps WHERE id=?',(mid,)).fetchone()
    _old_mp = dict(_old_mp) if _old_mp else {}
    _mp_diffs=[]
    for _k,_lbl in [('ref','Ref'),('title','Title'),('target','Target'),('freq','Frequency')]:
        _ov=str(_old_mp.get(_k,'')); _nv=str(d.get(_k,''))
        if _ov!=_nv: _mp_diffs.append(_lbl+': '+_ov+' → '+_nv)
    log_audit('MP_UPDATE','mp',mid,'; '.join(_mp_diffs) if _mp_diffs else 'MP updated')
    db.execute("UPDATE mps SET ref=?,title=?,target=?,freq=?,kpi_c=?,kpi_nc=?,kpi_total=? WHERE id=?",
               (d['ref'],d['title'],d.get('target',''),d.get('freq','Monthly'),d.get('kpi_c',0),d.get('kpi_nc',0),d.get('kpi_total',0),mid))
    db.execute("DELETE FROM mp_owners WHERE mp_id=?", (mid,))
    for eid in d.get('owner_ids',[]): db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (mid, eid))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/mps/import_excel', methods=['POST'])
def import_mps_excel():
    if not HAS_OPENPYXL: return json_error('pip install openpyxl')
    f = request.files.get('file')
    if not f: return json_error('No file')
    db = get_db(); wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
    hdrs = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
    def col(n,al=[]): return next((hdrs.index(a) for a in [n]+al if a in hdrs), None)
    ci=col('ref',['mp ref','mp_ref']); ct=col('title',['managing point'])
    cta=col('target',['sla']); cf=col('frequency',['freq']); ctot=col('kpi_total',['total'])
    cc=col('kpi_c',['compliant','c']); cnc=col('kpi_nc',['non_compliant','nc'])
    co=col('owner_codes',['owners','employee codes'])
    code_map = {r['emp_code']:r['id'] for r in db.execute("SELECT id,emp_code FROM employees WHERE emp_code IS NOT NULL")}
    imp = 0; errs = []
    for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
        try:
            ref = str(row[ci] or '').strip() if ci is not None else ''
            if not ref: continue
            title=str(row[ct] or '') if ct is not None else ''
            tgt=str(row[cta] or '') if cta is not None else ''
            freq=str(row[cf] or 'Monthly') if cf is not None else 'Monthly'
            tot=int(row[ctot] or 0) if ctot is not None else 0
            cv=int(row[cc] or 0) if cc is not None else 0
            nv=int(row[cnc] or 0) if cnc is not None else 0
            ex = db.execute("SELECT id FROM mps WHERE ref=?", (ref,)).fetchone()
            mid = ex['id'] if ex else uid()
            db.execute("INSERT OR REPLACE INTO mps(id,ref,title,target,freq,kpi_c,kpi_nc,kpi_total) VALUES(?,?,?,?,?,?,?,?)", (mid,ref,title,tgt,freq,cv,nv,tot))
            if co is not None and row[co]:
                codes = [x.strip() for x in str(row[co]).split(',') if x.strip()]
                db.execute("DELETE FROM mp_owners WHERE mp_id=?", (mid,))
                for code in codes:
                    eid = code_map.get(code)
                    if eid: db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (mid, eid))
            imp += 1
        except Exception as e: errs.append(f"Row {i}: {e}")
    db.commit()
    return jsonify({'imported': imp, 'errors': errs})

@app.route('/api/mps/template_excel')
def mp_excel_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Managing Points"
    ws.append(["Ref","Title","Target","Frequency","KPI_Total","KPI_C","KPI_NC","Owner_Codes"])
    ws.append(["HODL-1","Fully compliant and timely vehicle import","1 Day","Monthly",3299,3183,116,"EMP-002,EMP-003"])
    ws.append(["HODL-2","100% NIAC Claim resolution","120 Days","Fortnightly",0,0,0,"EMP-002"])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='MP_Template.xlsx')

# ── CPs ────────────────────────────────────────────────────────────────────
def enrich_cp(c, db):
    r = dict(c)
    r['owner_ids'] = [x['emp_id'] for x in db.execute("SELECT emp_id FROM cp_owners WHERE cp_id=?", (r['id'],))]
    return r

@app.route('/api/cps', methods=['GET','POST'])
def cps_api():
    db = get_db()
    if request.method == 'GET':
        return jsonify([enrich_cp(c, db) for c in db.execute("SELECT * FROM cps ORDER BY ref")])
    d = request.json or {}
    missing = validate_required(d, 'ref', 'title')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    cid = d.get('id') or uid()
    log_audit("CP_SAVE","cp",cid,"Saved CP "+str(cid))
    db.execute("INSERT OR REPLACE INTO cps(id,ref,title,target,freq,source,mp_id) VALUES(?,?,?,?,?,?,?)",
               (cid,d['ref'],d['title'],d.get('target',''),d.get('freq','Daily'),d.get('source',''),d.get('mp_id','')))
    db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
    for eid in d.get('owner_ids',[]): db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, eid))
    db.commit()
    return jsonify({'id': cid})

@app.route('/api/cps/<cid>', methods=['PUT','DELETE'])
def cp_api(cid):
    db = get_db()
    if request.method == 'DELETE':
        log_audit("CP_DELETE","cp",cid,"Deleted CP "+str(cid))
        db.execute("DELETE FROM cps WHERE id=?", (cid,))
        db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
        db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE cps SET ref=?,title=?,target=?,freq=?,source=?,mp_id=? WHERE id=?",
               (d['ref'],d['title'],d.get('target',''),d.get('freq','Daily'),d.get('source',''),d.get('mp_id',''),cid))
    db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
    for eid in d.get('owner_ids',[]): db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, eid))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/cps/import_excel', methods=['POST'])
def import_cps_excel():
    if not HAS_OPENPYXL: return json_error('pip install openpyxl')
    f = request.files.get('file')
    if not f: return json_error('No file')
    db = get_db(); wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
    hdrs = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
    def col(n,al=[]): return next((hdrs.index(a) for a in [n]+al if a in hdrs), None)
    ci=col('ref',['cp ref']); ct=col('title',['checking point'])
    cta=col('target',['sla']); cf=col('frequency',['freq']); cs=col('source',['report source'])
    cmpr=col('mp_ref',['mp ref']); co=col('owner_codes',['owners'])
    code_map = {r['emp_code']:r['id'] for r in db.execute("SELECT id,emp_code FROM employees WHERE emp_code IS NOT NULL")}
    mp_map   = {r['ref']:r['id'] for r in db.execute("SELECT id,ref FROM mps")}
    imp = 0; errs = []
    for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
        try:
            ref = str(row[ci] or '').strip() if ci is not None else ''
            if not ref: continue
            title=str(row[ct] or '') if ct is not None else ''
            tgt=str(row[cta] or '') if cta is not None else ''
            freq=str(row[cf] or 'Daily') if cf is not None else 'Daily'
            src=str(row[cs] or '') if cs is not None else ''
            mpref=str(row[cmpr] or '') if cmpr is not None else ''
            mp_id=mp_map.get(mpref,'')
            ex = db.execute("SELECT id FROM cps WHERE ref=?", (ref,)).fetchone()
            cid = ex['id'] if ex else uid()
            db.execute("INSERT OR REPLACE INTO cps(id,ref,title,target,freq,source,mp_id) VALUES(?,?,?,?,?,?,?)", (cid,ref,title,tgt,freq,src,mp_id))
            if co is not None and row[co]:
                codes = [x.strip() for x in str(row[co]).split(',') if x.strip()]
                db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
                for code in codes:
                    eid = code_map.get(code)
                    if eid: db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, eid))
            imp += 1
        except Exception as e: errs.append(f"Row {i}: {e}")
    db.commit()
    return jsonify({'imported': imp, 'errors': errs})

@app.route('/api/cps/template_excel')
def cp_excel_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Checking Points"
    ws.append(["Ref","Title","Target","Frequency","Source","MP_Ref","Owner_Codes"])
    ws.append(["LM-VEH-1","Tracking of New Vehicles till arrival at border","100%","Monthly","Manual","HODL-1","EMP-002"])
    ws.append(["LM-VEH-2-B","On time custom clearance within 3 days of CC memo","3 Days","Daily","System","HODL-1","EMP-002"])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='CP_Template.xlsx')

# ── ROLES ──────────────────────────────────────────────────────────────────
@app.route('/api/roles', methods=['GET','POST'])
def roles_api():
    db = get_db()
    if request.method == 'GET':
        res = []
        for r in db.execute("SELECT * FROM roles ORDER BY code"):
            role = dict(r)
            role['mp_ids'] = [x['mp_id'] for x in db.execute("SELECT mp_id FROM role_mps WHERE role_id=?", (r['id'],))]
            role['cp_ids'] = [x['cp_id'] for x in db.execute("SELECT cp_id FROM role_cps WHERE role_id=?", (r['id'],))]
            res.append(role)
        return jsonify(res)
    d = request.json or {}
    missing = validate_required(d, 'code', 'name')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    rid = d.get('id') or uid()
    db.execute("INSERT OR REPLACE INTO roles(id,code,name,description,color) VALUES(?,?,?,?,?)",
               (rid,d['code'],d['name'],d.get('description',''),d.get('color','#1d4ed8')))
    db.execute("DELETE FROM role_mps WHERE role_id=?", (rid,))
    db.execute("DELETE FROM role_cps WHERE role_id=?", (rid,))
    for mid in d.get('mp_ids',[]): db.execute("INSERT OR IGNORE INTO role_mps VALUES(?,?)", (rid, mid))
    for cid in d.get('cp_ids',[]): db.execute("INSERT OR IGNORE INTO role_cps VALUES(?,?)", (rid, cid))
    db.commit()
    return jsonify({'id': rid})

@app.route('/api/roles/<rid>', methods=['PUT','DELETE'])
def role_api(rid):
    db = get_db()
    if request.method == 'DELETE':
        for t,c in [('roles','id'),('role_mps','role_id'),('role_cps','role_id'),('emp_roles','role_id')]:
            db.execute(f"DELETE FROM {t} WHERE {c}=?", (rid,))
        db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE roles SET code=?,name=?,description=?,color=? WHERE id=?",
               (d['code'],d['name'],d.get('description',''),d.get('color','#1d4ed8'),rid))
    db.execute("DELETE FROM role_mps WHERE role_id=?", (rid,))
    db.execute("DELETE FROM role_cps WHERE role_id=?", (rid,))
    for mid in d.get('mp_ids',[]): db.execute("INSERT OR IGNORE INTO role_mps VALUES(?,?)", (rid, mid))
    for cid in d.get('cp_ids',[]): db.execute("INSERT OR IGNORE INTO role_cps VALUES(?,?)", (rid, cid))
    db.commit()
    return jsonify({'ok': True})

# ── CASCADE LINKS  (NEW) ───────────────────────────────────────────────────
@app.route('/api/cascade_links', methods=['GET','POST'])
def cascade_links_api():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute('''
            SELECT cl.*,
              se.name  as sup_name,  se.emp_code as sup_code,
              sub.name as sub_name, sub.emp_code as sub_code,
              cp.ref   as cp_ref,   cp.title     as cp_title,
              mp.ref   as mp_ref,   mp.title     as mp_title
            FROM cascade_links cl
            LEFT JOIN employees se   ON cl.superior_emp_id    = se.id
            LEFT JOIN employees sub  ON cl.subordinate_emp_id = sub.id
            LEFT JOIN cps       cp   ON cl.superior_cp_id     = cp.id
            LEFT JOIN mps       mp   ON cl.subordinate_mp_id  = mp.id
            ORDER BY se.name
        ''').fetchall()
        return jsonify(R(rows))

    d = request.json or {}
    missing = validate_required(d, 'superior_emp_id', 'superior_cp_id', 'subordinate_emp_id')
    if missing: return json_error(f"Missing: {', '.join(missing)}")

    # Duplicate check
    existing = db.execute(
        'SELECT id FROM cascade_links WHERE superior_cp_id=? AND subordinate_emp_id=?',
        (d['superior_cp_id'], d['subordinate_emp_id'])
    ).fetchone()
    if existing:
        return json_error('Cascade link already exists for this CP → Employee pair', 409)

    sub_mp_id    = d.get('subordinate_mp_id','').strip()
    auto_created = 0

    if not sub_mp_id:
        cp = db.execute('SELECT * FROM cps WHERE id=?', (d['superior_cp_id'],)).fetchone()
        if cp:
            new_ref   = f"AUTO-{cp['ref']}"
            new_title = f"[Auto] {(cp['title'] or cp['ref'] or 'Untitled')}"
            ex_mp = db.execute('SELECT id FROM mps WHERE ref=?', (new_ref,)).fetchone()
            if ex_mp:
                sub_mp_id = ex_mp['id']
            else:
                sub_mp_id = uid()
                db.execute('INSERT INTO mps(id,ref,title,target,freq,kpi_c,kpi_nc,kpi_total) VALUES(?,?,?,?,?,?,?,?)',
                           (sub_mp_id,new_ref,new_title,(cp['target'] or ''),(cp['freq'] or 'Monthly'),0,0,0))
                db.execute('INSERT OR IGNORE INTO mp_owners VALUES(?,?)',
                           (sub_mp_id, d['subordinate_emp_id']))
            auto_created = 1

    link_id = uid()
    db.execute('INSERT INTO cascade_links VALUES(?,?,?,?,?,?,?)',
               (link_id, d['superior_emp_id'], d['superior_cp_id'],
                d['subordinate_emp_id'], sub_mp_id, auto_created,
                datetime.datetime.now().isoformat()))
    db.commit()
    return jsonify({'id': link_id, 'subordinate_mp_id': sub_mp_id, 'auto_created': bool(auto_created)})


@app.route('/api/cascade_links/<lid>', methods=['PUT','DELETE'])
def cascade_link_detail(lid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute('DELETE FROM cascade_links WHERE id=?', (lid,))
        db.commit(); return jsonify({'ok': True})
    d = request.json or {}
    missing = validate_required(d, 'superior_emp_id', 'superior_cp_id', 'subordinate_emp_id')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    sub_mp_id = d.get('subordinate_mp_id','').strip()
    if not sub_mp_id:
        cp = db.execute('SELECT * FROM cps WHERE id=?', (d['superior_cp_id'],)).fetchone()
        if cp:
            new_ref = 'AUTO-'+cp['ref']
            ex_mp = db.execute('SELECT id FROM mps WHERE ref=?', (new_ref,)).fetchone()
            if ex_mp: sub_mp_id = ex_mp['id']
    db.execute(
        'UPDATE cascade_links SET superior_emp_id=?,superior_cp_id=?,subordinate_emp_id=?,subordinate_mp_id=? WHERE id=?',
        (d['superior_emp_id'],d['superior_cp_id'],d['subordinate_emp_id'],sub_mp_id,lid))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/fy_from_date')
def fy_from_date():
    date_str = request.args.get('date','')
    try:
        dt  = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        fy, bs_month = ad_date_to_fy_and_month(dt)
        return jsonify({'fy': fy, 'bs_month': bs_month, 'quarter': bs_q(bs_month)})
    except Exception as e:
        return json_error(f'Invalid date: {e}')

# ── CACHE ──────────────────────────────────────────────────────────────────
@app.route('/api/cache', methods=['GET','POST'])
def cache_api():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM perf_cache ORDER BY fy DESC").fetchall()
        res  = []
        for r in rows:
            item = dict(r)
            item['record_count'] = db.execute("SELECT COUNT(*) FROM perf WHERE fy=?", (r['fy'],)).fetchone()[0]
            res.append(item)
        return jsonify(res)
    d  = request.json or {}
    fy = d.get('fy','').strip()
    if not fy: return json_error('FY required')
    now = datetime.datetime.now().isoformat()
    db.execute("INSERT OR IGNORE INTO perf_cache VALUES(?,?,?,?,?,?)",
               (fy, d.get('label', f"FY {fy}"), 0, now, now, 0))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/cache/<fy>/clear', methods=['POST'])
def clear_cache(fy):
    db  = get_db()
    row = db.execute("SELECT locked FROM perf_cache WHERE fy=?", (fy,)).fetchone()
    if not row: return json_error('FY not found', 404)
    if row['locked']: return json_error('FY is locked', 400)
    db.execute("DELETE FROM perf WHERE fy=?", (fy,))
    _upd_cache(db, fy)
    db.commit()
    return jsonify({'cleared': True, 'fy': fy})

@app.route('/api/cache/<fy>/lock', methods=['POST'])
def toggle_lock(fy):
    db  = get_db()
    row = db.execute("SELECT locked FROM perf_cache WHERE fy=?", (fy,)).fetchone()
    if not row: return json_error('Not found', 404)
    nl  = 0 if row['locked'] else 1
    db.execute("UPDATE perf_cache SET locked=? WHERE fy=?", (nl, fy))
    db.commit()
    return jsonify({'locked': bool(nl)})

@app.route('/api/cache/<fy>', methods=['DELETE'])
def delete_cache(fy):
    db  = get_db()
    row = db.execute("SELECT locked FROM perf_cache WHERE fy=?", (fy,)).fetchone()
    if row and row['locked']: return json_error('Locked', 400)
    db.execute("DELETE FROM perf WHERE fy=?", (fy,))
    db.execute("DELETE FROM perf_cache WHERE fy=?", (fy,))
    db.commit()
    return jsonify({'deleted': True})

# ── PERF ───────────────────────────────────────────────────────────────────
@app.route('/api/perf', methods=['GET','POST'])
def perf_api():
    db = get_db()
    if request.method == 'GET':
        q    = "SELECT * FROM perf WHERE 1=1"; args = []
        for p,c in [('fy','fy'),('emp_code','emp_code'),('emp_id','emp_id'),
                    ('mp_ref','mp_ref'),('bs_month','bs_month'),('quarter','quarter')]:
            if request.args.get(p): q += f" AND {c}=?"; args.append(request.args[p])
        return jsonify(R(db.execute(q+' ORDER BY fy DESC,bs_month', args).fetchall()))

    d   = request.json or {}
    pid = d.get('id') or uid()
    eid = d.get('emp_id',''); ec = d.get('emp_code','')
    if not eid and ec:
        row = db.execute("SELECT id FROM employees WHERE emp_code=?", (ec,)).fetchone()
        if row: eid = row['id']
    bsm = norm_month(d.get('bs_month','Shrawan'))
    fy  = d.get('fy','2081-82')
    tot = int(d.get('total',0)); comp = int(d.get('compliant',0))
    nc  = int(d.get('non_compliant', tot-comp))
    pct_c  = round(comp/tot*100,2) if tot else 0
    pct_nc = round(nc/tot*100,2)   if tot else 0
    st  = d.get('status') or calc_status(d.get('actual_val',0), d.get('target_val',0), d.get('unit','%'))
    db.execute("""INSERT OR REPLACE INTO perf
               (id,fy,bs_month,quarter,emp_id,emp_code,mp_ref,cp_ref,
                metric,total,compliant,non_compliant,pct_compliant,pct_nc,
                target_val,actual_val,unit,status,notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
               (pid,fy,bsm,bs_q(bsm),eid,ec,d.get('mp_ref',''),d.get('cp_ref',''),d.get('metric',''),
                tot,comp,nc,pct_c,pct_nc,d.get('target_val',0),d.get('actual_val',0),d.get('unit','%'),st,d.get('notes','')))
    _upd_cache(db, fy)
    db.commit()
    return jsonify({'id': pid})

# ── QUICK PERF ENTRY  (NEW) ────────────────────────────────────────────────
@app.route('/api/perf/quick', methods=['POST'])
def perf_quick():
    d       = request.json or {}
    missing = validate_required(d, 'date', 'emp_code', 'cp_ref')
    if missing: return json_error(f"Required: {', '.join(missing)}")

    db  = get_db()
    ec  = str(d['emp_code']).strip()
    emp = db.execute('SELECT * FROM employees WHERE emp_code=?', (ec,)).fetchone()
    if not emp: return json_error(f'Employee not found: {ec}')
    eid = emp['id']

    cp = db.execute('SELECT * FROM cps WHERE ref=?', (d['cp_ref'],)).fetchone()
    if not cp: return json_error(f'CP not found: {d["cp_ref"]}')

    mp_ref = ''
    if cp['mp_id']:
        mp = db.execute('SELECT ref FROM mps WHERE id=?', (cp['mp_id'],)).fetchone()
        if mp: mp_ref = mp['ref']

    try:
        dt  = datetime.datetime.strptime(d['date'], '%Y-%m-%d')
        fy, bsm = ad_date_to_fy_and_month(dt)
    except:
        return json_error('Invalid date. Use YYYY-MM-DD')

    # Check FY lock
    lock = db.execute('SELECT locked FROM perf_cache WHERE fy=?', (fy,)).fetchone()
    if lock and lock['locked']:
        return json_error(f'FY {fy} is locked. Unlock first.', 403)

    target_val = float(d.get('target_val', 0) or 0)
    unit       = str(d.get('unit','%') or '%')[:20]
    total = int(d.get('total', 0) or 0)
    if total <= 0 and not d.get('status_override'): return json_error('Total must be > 0 (or set Status Override)')

    status_override = d.get('status_override','').strip().upper()

    if status_override == 'NA':
        compliant = 0; nc = 0; pct_c = 0; pct_nc = 0
    elif status_override in ('C','NC'):
        compliant = total if status_override == 'C' else 0
        nc = 0 if status_override == 'C' else total
        pct_c = 100.0 if status_override == 'C' else 0.0
        pct_nc = 0.0 if status_override == 'C' else 100.0
    else:
        mode = d.get('mode','count')
        if mode == 'percent':
            pct_c     = float(d.get('pct_achieved', 0) or 0)
            if not (0 <= pct_c <= 100): return json_error('Percentage must be 0–100')
            compliant = round(total * pct_c / 100)
        else:
            compliant = int(d.get('compliant', 0) or 0)
            if compliant > total: return json_error('Compliant count cannot exceed total')
            pct_c     = round(compliant / total * 100, 2)
        nc     = total - compliant
        pct_nc = round(nc / total * 100, 2)

    # Parse target from CP
    raw_tgt = (cp['target'] or '').strip()
    tgt_num = ''.join(c for c in raw_tgt if c.isdigit() or c == '.')
    try: tgt_val = float(tgt_num)
    except: tgt_val = 0.0

    actual_val = float(d.get('actual_val', 0) or 0)
    tgt_parts = raw_tgt.split()
    unit       = d.get('unit', tgt_parts[-1] if tgt_parts else '%')
    status     = calc_status(actual_val, tgt_val, unit)
    metric     = d.get('metric', (cp['title'] or '')[:60])

    # DUPLICATE CHECK
    existing = db.execute(
        'SELECT * FROM perf WHERE fy=? AND bs_month=? AND emp_id=? AND cp_ref=?',
        (fy, bsm, eid, d['cp_ref'])
    ).fetchone()

    if existing:
        total     = (existing['total'] or 0) + total
        compliant = (existing['compliant'] or 0) + compliant
        nc        = total - compliant
        pct_c     = round(compliant / total * 100, 2) if total else 0
        pct_nc    = round(nc / total * 100, 2) if total else 0
        actual_val = max(actual_val, existing['actual_val'] or 0)
        status    = 'C' if pct_c >= 95 else 'NC'

    pid = existing['id'] if existing else uid()
    entry_date = d.get('date', '')
    db.execute("""INSERT OR REPLACE INTO perf
               (id,fy,bs_month,quarter,emp_id,emp_code,mp_ref,cp_ref,
                metric,total,compliant,non_compliant,pct_compliant,pct_nc,
                target_val,actual_val,unit,status,notes,entry_date)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
               (pid, fy, bsm, bs_q(bsm), eid, ec, mp_ref, d['cp_ref'],
                metric, total, compliant, nc, pct_c, pct_nc,
                tgt_val, actual_val, unit, status, d.get('notes',''), entry_date))
    _upd_cache(db, fy)
    db.commit()
    return jsonify({
        'id': pid, 'fy': fy, 'bs_month': bsm, 'quarter': bs_q(bsm),
        'emp_code': ec, 'emp_id': eid, 'mp_ref': mp_ref, 'cp_ref': d['cp_ref'],
        'metric': metric, 'total': total, 'compliant': compliant,
        'non_compliant': nc, 'pct_compliant': pct_c, 'pct_nc': pct_nc,
        'status': status, 'unit': unit, 'notes': d.get('notes',''),
        'loc': '',
        'accumulated': existing is not None and not d.get('overwrite')
    })


@app.route('/api/cascade_mpcp', methods=['GET'])
def cascade_mpcp():
    """Return all AUTO- generated MPs and their CPs for the cascade MP/CP tab."""
    db = get_db()
    mps = R(db.execute(
        "SELECT * FROM mps WHERE ref LIKE 'AUTO-%' ORDER BY ref"
    ).fetchall())
    result = []
    for mp in mps:
        cps = R(db.execute(
            "SELECT * FROM cps WHERE mp_id=? ORDER BY ref", (mp['id'],)
        ).fetchall())
        # find which employee owns this mp
        owners = R(db.execute(
            "SELECT e.id,e.name,e.emp_code,e.level,e.dept FROM employees e "
            "JOIN mp_owners o ON o.emp_id=e.id WHERE o.mp_id=?", (mp['id'],)
        ).fetchall())
        # find the cascade link that created it
        link = db.execute(
            "SELECT cl.*,e2.name as superior_name,e2.emp_code as superior_code "
            "FROM cascade_links cl "
            "JOIN employees e2 ON e2.id=cl.superior_emp_id "
            "WHERE cl.subordinate_mp_id=?", (mp['id'],)
        ).fetchone()
        result.append({
            **mp,
            'cps': cps,
            'owners': owners,
            'source_link': dict(link) if link else None
        })
    return jsonify(result)

@app.route('/api/perf/exceptions')
def perf_exceptions():
    db  = get_db()
    fy  = request.args.get('fy', '')
    q2  = "SELECT p.*, e.name as emp_name FROM perf p LEFT JOIN employees e ON p.emp_id=e.id WHERE p.status='NC'"
    args= []
    if fy: q2 += " AND p.fy=?"; args.append(fy)
    rows = R(db.execute(q2 + " ORDER BY p.fy DESC,p.bs_month", args).fetchall())
    repeat = {}
    for r in rows:
        cp = r.get('cp_ref','')
        if cp: repeat[cp] = repeat.get(cp,0)+1
    return jsonify({'total_nc': len(rows), 'exceptions': rows[:50],
                    'repeat_failures': {k:v for k,v in repeat.items() if v>1}})

# ── MPCP CREATOR (outline-builder publish) ──────────────────────────────────
def _ensure_mpcp_status_columns(db):
    """
    init_db()/_migrate() exist elsewhere in this file but are never actually
    invoked (no call site found — confirmed by audit). CREATE TABLE IF NOT
    EXISTS in SCHEMA only affects brand-new DB files, so existing master.db
    and per-department .db files never get new columns automatically.
    This runs the same additive ALTER TABLE pattern directly against
    whichever DB this request's get_db() resolves to, so it self-heals
    both master and department databases the first time Publish touches them.
    """
    for table in ('mps', 'cps', 'roles'):
        cols = {row[1] for row in db.execute(f"PRAGMA table_info({table})").fetchall()}
        if 'status' not in cols:
            db.execute(f"ALTER TABLE {table} ADD COLUMN status TEXT DEFAULT 'active'")
    db.commit()

@app.route('/api/mpcp/publish', methods=['POST'])
def mpcp_publish():
    """
    Accepts a flattened outline (roles -> mps -> cps) built in the MPCP Creator
    canvas and writes it into the SAME mps/cps/roles/cascade_links tables used
    everywhere else in the app. Reuses the existing ref-keyed upsert pattern
    (INSERT OR REPLACE keyed by ref) already used by /api/mps/import_excel and
    /api/cps/import_excel, so re-publishing the same outline is idempotent.
    Nothing here is a new data model — it's the same rows, created a different way.
    """
    db = get_db()
    _ensure_mpcp_status_columns(db)
    d = request.json or {}
    nodes = d.get('nodes', [])
    if not nodes:
        return json_error('No nodes to publish')

    created = {'roles': 0, 'mps': 0, 'cps': 0, 'cascade_links': 0}
    errors = []
    mp_ref_to_id = {}

    # Pass 1: roles + MPs (CPs may reference an MP created in this same batch)
    for n in nodes:
        try:
            if n.get('kind') == 'role':
                rid = n.get('id') or uid()
                code = (n.get('ref') or n.get('text','ROLE'))[:40]
                db.execute("INSERT OR REPLACE INTO roles VALUES(?,?,?,?,?,?)",
                           (rid, code, n.get('text',''), n.get('description',''), n.get('color','#1d4ed8'),
                            n.get('status','active')))
                created['roles'] += 1
            elif n.get('kind') == 'mp':
                ref = n.get('ref','').strip()
                if not ref: continue
                ex = db.execute("SELECT id FROM mps WHERE ref=?", (ref,)).fetchone()
                mid = ex['id'] if ex else uid()
                db.execute("INSERT OR REPLACE INTO mps VALUES(?,?,?,?,?,?,?,?,?)",
                           (mid, ref, n.get('text',''), n.get('target',''), n.get('freq','Monthly'),
                            n.get('kpi_c',0), n.get('kpi_nc',0), n.get('kpi_total',0), n.get('status','active')))
                mp_ref_to_id[ref] = mid
                if n.get('role_id'):
                    db.execute("INSERT OR IGNORE INTO role_mps VALUES(?,?)", (n['role_id'], mid))
                created['mps'] += 1
        except Exception as e:
            errors.append(f"{n.get('kind')} '{n.get('text','')}': {e}")

    # Pass 2: CPs (now mp_ref_to_id is fully populated for this batch)
    code_map = {r['emp_code']: r['id'] for r in db.execute(
        "SELECT id, emp_code FROM employees WHERE emp_code IS NOT NULL")}
    for n in nodes:
        if n.get('kind') != 'cp': continue
        try:
            ref = n.get('ref','').strip()
            if not ref: continue
            mp_ref = n.get('parent_mp_ref','')
            mp_id = mp_ref_to_id.get(mp_ref) or (
                db.execute("SELECT id FROM mps WHERE ref=?", (mp_ref,)).fetchone() or {}).get('id', '')
            ex = db.execute("SELECT id FROM cps WHERE ref=?", (ref,)).fetchone()
            cid = ex['id'] if ex else uid()
            db.execute("INSERT OR REPLACE INTO cps VALUES(?,?,?,?,?,?,?,?)",
                       (cid, ref, n.get('text',''), n.get('target',''), n.get('freq','Daily'),
                        n.get('source',''), mp_id, n.get('status','active')))
            db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
            owner_code = n.get('owner_code','')
            owner_eid = code_map.get(owner_code)
            if owner_eid:
                db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (cid, owner_eid))
            created['cps'] += 1
        except Exception as e:
            errors.append(f"cp '{n.get('text','')}': {e}")

    db.commit()
    log_audit("MPCP_PUBLISH", "mpcp", "batch",
               f"Published {created['mps']} MP(s), {created['cps']} CP(s) via MPCP Creator")
    return jsonify({'ok': True, 'created': created, 'errors': errors})

@app.route('/api/mpcp/set_status', methods=['POST'])
def mpcp_set_status():
    """Close/reopen a Role, MP, or CP. Status only filters entry/dashboard views —
    never touches perf or cascade_links rows, per the 'always editable' requirement."""
    db = get_db()
    _ensure_mpcp_status_columns(db)
    d = request.json or {}
    kind, rid, status = d.get('kind'), d.get('id'), d.get('status')
    if kind not in ('role', 'mp', 'cp') or not rid or status not in ('active', 'closed'):
        return json_error('kind must be role/mp/cp, status must be active/closed')
    table = {'role': 'roles', 'mp': 'mps', 'cp': 'cps'}[kind]
    db.execute(f"UPDATE {table} SET status=? WHERE id=?", (status, rid))
    db.commit()
    log_audit("MPCP_STATUS", kind, rid, f"Set status={status}")
    return jsonify({'ok': True})

# ── MPCP CREATOR: LOOKUP + EXPORT ─────────────────────────────────────────
@app.route('/api/mpcp/lookup')
def mpcp_lookup():
    """Return existing roles, mps, cps, and employees for Creator dropdowns."""
    db = get_db()
    _ensure_mpcp_status_columns(db)
    q = (request.args.get('q') or '').strip().lower()
    roles = [dict(r) for r in db.execute(
        "SELECT id, code, name FROM roles ORDER BY name").fetchall()]
    mps = [dict(r) for r in db.execute(
        "SELECT id, ref, title, target, freq FROM mps ORDER BY ref").fetchall()]
    cps = [dict(r) for r in db.execute(
        "SELECT id, ref, title, target, freq FROM cps ORDER BY ref").fetchall()]
    emps = [dict(r) for r in db.execute(
        "SELECT id, emp_code, name, level FROM employees ORDER BY name").fetchall()]
    if q:
        roles = [r for r in roles if q in (r.get('name') or '').lower() or q in (r.get('code') or '').lower()]
        mps   = [m for m in mps   if q in (m.get('ref') or '').lower() or q in (m.get('title') or '').lower()]
        cps   = [c for c in cps   if q in (c.get('ref') or '').lower() or q in (c.get('title') or '').lower()]
        emps  = [e for e in emps  if q in (e.get('name') or '').lower() or q in (e.get('emp_code') or '').lower()]
    return jsonify({'roles': roles, 'mps': mps, 'cps': cps, 'employees': emps})


def _compiled_mpcp_workbook(nodes):
    """Build an openpyxl workbook in the original Compiled MPCP 15-column format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADERS = ['Role','S.N.','Managing Points','UoM','MP-Ref','Target','Freq',
               'S.N.','Checking Points','UoM','CP-Ref','Target','CP Resp','Freq','Report Source']
    COL_W   = [18, 6, 38, 8, 14, 10, 10, 6, 38, 8, 14, 10, 16, 10, 16]

    hdr_fill = PatternFill('solid', fgColor='1F3864')
    hdr_font = Font(bold=True, color='FFFFFF', size=9)
    role_fill = PatternFill('solid', fgColor='D6E4F7')
    mp_fill   = PatternFill('solid', fgColor='EBF3FB')
    thin = Side(border_style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='top')

    def _write_sheet(ws, sheet_nodes):
        ws.freeze_panes = 'A2'
        for ci, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
            cell = ws.cell(1, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w

        row = 2
        for rn in sheet_nodes:
            if rn.get('kind') != 'role': continue
            role_text = rn.get('text', '')
            for mi, mn in enumerate(rn.get('children', []), 1):
                if mn.get('kind') != 'mp': continue
                cps_list = [c for c in mn.get('children', []) if c.get('kind') == 'cp']
                if not cps_list:
                    cps_list = [{}]  # at least one blank CP row per MP
                mp_row_start = row
                for ci2, cn in enumerate(cps_list, 1):
                    sn_mp = f"{mi}" if ci2 == 1 else ''
                    role_cell_val = role_text if (mi == 1 and ci2 == 1) else ''
                    vals = [
                        role_cell_val,
                        sn_mp,
                        mn.get('text','') if ci2 == 1 else '',
                        mn.get('uom',''),
                        mn.get('ref','') if ci2 == 1 else '',
                        mn.get('target',''),
                        mn.get('freq',''),
                        f"{mi}.{ci2}",
                        cn.get('text',''),
                        cn.get('uom',''),
                        cn.get('ref',''),
                        cn.get('target',''),
                        cn.get('owner_name', cn.get('owner_code','')),
                        cn.get('freq',''),
                        cn.get('source',''),
                    ]
                    for col_i, v in enumerate(vals, 1):
                        cell = ws.cell(row, col_i, v)
                        cell.border = border; cell.alignment = wrap; cell.font = Font(size=9)
                        if col_i == 1: cell.fill = role_fill
                        elif col_i <= 7: cell.fill = mp_fill
                    ws.row_dimensions[row].height = 28
                    row += 1
                # Merge MP columns for rows spanning multiple CPs
                if len(cps_list) > 1:
                    for col_i in [1, 2, 3, 4, 5, 6, 7]:
                        ws.merge_cells(start_row=mp_row_start, start_column=col_i,
                                       end_row=row-1, end_column=col_i)
                        ws.cell(mp_row_start, col_i).alignment = Alignment(vertical='top', wrap_text=True)

    def _collect_owners(nodes):
        owners = {}
        def walk(ns):
            for n in ns:
                if n.get('kind') == 'cp':
                    code = n.get('owner_code','') or n.get('owner_name','')
                    name = n.get('owner_name','') or code
                    if code: owners[code] = name
                walk(n.get('children', []))
        walk(nodes)
        return owners

    wb = openpyxl.Workbook()
    # Master sheet
    ws_master = wb.active; ws_master.title = 'Master'
    _write_sheet(ws_master, nodes)

    # Per-owner sheets
    owners = _collect_owners(nodes)
    for code, name in owners.items():
        ws_name = (name or code)[:31]
        ws_p = wb.create_sheet(ws_name)
        # Filter tree to only CPs owned by this person
        filtered = []
        for rn in nodes:
            r_copy = dict(rn); r_copy['children'] = []
            for mn in rn.get('children', []):
                m_copy = dict(mn); m_copy['children'] = [
                    c for c in mn.get('children', [])
                    if (c.get('owner_code') or c.get('owner_name','')) == code
                ]
                if m_copy['children']:
                    r_copy['children'].append(m_copy)
            if r_copy['children']:
                filtered.append(r_copy)
        _write_sheet(ws_p, filtered)

    return wb


@app.route('/api/mpcp/export_draft', methods=['POST'])
def mpcp_export_draft():
    """Export the Creator draft (client-side tree) directly to Compiled MPCP Excel format."""
    if not HAS_OPENPYXL:
        return json_error('openpyxl not installed')
    d = request.json or {}
    nodes = d.get('nodes', [])
    if not nodes:
        return json_error('No nodes provided')
    wb = _compiled_mpcp_workbook(nodes)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='MPCP_Draft_Export.xlsx')


@app.route('/api/mpcp/export_published')
def mpcp_export_published():
    """Export live published MPs/CPs from DB in Compiled MPCP format."""
    if not HAS_OPENPYXL:
        return json_error('openpyxl not installed')
    db = get_db()
    _ensure_mpcp_status_columns(db)
    # Reconstruct tree from DB
    mps_rows = db.execute("SELECT id, ref, title AS text, target, freq FROM mps ORDER BY ref").fetchall()
    cps_rows  = db.execute("""
        SELECT c.id, c.ref, c.title AS text, c.target, c.freq, c.source, c.mp_id,
               e.emp_code AS owner_code, e.name AS owner_name
        FROM cps c
        LEFT JOIN cp_owners co ON co.cp_id = c.id
        LEFT JOIN employees e  ON e.id = co.owner_id
        ORDER BY c.ref
    """).fetchall()
    mp_map = {r['id']: dict(r, kind='mp', children=[]) for r in mps_rows}
    for r in cps_rows:
        cp = dict(r, kind='cp', children=[])
        if r['mp_id'] and r['mp_id'] in mp_map:
            mp_map[r['mp_id']]['children'].append(cp)
    nodes = [{'kind': 'role', 'text': 'All Published MPs/CPs', 'ref': 'ROOT',
               'children': list(mp_map.values())}]
    wb = _compiled_mpcp_workbook(nodes)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='MPCP_Published_Export.xlsx')


@app.route('/api/mpcp/export_template')
def mpcp_export_template():
    """Return the blank MPCP Creator upload template Excel."""
    if not HAS_OPENPYXL:
        return json_error('openpyxl not installed')
    import openpyxl as _xl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK='0A0E1A'; HDR='1A1F30'; ROLE='1F253A'; MP='141828'
    RED='FF2D3A'; BLUE='2D9EFF'; GREEN='10DF8A'; AMB='FFB020'; MUT='6B7A9F'; TXT='F0F4FF'; TXT2='B8C4E0'; BDR='252B40'
    thin = Side(border_style='thin', color=BDR)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    def sc(ws,r,c,v,bg,fg,bold=False,sz=9,wrap=True,center=False):
        cell=ws.cell(r,c,v)
        cell.font=Font(name='Calibri',bold=bold,color=fg,size=sz)
        cell.fill=PatternFill('solid',fgColor=bg)
        cell.alignment=Alignment(wrap_text=wrap,vertical='center',horizontal='center' if center else 'left')
        cell.border=bdr; return cell

    wb = _xl.Workbook(); ws = wb.active; ws.title='MPCP_Upload'
    ws.sheet_view.showGridLines=False
    cols=[(1,1,'KIND',RED,RED),(2,2,'ROLE / OBJECTIVE',ROLE,AMB),(3,7,'MANAGING POINT ▶',HDR,BLUE),(8,14,'CHECKING POINT ▶',MP,GREEN)]
    for s1,s2,lbl,bg,fg in cols:
        if s1<s2: ws.merge_cells(start_row=1,start_column=s1,end_row=1,end_column=s2)
        c=ws.cell(1,s1,lbl); c.font=Font(name='Calibri',bold=True,color=fg,size=8); c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
        for ci in range(s1+1,s2+1): ws.cell(1,ci).fill=PatternFill('solid',fgColor=bg); ws.cell(1,ci).border=bdr
    hdrs=[('KIND',8,RED),('ROLE',42,AMB),('MP REF',14,BLUE),('MP TEXT',42,BLUE),('MP UOM',9,BLUE),('MP TARGET',12,BLUE),('MP FREQ',12,BLUE),
          ('CP REF',14,GREEN),('CP TEXT',42,GREEN),('CP UOM',9,GREEN),('CP TARGET',12,GREEN),('CP FREQ',12,GREEN),('OWNER NAME',22,AMB),('SOURCE',14,MUT)]
    for i,(h,w,fg) in enumerate(hdrs,1):
        c=ws.cell(2,i,h); c.font=Font(name='Calibri',bold=True,color=fg,size=8)
        c.fill=PatternFill('solid',fgColor=DARK); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=False); c.border=bdr
        ws.column_dimensions[get_column_letter(i)].width=w
    hints=['role/mp/cp or blank','Full objective description','Unique MP code','MP description','Days/%/Nos','e.g. 3 Days','Daily/Monthly',
           'Unique CP code','CP description','Days/%/Nos','e.g. 3 Days','Daily/Weekly','Name or emp code','Manual/System']
    for i,h in enumerate(hints,1):
        c=ws.cell(3,i,h); c.font=Font(name='Calibri',italic=True,color=MUT,size=7)
        c.fill=PatternFill('solid',fgColor=DARK); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
    ws.row_dimensions[1].height=16; ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=14
    sample=[
        ('role','Your Role / Objective Description','','','','','','','','','','','',''),
        ('mp','','HOD-MP-01','Managing Point Description','Days','3 Days','Monthly','','','','','','',''),
        ('cp','','','','','','','HOD-CP-01.01','Checking Point Description','Days','3 Days','Daily','Owner Name','Manual'),
        ('cp','','','','','','','HOD-CP-01.02','Another Checking Point','%','100%','Weekly','Owner Name','System'),
        ('mp','','HOD-MP-02','Second Managing Point','%','100%','Monthly','','','','','','',''),
        ('cp','','','','','','','HOD-CP-02.01','CP under second MP','Nos','1','Daily','Owner Name','System'),
        ('role','Second Role / Objective','','','','','','','','','','','',''),
        ('mp','','HOD-MP-03','Managing Point under second role','Days','15 Days','Fortnightly','','','','','','',''),
        ('cp','','','','','','','HOD-CP-03.01','CP description','Days','15 Days','Weekly','Owner Name','System'),
    ]
    bgs={'role':ROLE,'mp':HDR,'cp':MP}; fgmap={1:(RED,RED,GREEN),2:(AMB,DARK,DARK)}
    for ri,row in enumerate(sample,4):
        kind=row[0]; bg=bgs.get(kind,MP)
        for ci,val in enumerate(row,1):
            if ci==1: fg=RED if kind=='role' else (BLUE if kind=='mp' else GREEN); cbg=bg
            elif ci==2: fg=AMB; cbg=ROLE if kind=='role' else bg
            elif 3<=ci<=7: fg=BLUE if kind=='mp' else MUT; cbg=HDR if kind=='mp' else bg
            else: fg=GREEN if kind=='cp' else MUT; cbg=MP if kind=='cp' else bg
            if ci==13: fg=AMB
            c=ws.cell(ri,ci,val); c.font=Font(name='Calibri',bold=(ci==1),color=fg,size=9)
            c.fill=PatternFill('solid',fgColor=cbg); c.alignment=Alignment(wrap_text=True,vertical='top',horizontal='center' if ci==1 else 'left'); c.border=bdr
        ws.row_dimensions[ri].height=22 if kind=='cp' else 26
    note=ws.max_row+2
    c=ws.cell(note,1,'NOTE:'); c.font=Font(bold=True,color=RED,size=8); c.fill=PatternFill('solid',fgColor=DARK); c.border=bdr
    c2=ws.cell(note,2,'Delete sample rows 4+ and enter your own. Keep rows 1-3 (headers). KIND is optional — auto-detected. Upload via MPCP Creator → Upload Excel.')
    c2.font=Font(color=MUT,size=8,italic=True); c2.fill=PatternFill('solid',fgColor=DARK)
    c2.alignment=Alignment(wrap_text=True); c2.border=bdr
    ws.merge_cells(f'B{note}:N{note}'); ws.row_dimensions[note].height=28
    ws.freeze_panes='A4'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='MPCP_Creator_Template.xlsx')


@app.route('/api/mpcp/upload_excel', methods=['POST'])
def mpcp_upload_excel():
    """
    Parse a MPCP Creator upload Excel (flat format with kind/role/mp_ref/mp_text/
    mp_uom/mp_target/mp_freq/cp_ref/cp_text/cp_uom/cp_target/cp_freq/owner_name/source
    columns) and return structured nodes ready for preview or direct publish.

    The parser is intentionally forgiving:
      - kind column is inferred if blank (role row = role col filled + no mp_ref/cp_ref,
        mp row = mp_ref filled + no cp_ref, cp row = cp_ref filled)
      - Merged/carry-down handled by fill-forward on role and mp columns
      - Returns {nodes: [...], warnings: [...], stats: {roles, mps, cps}}
    """
    if not HAS_OPENPYXL:
        return json_error('openpyxl not installed on server')

    f = request.files.get('file')
    if not f:
        return json_error('No file uploaded')

    try:
        import openpyxl as _xl
        buf = io.BytesIO(f.read())
        wb = _xl.load_workbook(buf, data_only=True)
    except Exception as e:
        return json_error(f'Cannot open Excel file: {e}')

    # Find the upload sheet — prefer 'MPCP_Upload', else first non-Instructions sheet
    target_sheet = None
    for name in wb.sheetnames:
        if 'upload' in name.lower() or 'mpcp' in name.lower():
            target_sheet = wb[name]; break
    if target_sheet is None:
        target_sheet = wb[wb.sheetnames[0]]

    ws = target_sheet

    def cv(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ''

    # Detect header row — scan first 6 rows, pick the one with MOST recognised
    # field names. Handles two-row headers: row 1 = group labels, row 2 = fields.
    KNOWN_HEADERS = {
        'KIND','ROLE','ROLE / OBJECTIVE','ROLE/OBJECTIVE',
        'MP REF','MP_REF','MP TEXT','MP_TEXT','MANAGING POINTS','MANAGING POINT',
        'MP UOM','MP_UOM','UOM','MP TARGET','MP_TARGET','MP FREQ','MP_FREQ',
        'CP REF','CP_REF','CP TEXT','CP_TEXT','CHECKING POINTS','CHECKING POINT',
        'CP UOM','CP_UOM','CP TARGET','CP_TARGET','CP FREQ','CP_FREQ',
        'OWNER NAME','OWNER_NAME','OWNER','CP RESP','CP_RESP','SOURCE','REPORT SOURCE',
    }
    header_row = None; best_score = 0
    for r in range(1, 7):
        row_vals = [cv(r, c).upper().strip() for c in range(1, 16)]
        score = sum(1 for v in row_vals if v in KNOWN_HEADERS)
        if score > best_score:
            best_score = score; header_row = r
    if header_row is None or best_score == 0:
        return json_error('Could not find header row. Ensure the file has columns like MP TEXT, CP TEXT, OWNER NAME.')

    # Map column names → indices
    col_map = {}
    aliases = {
        'kind':['KIND'],'role':['ROLE','ROLE / OBJECTIVE','ROLE/OBJECTIVE'],
        'mp_ref':['MP REF','MP_REF','MP-REF','MPREF'],
        'mp_text':['MP TEXT','MP_TEXT','MANAGING POINTS','MANAGING POINT'],
        'mp_uom':['MP UOM','MP_UOM','UOM'],
        'mp_target':['MP TARGET','MP_TARGET'],
        'mp_freq':['MP FREQ','MP_FREQ'],
        'cp_ref':['CP REF','CP_REF','CP-REF','CPREF'],
        'cp_text':['CP TEXT','CP_TEXT','CHECKING POINTS','CHECKING POINT'],
        'cp_uom':['CP UOM','CP_UOM'],
        'cp_target':['CP TARGET','CP_TARGET'],
        'cp_freq':['CP FREQ','CP_FREQ'],
        'owner_name':['OWNER NAME','OWNER_NAME','OWNER','CP RESP','CP_RESP'],
        'source':['SOURCE','REPORT SOURCE'],
    }
    for c in range(1, 20):
        hdr = cv(header_row, c).upper().strip()
        for field, names in aliases.items():
            if hdr in names and field not in col_map:
                col_map[field] = c

    # mp_ref and cp_ref are OPTIONAL — system auto-generates them when absent
    has_mp_ref_col = 'mp_ref' in col_map
    has_cp_ref_col = 'cp_ref' in col_map

    # Minimum requirement: at least one of mp_text or cp_text must exist
    if 'mp_text' not in col_map and 'cp_text' not in col_map and 'role' not in col_map:
        return json_error(f'Cannot find usable columns. Need at least one of: MP TEXT, CP TEXT, ROLE. Found: {list(col_map.keys())}')

    # Auto-ref counters per module prefix
    _role_counter = [0]
    _mp_counters  = {}   # role_idx -> mp_count
    _cp_counters  = {}   # mp_ref  -> cp_count

    def _auto_mp_ref(role_idx, mp_uom_hint=''):
        _mp_counters.setdefault(role_idx, 0)
        _mp_counters[role_idx] += 1
        prefix = 'HOD-MP'
        return f'{prefix}-{role_idx:02d}.{_mp_counters[role_idx]:02d}'

    def _auto_cp_ref(mp_ref, cp_uom_hint=''):
        _cp_counters.setdefault(mp_ref, 0)
        _cp_counters[mp_ref] += 1
        return f'{mp_ref}-CP{_cp_counters[mp_ref]:02d}'

    SKIP_HINTS = {'role/mp/cp or blank','kind','hint','←','note','example','your role','[your','managing point description','[managing'}

    nodes = []
    warnings = []
    prev_role = ''; prev_role_idx = 0
    prev_mp_ref = ''; prev_mp_text = ''; prev_mp_uom = ''; prev_mp_target = ''; prev_mp_freq = ''
    seen_roles = {}; seen_mps = {}; seen_cps = {}
    stats = {'roles': 0, 'mps': 0, 'cps': 0, 'skipped': 0}
    skipped = []  # detailed skip log: [{row, reason, preview}]

    def get(row, field, default=''):
        c = col_map.get(field)
        return cv(row, c) if c else default

    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [cv(r, c) for c in range(1, 16)]
        if not any(row_vals): continue

        kind_raw = get(r, 'kind').lower().strip()
        role_col = get(r, 'role')
        mp_ref   = get(r, 'mp_ref')
        mp_text  = get(r, 'mp_text')
        mp_uom   = get(r, 'mp_uom')
        mp_tgt   = get(r, 'mp_target')
        mp_freq  = get(r, 'mp_freq') or 'Monthly'
        cp_ref   = get(r, 'cp_ref')
        cp_text  = get(r, 'cp_text')
        cp_uom   = get(r, 'cp_uom')
        cp_tgt   = get(r, 'cp_target')
        cp_freq  = get(r, 'cp_freq') or 'Daily'
        owner    = get(r, 'owner_name')
        source   = get(r, 'source')

        # Skip hint/placeholder rows
        all_text = ' '.join(filter(None,[kind_raw,role_col,mp_text,cp_text])).lower()
        if any(h in all_text for h in SKIP_HINTS):
            skipped.append({'row':r,'reason':'Header/hint row','preview':all_text[:60]})
            stats['skipped'] += 1; continue
        if not any([role_col,mp_ref,mp_text,cp_ref,cp_text]):
            skipped.append({'row':r,'reason':'Completely empty row','preview':''})
            stats['skipped'] += 1; continue

        # ── Infer kind ────────────────────────────────────────────────
        if not kind_raw:
            if role_col and not mp_text and not cp_text and not mp_ref and not cp_ref:
                kind_raw = 'role'
            elif mp_text and not cp_text:
                kind_raw = 'mp'
            elif cp_text and not mp_text:
                kind_raw = 'cp'
            elif mp_text and cp_text:
                # Both filled on same row — treat mp_text as context, cp_text as CP
                kind_raw = 'cp'
            elif mp_ref and not cp_ref:
                kind_raw = 'mp'
            elif cp_ref:
                kind_raw = 'cp'
            elif role_col:
                kind_raw = 'role'
            else:
                skipped.append({'row':r,'reason':'Cannot determine kind (role/mp/cp) — no text or ref found','preview':all_text[:60]})
                stats['skipped'] += 1; continue

        # Fill-forward role context
        if role_col: prev_role = role_col
        if mp_ref or mp_text:
            if mp_ref: prev_mp_ref = mp_ref
            if mp_text: prev_mp_text = mp_text
            if mp_uom: prev_mp_uom = mp_uom
            if mp_tgt: prev_mp_target = mp_tgt
            if mp_freq: prev_mp_freq = mp_freq

        if kind_raw == 'role':
            role_text = role_col or prev_role
            if not role_text: skipped.append({'row':r,'reason':'Role row has no text','preview':''}); stats['skipped'] += 1; continue
            if role_text not in seen_roles:
                _role_counter[0] += 1
                node = {'kind':'role','text':role_text,'ref':'','status':'active','children':[]}
                seen_roles[role_text] = (node, _role_counter[0])
                nodes.append(node)
                stats['roles'] += 1
            prev_role_idx = seen_roles[role_text][1]

        elif kind_raw == 'mp':
            text = mp_text or mp_ref
            if not text: skipped.append({'row':r,'reason':'MP row has no title text','preview':str(mp_ref)}); stats['skipped'] += 1; continue
            # Auto-generate ref if missing
            if not mp_ref:
                mp_ref = _auto_mp_ref(prev_role_idx, mp_uom)
                warnings.append(f'Row {r}: Auto-generated MP ref "{mp_ref}" for "{text[:40]}"')
            if mp_ref not in seen_mps:
                role_text = prev_role
                if role_text not in seen_roles:
                    _role_counter[0] += 1
                    rn = {'kind':'role','text':role_text or 'General','ref':'','status':'active','children':[]}
                    seen_roles[role_text] = (rn, _role_counter[0])
                    nodes.append(rn); stats['roles'] += 1
                prev_role_idx = seen_roles[role_text][1]
                mp_node = {'kind':'mp','ref':mp_ref,'text':text,'uom':mp_uom,'targetVal':mp_tgt,
                           'unit':'','freq':mp_freq,'status':'active','children':[]}
                seen_mps[mp_ref] = mp_node
                seen_roles[role_text][0]['children'].append(mp_node)
                stats['mps'] += 1
            prev_mp_ref = mp_ref

        elif kind_raw == 'cp':
            text = cp_text or cp_ref
            if not text: skipped.append({'row':r,'reason':'CP row has no description text','preview':str(cp_ref)}); stats['skipped'] += 1; continue
            # Determine parent MP — may come from same row's mp_text/mp_ref context
            if mp_text and mp_text != prev_mp_text:
                # Inline MP context on CP row — auto-create MP if not seen
                inline_mp_ref = mp_ref or _auto_mp_ref(prev_role_idx, mp_uom)
                if inline_mp_ref not in seen_mps:
                    role_text = prev_role
                    if role_text not in seen_roles:
                        _role_counter[0] += 1
                        rn = {'kind':'role','text':role_text or 'General','ref':'','status':'active','children':[]}
                        seen_roles[role_text] = (rn, _role_counter[0])
                        nodes.append(rn); stats['roles'] += 1
                    prev_role_idx = seen_roles[role_text][1]
                    mp_node = {'kind':'mp','ref':inline_mp_ref,'text':mp_text,'uom':mp_uom,
                               'targetVal':mp_tgt,'unit':'','freq':mp_freq,'status':'active','children':[]}
                    seen_mps[inline_mp_ref] = mp_node
                    seen_roles[role_text][0]['children'].append(mp_node)
                    stats['mps'] += 1
                    warnings.append(f'Row {r}: Auto-created MP "{inline_mp_ref}" from inline context')
                prev_mp_ref = inline_mp_ref
                prev_mp_text = mp_text

            parent_ref = prev_mp_ref
            # Auto-generate CP ref if missing
            if not cp_ref:
                cp_ref = _auto_cp_ref(parent_ref or 'ROOT', cp_uom)
                warnings.append(f'Row {r}: Auto-generated CP ref "{cp_ref}" for "{text[:40]}"')

            if cp_ref in seen_cps:
                existing = seen_cps[cp_ref]
                if owner and owner != existing.get('owner_name',''):
                    existing.setdefault('extra_owners',[]).append(owner)
                    warnings.append(f'Row {r}: CP {cp_ref} duplicate — added owner "{owner}"')
            else:
                cp_node = {'kind':'cp','ref':cp_ref,'text':text,'uom':cp_uom,'targetVal':cp_tgt,
                           'unit':'','freq':cp_freq,'owner_name':owner,'owner_code':'',
                           'source':source,'parent_mp_ref':parent_ref,'status':'active','children':[]}
                seen_cps[cp_ref] = cp_node
                if parent_ref and parent_ref in seen_mps:
                    seen_mps[parent_ref]['children'].append(cp_node)
                elif nodes and nodes[-1]['children']:
                    nodes[-1]['children'][-1]['children'].append(cp_node)
                    warnings.append(f'Row {r}: CP {cp_ref} attached to last MP (no explicit parent)')
                else:
                    warnings.append(f'Row {r}: CP {cp_ref} orphaned — no parent MP found')
                stats['cps'] += 1
        else:
            skipped.append({'row':r,'reason':f'Unrecognised kind value: "{kind_raw}"','preview':all_text[:60]})
            stats['skipped'] += 1

        # keep seen_roles consistent for fill-forward
        if kind_raw in ('mp','cp') and prev_role in seen_roles:
            prev_role_idx = seen_roles[prev_role][1]

    if not nodes:
        return json_error('No valid data found. Check that MP TEXT or CP TEXT columns have content.')

    return jsonify({'ok': True, 'nodes': nodes, 'warnings': warnings, 'skipped': skipped, 'stats': stats})





# ── ORG TREE ──────────────────────────────────────────────────────────────
@app.route('/api/perf/simple_template')
def perf_simple_template():
    """Generate a rich Excel template pre-filled with real employees & CPs from DB."""
    db   = get_db()
    _ensure_mpcp_status_columns(db)
    emps = db.execute("SELECT emp_code, name FROM employees ORDER BY emp_code").fetchall()
    cps  = db.execute("SELECT ref, title, target FROM cps WHERE status='active' OR status IS NULL ORDER BY ref").fetchall()

    if not HAS_OPENPYXL:
        # Fallback to CSV if openpyxl not installed
        out = io.StringIO()
        w   = csv.writer(out)
        w.writerow(['Date','Emp_Code','CP_Ref','Total','Compliant','Actual_Val','Notes'])
        for e in emps[:3]:
            for c in cps[:2]:
                w.writerow([datetime.date.today().isoformat(), e['emp_code'], c['ref'], '', '', '', ''])
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode()), mimetype='text/csv',
                         as_attachment=True, download_name='MPCP_Simple_Template.csv')

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data Entry ─────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Performance Data'

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    HDR_FILL = PatternFill('solid', fgColor='0A1628')
    HDR_FONT = Font(color='FFFFFF', bold=True, size=11)
    INFO_FILL = PatternFill('solid', fgColor='DBEAFE')
    INFO_FONT = Font(color='1E3A8A', size=10)
    REQ_FILL  = PatternFill('solid', fgColor='FEF9C3')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Date *', 'Emp_Code *', 'CP_Ref *', 'Total *', 'Compliant *', 'Actual_Val', 'Notes']
    col_widths = [14, 16, 20, 10, 12, 13, 30]

    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = '⚡ MPCP Performance Upload Template  —  Fill Date · Emp_Code · CP_Ref · Total · Compliant. Everything else is auto-calculated on import.'
    ws['A1'].font      = Font(bold=True, color='1E3A8A', size=11)
    ws['A1'].fill      = INFO_FILL
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Header row
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Hint row
    hints = [
        'YYYY-MM-DD  e.g. 2025-07-16',
        'Employee code  (see Emp List sheet)',
        'CP reference  (see CP List sheet)',
        'Total units done this period',
        'How many were compliant/on-time',
        'Actual measured value (optional)',
        'Optional remarks'
    ]
    for col_idx, hint in enumerate(hints, 1):
        cell = ws.cell(row=3, column=col_idx, value=hint)
        cell.font      = Font(color='6B7A99', italic=True, size=9)
        cell.fill      = PatternFill('solid', fgColor='F8FAFC')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border    = border
    ws.row_dimensions[3].height = 18

    # Sample rows using real data
    today = datetime.date.today().isoformat()
    sample_rows = []
    for e in emps[:3]:
        for c in cps[:2]:
            sample_rows.append([today, e['emp_code'], c['ref'], 100, 95, '', ''])
    if not sample_rows:
        sample_rows = [[today, 'EMP-001', 'CP-REF', 100, 95, '', '']]

    for row_idx, row_data in enumerate(sample_rows, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border    = border
            if col_idx <= 5:
                cell.fill = REQ_FILL
        ws.row_dimensions[row_idx].height = 16

    # Blank data rows (rows 4+sample to 53)
    start_blank = 4 + len(sample_rows)
    for row_idx in range(start_blank, 54):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border    = border
            if col_idx <= 5:
                cell.fill = REQ_FILL

    # Data validation for Emp_Code column (col B = 2)
    if emps:
        emp_codes = ','.join(f'"{e["emp_code"]}"' for e in emps[:50])
        dv_emp = DataValidation(type='list', formula1=f'"{emp_codes}"', allow_blank=True,
                                showDropDown=False, showErrorMessage=True,
                                error='Use an employee code from the Emp List sheet',
                                errorTitle='Invalid Employee Code')
        ws.add_data_validation(dv_emp)
        dv_emp.sqref = f'B4:B200'

    # Data validation for CP_Ref column (col C = 3)
    if cps:
        cp_refs = ','.join(f'"{c["ref"]}"' for c in cps[:50])
        dv_cp = DataValidation(type='list', formula1=f'"{cp_refs}"', allow_blank=True,
                               showDropDown=False, showErrorMessage=True,
                               error='Use a CP reference from the CP List sheet',
                               errorTitle='Invalid CP Reference')
        ws.add_data_validation(dv_cp)
        dv_cp.sqref = f'C4:C200'

    ws.freeze_panes = 'A4'

    # ── Sheet 2: Employee List ──────────────────────────────────────────
    ws2 = wb.create_sheet('Emp List')
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 30
    ws2.cell(1,1,'Emp_Code').font = Font(bold=True); ws2.cell(1,1).fill = HDR_FILL; ws2.cell(1,1).font = HDR_FONT
    ws2.cell(1,2,'Name').font     = HDR_FONT;        ws2.cell(1,2).fill = HDR_FILL; ws2.cell(1,2).font = HDR_FONT
    for i, e in enumerate(emps, 2):
        ws2.cell(i, 1, e['emp_code'])
        ws2.cell(i, 2, e['name'])

    # ── Sheet 3: CP List ───────────────────────────────────────────────
    ws3 = wb.create_sheet('CP List')
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 16
    for col, hdr in enumerate(['CP_Ref','Title','Target'], 1):
        c = ws3.cell(1, col, hdr); c.fill = HDR_FILL; c.font = HDR_FONT
    for i, c in enumerate(cps, 2):
        ws3.cell(i, 1, c['ref'])
        ws3.cell(i, 2, c['title'])
        ws3.cell(i, 3, c['target'] or '')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='MPCP_Upload_Template.xlsx')


@app.route('/api/perf/template')
def perf_template():
    out = io.StringIO(); w = csv.writer(out); w.writerow(PHDR)
    for s in [
        ["2081-82","Shrawan","EMP-002","HODL-1","LM-VEH-1","Vehicle Border Tracking",410,398,12,97.07,2.93,100,97,"%","C","All tracked"],
        ["2081-82","Shrawan","EMP-002","HODL-1","LM-VEH-2-B","CC within 3 Days",410,390,20,95.12,4.88,3,2.6,"Days","C","On track"],
        ["2081-82","Bhadra","EMP-004","HODL-3","LM-VEH-5-A","Registration 15 Days",380,370,10,97.37,2.63,15,13,"Days","C","Improved"],
    ]: w.writerow(s)
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='MPCP_Perf_Template.csv')


@app.route('/api/perf/export')
def export_perf():
    db = get_db(); fy = request.args.get('fy')
    q  = "SELECT p.*,e.name as emp_name FROM perf p LEFT JOIN employees e ON p.emp_id=e.id"
    args = []
    if fy: q += " WHERE p.fy=?"; args.append(fy)
    rows = db.execute(q+' ORDER BY p.fy DESC,p.bs_month', args).fetchall()
    out  = io.StringIO(); w = csv.writer(out)
    w.writerow(['FY','BS_Month','Quarter','Emp_Code','Emp_Name','MP_Ref','CP_Ref','Metric',
                'Total','Compliant','Non_Compliant','Pct_Compliant','Pct_NC',
                'Target_Val','Actual_Val','Unit','Status','Notes'])
    for r in rows:
        w.writerow([r['fy'],r['bs_month'],r['quarter'],r['emp_code'],r['emp_name'] or '',
                    r['mp_ref'],r['cp_ref'],r['metric'],r['total'],r['compliant'],r['non_compliant'],
                    r['pct_compliant'],r['pct_nc'],r['target_val'],r['actual_val'],r['unit'],r['status'],r['notes']])
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name=f'MPCP_Perf_{fy or "All"}.csv')


@app.route('/api/perf/import', methods=['POST'])
def import_perf():
    f = request.files.get('file')
    if not f: return json_error('No file')
    text   = f.read().decode('utf-8-sig'); reader = csv.DictReader(io.StringIO(text))
    db     = get_db()
    code_map = {r['emp_code']:r['id'] for r in db.execute("SELECT id,emp_code FROM employees WHERE emp_code IS NOT NULL")}
    count  = 0; errs = []; fys_seen = set()
    for i,row in enumerate(reader,2):
        try:
            fy = str(row.get('FY','') or row.get('fy','')).strip()
            if not fy: continue
            fys_seen.add(fy)
            bsm = norm_month(row.get('BS_Month','') or row.get('bs_month',''))
            ec  = str(row.get('Emp_Code','') or row.get('emp_code','')).strip()
            eid = code_map.get(ec,'')
            tot = int(float(row.get('Total','0') or 0)); comp = int(float(row.get('Compliant','0') or 0))
            nc  = tot - comp
            pct_c  = float(row.get('Pct_Compliant','0') or (round(comp/tot*100,2) if tot else 0))
            pct_nc = float(row.get('Pct_NC','0') or (round(nc/tot*100,2) if tot else 0))
            tgt = float(row.get('Target_Val','0') or 0); act = float(row.get('actual_val','0') or 0)
            unit = str(row.get('Unit','%') or '%')
            st   = str(row.get('Status','') or calc_status(act,tgt,unit))
            if st not in ('C','NC'): st = calc_status(act,tgt,unit)
            db.execute("""INSERT OR IGNORE INTO perf
                       (id,fy,bs_month,quarter,emp_id,emp_code,mp_ref,cp_ref,
                        metric,total,compliant,non_compliant,pct_compliant,pct_nc,
                        target_val,actual_val,unit,status,notes)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                       (uid(),fy,bsm,bs_q(bsm),eid,ec,str(row.get('MP_Ref','') or ''),str(row.get('CP_Ref','') or ''),
                        str(row.get('Metric','') or ''),tot,comp,nc,pct_c,pct_nc,tgt,act,unit,st,str(row.get('Notes','') or '')))
            count += 1
        except Exception as e: errs.append(f"Row {i}: {e}")
    for fy in fys_seen: _upd_cache(db, fy)
    db.commit()
    return jsonify({'imported': count, 'errors': errs[:10]})

# ── ANALYTICS — SQL-based  (UPGRADED) ─────────────────────────────────────
@app.route('/api/perf/<pid>', methods=['PUT','DELETE'])
def perf_record(pid):
    db = get_db()
    if request.method == 'DELETE':
        row = db.execute("SELECT fy FROM perf WHERE id=?", (pid,)).fetchone()
        db.execute("DELETE FROM perf WHERE id=?", (pid,))
        if row: _upd_cache(db, row['fy'])
        db.commit()
        return jsonify({'ok': True})
    d      = request.json or {}
    bsm    = norm_month(d.get('bs_month','Shrawan'))
    tot    = int(d.get('total',0)); comp = int(d.get('compliant',0))
    nc     = int(d.get('non_compliant', tot-comp))
    pct_c  = round(comp/tot*100,2) if tot else 0
    pct_nc = round(nc/tot*100,2)   if tot else 0
    db.execute("""UPDATE perf SET fy=?,bs_month=?,quarter=?,emp_id=?,emp_code=?,mp_ref=?,cp_ref=?,
                  metric=?,total=?,compliant=?,non_compliant=?,pct_compliant=?,pct_nc=?,
                  target_val=?,actual_val=?,unit=?,status=?,notes=? WHERE id=?""",
               (d.get('fy'),bsm,bs_q(bsm),d.get('emp_id',''),d.get('emp_code',''),d.get('mp_ref',''),d.get('cp_ref',''),
                d.get('metric',''),tot,comp,nc,pct_c,pct_nc,d.get('target_val',0),d.get('actual_val',0),
                d.get('unit','%'),d.get('status','C'),d.get('notes',''),pid))
    db.commit()
    return jsonify({'ok': True})

PHDR = ['FY','BS_Month','Emp_Code','MP_Ref','CP_Ref','Metric',
        'Total','Compliant','Non_Compliant','Pct_Compliant','Pct_NC',
        'Target_Val','Actual_Val','Unit','Status','Notes']




@app.route('/api/analytics/summary')
def analytics_summary():
    db    = get_db()
    fys   = request.args.get('fys','')
    flist = [x.strip() for x in fys.split(',') if x.strip()]
    ph    = ','.join('?'*len(flist)) if flist else "''"
    wh    = f"WHERE fy IN ({ph})" if flist else ""
    wh_and = f"WHERE fy IN ({ph}) AND" if flist else "WHERE"

    by_fy = R(db.execute(f'''
        SELECT fy,
          SUM(total) as total, SUM(compliant) as compliant, SUM(non_compliant) as nc,
          ROUND(SUM(compliant)*100.0/NULLIF(SUM(total),0),2) as pct_c,
          ROUND(SUM(non_compliant)*100.0/NULLIF(SUM(total),0),2) as pct_nc
        FROM perf {wh} GROUP BY fy ORDER BY fy
    ''', flist).fetchall())

    by_month = R(db.execute(f'''
        SELECT fy, bs_month, quarter,
          SUM(total) as total, SUM(compliant) as compliant,
          ROUND(SUM(compliant)*100.0/NULLIF(SUM(total),0),2) as pct_c
        FROM perf {wh} GROUP BY fy,bs_month ORDER BY fy,bs_month
    ''', flist).fetchall())

    by_mp = R(db.execute(f'''
        SELECT fy, mp_ref,
          SUM(total) as total, SUM(compliant) as compliant,
          ROUND(SUM(compliant)*100.0/NULLIF(SUM(total),0),2) as pct_c
        FROM perf {wh_and} mp_ref!=''
        GROUP BY fy,mp_ref ORDER BY fy,mp_ref
    ''', flist).fetchall())

    by_cp = R(db.execute(f'''
        SELECT fy, mp_ref, cp_ref,
          SUM(total) as total, SUM(compliant) as compliant,
          ROUND(SUM(compliant)*100.0/NULLIF(SUM(total),0),2) as pct_c
        FROM perf {wh_and} cp_ref!=''
        GROUP BY fy,mp_ref,cp_ref ORDER BY fy,mp_ref,cp_ref
    ''', flist).fetchall())

    by_emp = R(db.execute(f'''
        SELECT fy, emp_code,
          SUM(total) as total, SUM(compliant) as compliant,
          ROUND(SUM(compliant)*100.0/NULLIF(SUM(total),0),2) as pct_c
        FROM perf {wh_and} emp_code!=''
        GROUP BY fy,emp_code ORDER BY fy,emp_code
    ''', flist).fetchall())

    return jsonify({'by_fy': by_fy, 'by_month': by_month,
                    'by_mp': by_mp, 'by_cp': by_cp, 'by_emp': by_emp})

@app.route('/api/calendar')
def calendar_api():
    return jsonify({'bs_months': BS_MONTHS, 'quarter_map': BS_Q})

@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'index.html'), 'r', encoding='utf-8') as f:
        return f.read()

# ── SAMPLE FILE DOWNLOADS ──────────────────────────────────────────────────
@app.route('/api/samples/<fname>')
def sample_file(fname):
    safe = {'employees':'Sample_Employees.xlsx',
            'mps':'Sample_ManagingPoints.xlsx',
            'cps':'Sample_CheckingPoints.xlsx',
            'perf':'Sample_Performance.csv'}
    if fname not in safe: return 'Not found', 404
    path = os.path.join(os.path.dirname(__file__), safe[fname])
    if not os.path.exists(path): return 'File not found', 404
    return send_file(path, as_attachment=True, download_name=safe[fname])



# ── CASCADE LINKS (frontend-compatible URLs) ──────────────────────────────
# Frontend calls /api/cascade  (not /api/cascade_links)

@app.route('/api/cascade', methods=['GET', 'POST'])
def cascade_api():
    db = get_db()
    if request.method == 'GET':
        rows = db.execute("""
            SELECT cl.*,
              se.name  as sup_name,  se.emp_code as sup_code,
              sub.name as sub_name, sub.emp_code as sub_code,
              cp.ref   as cp_ref,   cp.title     as cp_title,
              mp.ref   as mp_ref,   mp.title     as mp_title
            FROM cascade_links cl
            LEFT JOIN employees se   ON cl.superior_emp_id    = se.id
            LEFT JOIN employees sub  ON cl.subordinate_emp_id = sub.id
            LEFT JOIN cps       cp   ON cl.superior_cp_id     = cp.id
            LEFT JOIN mps       mp   ON cl.subordinate_mp_id  = mp.id
            ORDER BY se.name
        """).fetchall()
        return jsonify(R(rows))

    d = request.json or {}
    # Support both field naming conventions from frontend
    sup_emp  = d.get('superior_emp_id') or d.get('sup_emp_id') or d.get('parent_emp_id')
    sup_cp   = d.get('superior_cp_id')  or d.get('sup_cp_id')  or d.get('cp_id')
    sub_emp  = d.get('subordinate_emp_id') or d.get('sub_emp_id') or d.get('child_emp_id')
    sub_mp   = d.get('subordinate_mp_id') or d.get('sub_mp_id') or d.get('mp_id') or ''

    if not sup_emp or not sup_cp or not sub_emp:
        return jsonify({'error': 'Missing required fields: superior_emp_id, superior_cp_id, subordinate_emp_id'}), 400

    # Duplicate check
    existing = db.execute(
        'SELECT id FROM cascade_links WHERE superior_cp_id=? AND subordinate_emp_id=?',
        (sup_cp, sub_emp)
    ).fetchone()
    if existing:
        return jsonify({'error': 'Link already exists', 'id': existing['id']}), 409

    auto_created = 0
    sub_mp = sub_mp.strip()
    if not sub_mp:
        cp = db.execute('SELECT * FROM cps WHERE id=?', (sup_cp,)).fetchone()
        if not cp:
            # Try by ref
            cp = db.execute('SELECT * FROM cps WHERE ref=?', (sup_cp,)).fetchone()
        if cp:
            new_ref   = f"AUTO-{cp['ref']}"
            new_title = f"[Auto] {(cp['title'] or cp['ref'] or 'Untitled')}"
            ex_mp = db.execute('SELECT id FROM mps WHERE ref=?', (new_ref,)).fetchone()
            if ex_mp:
                sub_mp = ex_mp['id']
            else:
                sub_mp = uid()
                db.execute('INSERT INTO mps(id,ref,title,target,freq,kpi_c,kpi_nc,kpi_total) VALUES(?,?,?,?,?,?,?,?)',
                           (sub_mp, new_ref, new_title, (cp['target'] or ''), (cp['freq'] or 'Monthly'), 0, 0, 0))
                db.execute('INSERT OR IGNORE INTO mp_owners VALUES(?,?)', (sub_mp, sub_emp))
            auto_created = 1

    link_id = uid()
    db.execute('INSERT INTO cascade_links VALUES(?,?,?,?,?,?,?)',
               (link_id, sup_emp, sup_cp, sub_emp, sub_mp, auto_created,
                datetime.datetime.now().isoformat()))
    db.commit()
    return jsonify({
        'id': link_id, 'subordinate_mp_id': sub_mp,
        'auto_created': bool(auto_created), 'ok': True
    })


@app.route('/api/cascade/<lid>', methods=['DELETE'])
def cascade_del(lid):
    db = get_db()
    row = db.execute('SELECT * FROM cascade_links WHERE id=?', (lid,)).fetchone()
    if not row: return jsonify({'error': 'Not found'}), 404
    if row['auto_created']:
        db.execute('DELETE FROM mps WHERE id=?', (row['subordinate_mp_id'],))
        db.execute('DELETE FROM mp_owners WHERE mp_id=?', (row['subordinate_mp_id'],))
    db.execute('DELETE FROM cascade_links WHERE id=?', (lid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/cascade/tree')
def cascade_tree():
    db  = get_db()
    eid = request.args.get('emp_id')

    # Build tree: emp → roles → mps → cps → cascade children
    emp_q = "SELECT * FROM employees" + (" WHERE id=?" if eid else "") + " ORDER BY level, name"
    emps  = R(db.execute(emp_q, (eid,) if eid else ()).fetchall())

    links = R(db.execute("""
        SELECT cl.*,
          se.name as sup_name, sub.name as sub_name,
          cp.ref as cp_ref, cp.title as cp_title,
          mp.ref as mp_ref
        FROM cascade_links cl
        LEFT JOIN employees se  ON cl.superior_emp_id    = se.id
        LEFT JOIN employees sub ON cl.subordinate_emp_id = sub.id
        LEFT JOIN cps cp        ON cl.superior_cp_id     = cp.id
        LEFT JOIN mps mp        ON cl.subordinate_mp_id  = mp.id
    """).fetchall())

    # Index links by superior_cp_id for fast lookup
    by_cp = {}
    for lnk in links:
        key = lnk['superior_cp_id']
        by_cp.setdefault(key, []).append(lnk)

    tree = []
    for emp in emps:
        eid_  = emp['id']
        roles = R(db.execute("""
            SELECT r.* FROM roles r
            JOIN emp_roles er ON r.id = er.role_id
            WHERE er.emp_id=?
        """, (eid_,)).fetchall())

        role_nodes = []
        for role in roles:
            mps_ = R(db.execute("""
                SELECT m.* FROM mps m
                JOIN role_mps rm ON m.id = rm.mp_id
                WHERE rm.role_id=?
            """, (role['id'],)).fetchall())

            mp_nodes = []
            for mp in mps_:
                cps_ = R(db.execute("""
                    SELECT c.* FROM cps c WHERE c.mp_id=?
                """, (mp['id'],)).fetchall())

                cp_nodes = []
                for cp in cps_:
                    children = by_cp.get(cp['id'], [])
                    cp_nodes.append({**cp, 'cascade_children': children})

                mp_nodes.append({**mp, 'cps': cp_nodes})
            role_nodes.append({**role, 'mps': mp_nodes})

        tree.append({**emp, 'roles': role_nodes})

    return jsonify({'tree': tree, 'links': links})


@app.route('/api/org/cascade_assign', methods=['POST'])
def cascade_assign():
    """Alternative cascade creation endpoint used by org chart view"""
    d       = request.json or {}
    sup_emp = d.get('parent_emp_id') or d.get('superior_emp_id')
    cp_id   = d.get('cp_id') or d.get('superior_cp_id')
    sub_emp = d.get('child_emp_id') or d.get('subordinate_emp_id')
    sub_mp  = (d.get('mp_id') or d.get('subordinate_mp_id') or '').strip()

    if not sup_emp or not cp_id or not sub_emp:
        return jsonify({'error': 'Missing fields: superior_emp_id, superior_cp_id, subordinate_emp_id'}), 400

    db = get_db()

    # Duplicate check
    existing = db.execute(
        'SELECT id FROM cascade_links WHERE superior_cp_id=? AND subordinate_emp_id=?',
        (cp_id, sub_emp)
    ).fetchone()
    if existing:
        return jsonify({'ok': True, 'id': existing['id'], 'existing': True})

    auto_created = 0
    if not sub_mp:
        cp = db.execute('SELECT * FROM cps WHERE id=?', (cp_id,)).fetchone()
        if not cp:
            cp = db.execute('SELECT * FROM cps WHERE ref=?', (cp_id,)).fetchone()
        if cp:
            new_ref   = f"AUTO-{cp['ref'] or uid()}"
            new_title = f"[Auto] {(cp['title'] or cp['ref'] or 'Untitled')}"
            ex_mp = db.execute('SELECT id FROM mps WHERE ref=?', (new_ref,)).fetchone()
            if ex_mp:
                sub_mp = ex_mp['id']
            else:
                sub_mp = uid()
                db.execute('INSERT INTO mps(id,ref,title,target,freq,kpi_c,kpi_nc,kpi_total) VALUES(?,?,?,?,?,?,?,?)',
                           (sub_mp, new_ref, new_title,
                            (cp['target'] or ''), (cp['freq'] or 'Monthly'), 0, 0, 0))
                db.execute('INSERT OR IGNORE INTO mp_owners VALUES(?,?)', (sub_mp, sub_emp))
            auto_created = 1

    lid = uid()
    db.execute('INSERT INTO cascade_links VALUES(?,?,?,?,?,?,?)',
               (lid, sup_emp, cp_id, sub_emp, sub_mp, auto_created,
                datetime.datetime.now().isoformat()))
    db.commit()
    return jsonify({'ok': True, 'id': lid, 'auto_created': bool(auto_created)})


@app.route('/api/bs_today')
def bs_today():
    dt  = datetime.datetime.now()
    fy, bsm = ad_date_to_fy_and_month(dt)
    return jsonify({'date': dt.strftime('%Y-%m-%d'), 'fy': fy,
                    'bs_month': bsm, 'quarter': bs_q(bsm)})

# ── SECTORS (stub — keeps frontend from 404-ing) ──────────────────────────
@app.route('/api/sectors', methods=['GET','POST'])
def sectors_api():
    db = get_db()
    # Ensure table exists
    db.execute('''CREATE TABLE IF NOT EXISTS sectors(
        id TEXT PRIMARY KEY, code TEXT NOT NULL, name TEXT NOT NULL,
        description TEXT DEFAULT '', color TEXT DEFAULT '#475569',
        sort_order INTEGER DEFAULT 0)''')
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM sectors ORDER BY sort_order,name").fetchall()
        return jsonify(R(rows))
    d   = request.json or {}
    sid = d.get('id') or uid()
    db.execute("INSERT OR REPLACE INTO sectors(id,code,name,description,color,sort_order) VALUES(?,?,?,?,?,?)",
               (sid, d.get('code',''), d.get('name',''), d.get('description',''),
                d.get('color','#475569'), d.get('sort_order',0)))
    db.commit()
    return jsonify({'id': sid})

@app.route('/api/sectors/<sid>', methods=['PUT','DELETE'])
def sector_api(sid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM sectors WHERE id=?", (sid,)); db.commit()
        return jsonify({'ok': True})
    d = request.json or {}
    db.execute("UPDATE sectors SET code=?,name=?,description=?,color=?,sort_order=? WHERE id=?",
               (d.get('code',''), d.get('name',''), d.get('description',''),
                d.get('color','#475569'), d.get('sort_order',0), sid))
    db.commit()
    return jsonify({'ok': True})

# ── LOCATIONS ─────────────────────────────────────────────────────────────

# ── LOCATIONS ─────────────────────────────────────────────────────────────

@app.route('/api/locations', methods=['GET','POST'])
def locations_api():
    db = get_db()
    # Ensure table exists with correct schema
    db.execute("""CREATE TABLE IF NOT EXISTS locations(
        id TEXT PRIMARY KEY,
        code TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        address TEXT DEFAULT '',
        type TEXT DEFAULT 'Branch',
        dept TEXT DEFAULT 'Ops',
        active INTEGER DEFAULT 1
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS location_employees(
        loc_id TEXT, emp_id TEXT, PRIMARY KEY(loc_id, emp_id)
    )""")
    # Add missing columns if upgrading from old schema
    existing = [r[1] for r in db.execute("PRAGMA table_info(locations)").fetchall()]
    if 'dept'   not in existing: db.execute("ALTER TABLE locations ADD COLUMN dept TEXT DEFAULT 'Ops'")
    if 'active' not in existing: db.execute("ALTER TABLE locations ADD COLUMN active INTEGER DEFAULT 1")
    db.commit()

    if request.method == 'GET':
        locs = R(db.execute("SELECT * FROM locations ORDER BY code").fetchall())
        for loc in locs:
            loc['emp_ids'] = [r['emp_id'] for r in db.execute(
                "SELECT emp_id FROM location_employees WHERE loc_id=?", (loc['id'],))]
        return jsonify(locs)

    d      = request.json or {}
    lid    = d.get('id') or uid()
    code   = d.get('code','').strip()
    name   = d.get('name','').strip()
    if not code or not name:
        return jsonify({'error':'code and name required'}), 400
    log_audit("LOC_SAVE","location",lid,"Saved location "+str(lid))
    db.execute(
        "INSERT OR REPLACE INTO locations(id,code,name,address,type,dept,active) VALUES(?,?,?,?,?,?,?)",
        (lid, code, name, d.get('address',''), d.get('type','Branch'), d.get('dept','Ops'), 1 if d.get('active',True) else 0)
    )
    db.execute("DELETE FROM location_employees WHERE loc_id=?", (lid,))
    for eid in d.get('emp_ids',[]):
        db.execute("INSERT OR IGNORE INTO location_employees(loc_id,emp_id) VALUES(?,?)", (lid, eid))
    db.commit()
    return jsonify({'id': lid})


@app.route('/api/locations/<lid>', methods=['PUT','DELETE'])
def location_api(lid):
    db = get_db()
    if request.method == 'DELETE':
        db.execute("DELETE FROM locations WHERE id=?", (lid,))
        db.execute("DELETE FROM location_employees WHERE loc_id=?", (lid,))
        db.commit()
        return jsonify({'ok': True})
    d      = request.json or {}
    code   = d.get('code','').strip()
    name   = d.get('name','').strip()
    if not code or not name:
        return jsonify({'error':'code and name required'}), 400
    db.execute(
        "UPDATE locations SET code=?,name=?,address=?,type=?,dept=?,active=? WHERE id=?",
        (code, name, d.get('address',''), d.get('type','Branch'), d.get('dept','Ops'), 1 if d.get('active',True) else 0, lid)
    )
    db.execute("DELETE FROM location_employees WHERE loc_id=?", (lid,))
    for eid in d.get('emp_ids',[]):
        db.execute("INSERT OR IGNORE INTO location_employees(loc_id,emp_id) VALUES(?,?)", (lid, eid))
    db.commit()
    return jsonify({'ok': True})


# ── BULK DELETES ───────────────────────────────────────────────────────────
@app.route('/api/perf/bulk_delete', methods=['POST'])
def perf_bulk_delete():
    d   = request.json or {}
    fy  = d.get('fy','').strip()
    if not fy: return json_error('FY required')
    lock = get_db().execute('SELECT locked FROM perf_cache WHERE fy=?', (fy,)).fetchone()
    if lock and lock['locked']: return json_error(f'FY {fy} is locked', 403)
    db   = get_db()
    q2   = "SELECT id FROM perf WHERE fy=?"; args = [fy]
    if d.get('month'):    q2 += " AND bs_month=?";  args.append(d['month'])
    if d.get('emp_code'): q2 += " AND emp_code=?";  args.append(d['emp_code'])
    if d.get('mp_ref'):   q2 += " AND mp_ref=?";    args.append(d['mp_ref'])
    if d.get('cp_ref'):   q2 += " AND cp_ref=?";    args.append(d['cp_ref'])
    rows = db.execute(q2, args).fetchall()
    if d.get('preview'): return jsonify({'count': len(rows)})
    db.execute(q2.replace("SELECT id","DELETE"), args)
    _upd_cache(db, fy); db.commit()
    return jsonify({'deleted': len(rows)})

@app.route('/api/perf/bulk_delete_fy', methods=['POST'])
def perf_bulk_delete_fy():
    d  = request.json or {}
    fy = d.get('fy','').strip()
    if not fy: return json_error('FY required')
    db = get_db()
    cnt = db.execute("SELECT COUNT(*) FROM perf WHERE fy=?", (fy,)).fetchone()[0]
    db.execute("DELETE FROM perf WHERE fy=?", (fy,))
    _upd_cache(db, fy); db.commit()
    return jsonify({'deleted': cnt})

@app.route('/api/employees/bulk_delete', methods=['POST'])
def employees_bulk_delete():
    ids = (request.json or {}).get('ids', [])
    if not ids: return jsonify({'deleted': 0})
    db = get_db()
    for eid in ids:
        for t,c in [('employees','id'),('mp_owners','emp_id'),('cp_owners','emp_id'),
                    ('emp_roles','emp_id'),('emp_mps','emp_id'),('emp_cps','emp_id')]:
            db.execute(f"DELETE FROM {t} WHERE {c}=?", (eid,))
    db.commit(); return jsonify({'deleted': len(ids)})

@app.route('/api/mps/bulk_delete', methods=['POST'])
def mps_bulk_delete():
    ids = (request.json or {}).get('ids', [])
    db  = get_db()
    for mid in ids:
        db.execute("DELETE FROM mps WHERE id=?", (mid,))
        db.execute("DELETE FROM mp_owners WHERE mp_id=?", (mid,))
        db.execute("UPDATE cps SET mp_id='' WHERE mp_id=?", (mid,))
    db.commit(); return jsonify({'deleted': len(ids)})

@app.route('/api/cps/bulk_delete', methods=['POST'])
def cps_bulk_delete():
    ids = (request.json or {}).get('ids', [])
    db  = get_db()
    for cid in ids:
        db.execute("DELETE FROM cps WHERE id=?", (cid,))
        db.execute("DELETE FROM cp_owners WHERE cp_id=?", (cid,))
    db.commit(); return jsonify({'deleted': len(ids)})

# ── SIMPLE PERF IMPORT TEMPLATE ────────────────────────────────────────────

@app.route('/api/perf/import_simple', methods=['POST'])
def import_perf_simple():
    f = request.files.get('file')
    if not f: return json_error('No file')
    ext  = f.filename.lower().split('.')[-1]
    db   = get_db()
    code_map = {r['emp_code']:r['id'] for r in db.execute(
        "SELECT id,emp_code FROM employees WHERE emp_code IS NOT NULL")}
    cp_map   = {r['ref']:dict(r) for r in db.execute("SELECT * FROM cps")}
    mp_map   = {r['id']:r['ref'] for r in db.execute("SELECT id,ref FROM mps")}
    count = 0; errs = []; total_rows = 0; fys_seen = set()
    rows = []
    if ext in ('xlsx','xls'):
        if not HAS_OPENPYXL: return json_error('pip install openpyxl')
        wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
        hdrs = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: str(v if v is not None else '').strip() for h,v in zip(hdrs, row)})
    else:
        text = f.read().decode('utf-8-sig')
        rows = list(csv.DictReader(io.StringIO(text)))
    total_rows = len(rows)
    for i, row in enumerate(rows, 2):
        try:
            row = {k.lower().strip(): v for k, v in row.items()}  # normalize to lowercase
            date_str = (row.get('date') or '').strip()
            ec  = (row.get('emp_code') or '').strip()
            cpr = (row.get('cp_ref')   or '').strip()
            if not date_str or not ec or not cpr: continue
            # Skip hint/instruction rows
            if date_str.lower().startswith('yyyy') or ec.lower().startswith('employee'): continue
            try:
                dt  = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                fy, bsm = ad_date_to_fy_and_month(dt)
            except:
                errs.append(f"Row {i}: invalid date '{date_str}'"); continue
            fys_seen.add(fy)
            lock = db.execute('SELECT locked FROM perf_cache WHERE fy=?', (fy,)).fetchone()
            if lock and lock['locked']:
                errs.append(f"Row {i}: FY {fy} is locked"); continue
            eid = code_map.get(ec,'')
            cp  = cp_map.get(cpr)
            if not cp: errs.append(f"Row {i}: CP '{cpr}' not found"); continue
            mp_ref = mp_map.get(cp['mp_id'],'') if cp.get('mp_id') else ''
            raw_tgt = (cp.get('target') or '').strip()
            tgt_num = ''.join(c for c in raw_tgt if c.isdigit() or c=='.')
            try: tgt_val = float(tgt_num)
            except: tgt_val = 0.0
            unit = raw_tgt.split()[-1] if raw_tgt else '%'
            total    = int(float(row.get('total','0')    or 0))
            compliant= int(float(row.get('compliant','0')or 0))
            if total <= 0: errs.append(f"Row {i}: total must be > 0"); continue
            if compliant > total: compliant = total
            nc     = total - compliant
            pct_c  = round(compliant/total*100,2)
            pct_nc = round(nc/total*100,2)
            act_val= float(row.get('Actual_Val','0') or 0)
            status = calc_status(act_val, tgt_val, unit)
            metric = (cp['title'] or cp['ref'] or '')[:60]
            notes  = (row.get('notes') or '').strip()
            existing = db.execute(
                'SELECT id FROM perf WHERE fy=? AND bs_month=? AND emp_id=? AND cp_ref=?',
                (fy, bsm, eid, cpr)).fetchone()
            pid = existing['id'] if existing else uid()
            db.execute("""INSERT OR REPLACE INTO perf
               (id,fy,bs_month,quarter,emp_id,emp_code,mp_ref,cp_ref,
                metric,total,compliant,non_compliant,pct_compliant,pct_nc,
                target_val,actual_val,unit,status,notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                       (pid,fy,bsm,bs_q(bsm),eid,ec,mp_ref,cpr,metric,
                        total,compliant,nc,pct_c,pct_nc,tgt_val,act_val,unit,status,notes))
            count += 1
        except Exception as e: errs.append(f"Row {i}: {e}")
    for fy in fys_seen: _upd_cache(db, fy)
    db.commit()
    return jsonify({'imported': count, 'total_rows': total_rows, 'errors': errs[:20]})

# ── PERF EXCEPTIONS ────────────────────────────────────────────────────────

@app.route('/api/org/tree')
def org_tree():
    db   = get_db()
    fy   = request.args.get('fy', '2081-82')
    loc  = request.args.get('loc','')
    sect = request.args.get('sect','')
    emps = R(db.execute("SELECT * FROM employees ORDER BY level,name").fetchall())
    perf_rows = R(db.execute("SELECT * FROM perf WHERE fy=?", (fy,)).fetchall())
    # Build compliance per emp
    def emp_perf(eid, ec):
        rows = [p for p in perf_rows if p['emp_id']==eid or p['emp_code']==ec]
        tot  = sum(r.get('total',1) for r in rows)
        comp = sum(r.get('compliant', 1 if r.get('status')=='C' else 0) for r in rows)
        pct  = round(comp/tot*100,1) if tot else None
        return {'tot':tot,'comp':comp,'pct':pct}
    def build(eid):
        emp  = next((e for e in emps if e['id']==eid), None)
        if not emp: return None
        children_raw = [e for e in emps if e.get('manager_id')==eid]
        children = [build(c['id']) for c in children_raw]
        children = [c for c in children if c]
        own = emp_perf(eid, emp.get('emp_code',''))
        # rollup
        def rollup(node):
            t = node['own']['tot']; c = node['own']['comp']
            for ch in (node.get('children') or []):
                rt,rc = rollup(ch); t+=rt; c+=rc
            return t,c
        rt,rc = rollup({'own':own,'children':children})
        rpct  = round(rc/rt*100,1) if rt else None
        mps_  = R(db.execute("SELECT m.* FROM mps m JOIN mp_owners o ON m.id=o.mp_id WHERE o.emp_id=?", (eid,)).fetchall())
        roles_= R(db.execute("SELECT r.* FROM roles r JOIN emp_roles er ON r.id=er.role_id WHERE er.emp_id=?", (eid,)).fetchall())
        for mp in mps_:
            mp['cps'] = R(db.execute("SELECT * FROM cps WHERE mp_id=?", (mp['id'],)).fetchall())
        return {**emp, 'own':own, 'rollup':{'tot':rt,'comp':rc,'pct':rpct},
                'children':children, 'mps':mps_, 'roles':roles_, 'locations':[]}
    roots = [e for e in emps if not e.get('manager_id') or
             not any(x['id']==e['manager_id'] for x in emps)]
    tree  = [build(r['id']) for r in roots]
    tree  = [t for t in tree if t]
    return jsonify(tree)

@app.route('/api/org/move', methods=['POST'])
def org_move():
    d   = request.json or {}
    eid = d.get('emp_id'); mgr = d.get('new_manager_id')
    if not eid: return json_error('emp_id required')
    db  = get_db()
    db.execute("UPDATE employees SET manager_id=? WHERE id=?", (mgr or None, eid))
    db.commit(); return jsonify({'ok': True})

@app.route('/api/org/assign_mp', methods=['POST'])
def org_assign_mp():
    d  = request.json or {}
    db = get_db()
    db.execute("INSERT OR IGNORE INTO emp_mps VALUES(?,?)", (d.get('emp_id',''), d.get('mp_id','')))
    db.execute("INSERT OR IGNORE INTO mp_owners VALUES(?,?)", (d.get('mp_id',''), d.get('emp_id','')))
    db.commit(); return jsonify({'ok': True})

@app.route('/api/org/assign_cp', methods=['POST'])
def org_assign_cp():
    d  = request.json or {}
    db = get_db()
    db.execute("INSERT OR IGNORE INTO emp_cps VALUES(?,?)", (d.get('emp_id',''), d.get('cp_id','')))
    db.execute("INSERT OR IGNORE INTO cp_owners VALUES(?,?)", (d.get('cp_id',''), d.get('emp_id','')))
    db.commit(); return jsonify({'ok': True})

# ── ANALYTICS WIDGET ───────────────────────────────────────────────────────
@app.route('/api/analytics/widget')
def analytics_widget():
    db  = get_db()
    fy  = request.args.get('fy','2081-82')
    metric = request.args.get('metric','overall')
    loc    = request.args.get('loc','')
    emp_id = request.args.get('emp_id','')
    sector = request.args.get('sector','')
    mp_ref = request.args.get('mp_ref','')
    q2  = "SELECT * FROM perf WHERE fy=?"; args=[fy]
    if emp_id: q2+=" AND emp_id=?";  args.append(emp_id)
    if mp_ref: q2+=" AND mp_ref=?";  args.append(mp_ref)
    rows = R(db.execute(q2, args).fetchall())
    tot  = sum(r.get('total',1) for r in rows)
    comp = sum(r.get('compliant',0) for r in rows)
    nc   = tot - comp
    pct  = round(comp/tot*100,2) if tot else 0
    summary = {'tot':tot,'comp':comp,'nc':nc,'pct':pct}
    if metric == 'overall':
        return jsonify({'summary': summary, 'labels': ['Compliant','NC'], 'values': [comp, nc]})
    if metric == 'by_month':
        BS_M = ["Shrawan","Bhadra","Ashwin","Kartik","Mangsir","Poush","Magh","Falgun","Chaitra","Baisakh","Jestha","Ashadh"]
        labels=[]; values=[]
        for m in BS_M:
            mr = [r for r in rows if r['bs_month']==m]
            if not mr: continue
            mt = sum(r.get('total',1) for r in mr)
            mc = sum(r.get('compliant',0) for r in mr)
            labels.append(m[:3]); values.append(round(mc/mt*100,1) if mt else 0)
        return jsonify({'summary':summary,'labels':labels,'values':values})
    if metric == 'by_mp':
        by_mp = {}
        for r in rows:
            ref = r.get('mp_ref','')
            if not ref: continue
            by_mp.setdefault(ref, {'tot':0,'comp':0})
            by_mp[ref]['tot']+=r.get('total',1); by_mp[ref]['comp']+=r.get('compliant',0)
        labels=[]; values=[]
        for ref,d2 in sorted(by_mp.items()):
            labels.append(ref)
            values.append(round(d2['comp']/d2['tot']*100,1) if d2['tot'] else 0)
        return jsonify({'summary':summary,'labels':labels,'values':values})
    if metric == 'by_emp':
        by_e = {}
        for r in rows:
            ec = r.get('emp_code',''); 
            if not ec: continue
            by_e.setdefault(ec, {'tot':0,'comp':0})
            by_e[ec]['tot']+=r.get('total',1); by_e[ec]['comp']+=r.get('compliant',0)
        labels=[]; values=[]
        for ec,d2 in sorted(by_e.items()):
            labels.append(ec)
            values.append(round(d2['comp']/d2['tot']*100,1) if d2['tot'] else 0)
        return jsonify({'summary':summary,'labels':labels,'values':values})
    return jsonify({'summary':summary,'labels':[],'values':[]})


@app.route('/api/analytics/employee_trend/<emp_code>')
def employee_trend(emp_code):
    db = get_db()
    fy = request.args.get('fy', '')
    BS_MONTHS = ['Shrawan','Bhadra','Ashwin','Kartik','Mangsir','Poush',
                 'Magh','Falgun','Chaitra','Baisakh','Jestha','Ashadh']
    params = [emp_code]
    sql = ("SELECT bs_month, fy, COUNT(*) tot, "
           "SUM(CASE WHEN status='C' THEN 1 ELSE 0 END) comp "
           "FROM perf WHERE emp_code=?")
    if fy: sql += " AND fy=?"; params.append(fy)
    sql += " GROUP BY fy, bs_month"
    rows = db.execute(sql, params).fetchall()
    # Build month-indexed result
    result = {}
    for r in rows:
        key = r['fy']
        if key not in result: result[key] = {m: None for m in BS_MONTHS}
        tot = r['tot'] or 0
        comp = r['comp'] or 0
        result[key][r['bs_month']] = round(comp/tot*100, 1) if tot else 0
    return jsonify({'months': BS_MONTHS, 'data': result})

@app.route('/api/analytics/by_location')
def analytics_by_location():
    db = get_db()
    fy = request.args.get('fy','')
    params = []
    sql = """
        SELECT
            COALESCE(NULLIF(p.loc,''),
                (SELECT l.name FROM employees e
                 JOIN emp_locations el ON el.emp_id = e.id
                 JOIN locations l ON l.id = el.loc_id
                 WHERE e.emp_code = p.emp_code
                 ORDER BY el.is_primary DESC LIMIT 1),
                'Unassigned') grp,
            COUNT(*) tot,
            SUM(CASE WHEN p.status='C' THEN 1 ELSE 0 END) comp
        FROM perf p WHERE 1=1
    """
    if fy:
        sql += " AND p.fy=?"; params.append(fy)
    sql += " GROUP BY 1 ORDER BY 1"
    try:
        rows = db.execute(sql, params).fetchall()
    except Exception:
        rows = []
    result = {}
    for r in rows:
        tot = r[1] or 0; comp = r[2] or 0
        loc = r[0] or 'Unassigned'
        result[loc] = {
            'total': tot, 'compliant': comp,
            'non_compliant': tot - comp,
            'pct': round(comp/tot*100, 1) if tot else 0
        }
    return jsonify(result)

# ── DASHBOARD LAYOUTS ──────────────────────────────────────────────────────
@app.route('/api/analytics/by_sector')
def analytics_by_sector():
    db = get_db()
    fy = request.args.get('fy','')
    params = []
    sql = ('SELECT COALESCE(NULLIF(e.dept,\'\'),\'Unassigned\') grp,'
           ' COUNT(DISTINCT p.emp_code) emp_count,'
           ' COUNT(*) tot,'
           ' SUM(CASE WHEN p.status=\'C\' THEN 1 ELSE 0 END) comp'
           ' FROM perf p'
           ' LEFT JOIN employees e ON e.emp_code=p.emp_code'
           ' WHERE 1=1')
    if fy:
        sql += ' AND p.fy=?'; params.append(fy)
    sql += ' GROUP BY 1 ORDER BY 1'
    rows = db.execute(sql, params).fetchall()
    result = {}
    for r in rows:
        tot = r[2] or 0; comp = r[3] or 0
        grp = r[0] or 'Unassigned'
        result[grp] = {
            'name': grp, 'emp_count': r[1] or 0,
            'total': tot, 'compliant': comp,
            'nc': tot - comp,
            'pct': round(comp/tot*100,1) if tot else 0,
            'color': '#475569'
        }
    return jsonify(result)


@app.route('/api/dashboard_layouts', methods=['GET','POST'])
def dashboard_layouts():
    db = get_db()
    # Create table if missing — use layout_json to match existing DB schema
    db.execute('''CREATE TABLE IF NOT EXISTS dashboard_layouts(
        id TEXT PRIMARY KEY, name TEXT NOT NULL, user TEXT DEFAULT 'default',
        layout_json TEXT DEFAULT '[]', created_at TEXT, updated_at TEXT)''')
    # Detect which column name exists: layout_json or layout
    cols = {r[1] for r in db.execute("PRAGMA table_info(dashboard_layouts)").fetchall()}
    layout_col = 'layout_json' if 'layout_json' in cols else 'layout'
    # Add layout_json column if only layout exists
    if layout_col == 'layout' and 'layout_json' not in cols:
        try: db.execute("ALTER TABLE dashboard_layouts ADD COLUMN layout_json TEXT DEFAULT '[]'")
        except Exception: pass
        layout_col = 'layout'
    if request.method == 'GET':
        user = request.args.get('user','default')
        rows = R(db.execute("SELECT * FROM dashboard_layouts WHERE user=? ORDER BY created_at", (user,)).fetchall())
        for r in rows:
            raw = r.get('layout_json') or r.get('layout') or '[]'
            try: r['layout'] = json.loads(raw)
            except: r['layout'] = []
        return jsonify(rows)
    d   = request.json or {}
    lid = d.get('id') or uid()
    now = datetime.datetime.now().isoformat()
    layout_data = json.dumps(d.get('layout', []))
    try:
        db.execute(f"INSERT OR REPLACE INTO dashboard_layouts(id,name,user,{layout_col},created_at,updated_at) VALUES(?,?,?,?,?,?)",
                   (lid, d.get('name','Layout'), d.get('user','default'), layout_data, now, now))
    except Exception:
        db.execute("INSERT OR REPLACE INTO dashboard_layouts VALUES(?,?,?,?,?,?)",
                   (lid, d.get('name','Layout'), d.get('user','default'), layout_data, now, now))
    db.commit()
    return jsonify({'id': lid})

@app.route('/api/dashboard_layouts/<lid>', methods=['DELETE'])
def delete_dashboard_layout(lid):
    db = get_db()
    db.execute("DELETE FROM dashboard_layouts WHERE id=?", (lid,)); db.commit()
    return jsonify({'ok': True})


@app.route('/api/analytics/summary_yoy')
def analytics_summary_yoy():
    """Returns analytics indexed by FY for YoY report."""
    db = get_db()
    all_fys = [r['fy'] for r in db.execute("SELECT DISTINCT fy FROM perf ORDER BY fy").fetchall()]
    result = {}
    for fy in all_fys:
        rows = R(db.execute("SELECT * FROM perf WHERE fy=?", (fy,)).fetchall())
        tot  = sum(r.get('total',1) for r in rows)
        comp = sum(r.get('compliant',0) for r in rows)
        nc   = tot-comp
        pct_c = round(comp/tot*100,2) if tot else 0
        pct_nc= round(nc/tot*100,2) if tot else 0
        # by_month
        by_month = {}
        for r in rows:
            m = r.get('bs_month','')
            by_month.setdefault(m,{'total':0,'compliant':0,'nc':0})
            by_month[m]['total']+=r.get('total',1)
            by_month[m]['compliant']+=r.get('compliant',0)
            by_month[m]['nc']+=r.get('non_compliant',0)
        # by_mp
        by_mp = {}
        for r in rows:
            ref = r.get('mp_ref','')
            if not ref: continue
            by_mp.setdefault(ref,{'total':0,'compliant':0,'nc':0})
            by_mp[ref]['total']+=r.get('total',1)
            by_mp[ref]['compliant']+=r.get('compliant',0)
            by_mp[ref]['nc']+=r.get('non_compliant',0)
        for ref in by_mp:
            t=by_mp[ref]['total']
            c=by_mp[ref]['compliant']
            by_mp[ref]['pct_c'] = round(c/t*100,2) if t else 0
        result[fy] = {'total':tot,'compliant':comp,'nc':nc,'pct_c':pct_c,'pct_nc':pct_nc,
                      'by_month':by_month,'by_mp':by_mp}
    return jsonify(result)


# ══════════════════════════════════════════════════════════
# EXPORT ROUTES
# ══════════════════════════════════════════════════════════

def _xl_sheet(cols, rows_data):
    """Build xlsx bytes with zero dependencies."""
    import io as _io, zipfile as _zf
    def ex(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
    ws = '<row r="1">'+''.join(f'<c r="{chr(65+i)}1" t="inlineStr"><is><t>{ex(col)}</t></is></c>' for i,col in enumerate(cols))+'</row>'
    for ri,row in enumerate(rows_data, 2):
        ws += f'<row r="{ri}">'+''.join(f'<c r="{chr(65+i)}{ri}" t="inlineStr"><is><t>{ex(v)}</t></is></c>' for i,v in enumerate(row))+'</row>'
    ws_xml = f'<?xml version="1.0" encoding="UTF-8"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>{ws}</sheetData></worksheet>'
    wb_xml = '<?xml version="1.0" encoding="UTF-8"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Data" sheetId="1" r:id="rId1"/></sheets></workbook>'
    rels   = '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>'
    ct     = '<?xml version="1.0" encoding="UTF-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>'
    rels_root = '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'
    buf = _io.BytesIO()
    with _zf.ZipFile(buf,'w',_zf.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', ct)
        z.writestr('_rels/.rels', rels_root)
        z.writestr('xl/_rels/workbook.xml.rels', rels)
        z.writestr('xl/workbook.xml', wb_xml)
        z.writestr('xl/worksheets/sheet1.xml', ws_xml)
    buf.seek(0); return buf.read()

def _xl_response(data, filename):
    return data, 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{filename}"'}

def _html_page(title, body, subtitle=''):
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><style>
body{{font-family:Arial,sans-serif;margin:0;padding:24px;color:#0f1a2e;font-size:12px}}
h1{{font-size:16px;font-weight:700;color:#0a1628;margin:0 0 2px}}
.sub{{font-size:11px;color:#6b7a99;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse;margin-bottom:24px}}
thead th{{background:#0a1628;color:#fff;padding:7px 10px;text-align:left;font-size:11px;font-weight:700}}
tbody tr:nth-child(even){{background:#f7f8fb}}
td{{padding:6px 10px;border-bottom:1px solid #e8eaf0;font-size:11px;vertical-align:top}}
.ref{{font-family:monospace;color:#1d4ed8;font-weight:700}}
.badge{{display:inline-block;padding:1px 6px;border-radius:4px;font-size:10px;font-weight:700}}
.l1{{background:#e0e7ff;color:#3730a3}}.l2{{background:#dcfce7;color:#166534}}.l3{{background:#fef9c3;color:#854d0e}}
.role-row{{background:#eef2ff!important;font-weight:700}}
.mp-row{{background:#f0fdf4!important}}
.no-print{{margin-bottom:12px}}
@media print{{.no-print{{display:none}}@page{{margin:12mm}}.page-break{{page-break-before:always}}}}
.tab-bar{{display:flex;gap:4px;margin-bottom:20px;border-bottom:2px solid #EEEEEE;padding-bottom:0}}
.tab-btn{{font-family:'Montserrat',sans-serif;font-size:11px;font-weight:700;padding:8px 18px;border:none;background:none;cursor:pointer;color:#777;border-bottom:3px solid transparent;margin-bottom:-2px;text-transform:uppercase;letter-spacing:.3px}}
.tab-btn.active{{color:#ED1C24;border-bottom-color:#ED1C24}}
.tab-btn:hover{{color:#ED1C24}}
.tab-panel{{display:none}}.tab-panel.active{{display:block}}
.audit-table td{{font-size:11px;padding:8px 12px}}
.audit-action{{font-family:'Montserrat',sans-serif;font-weight:700;font-size:10px;padding:2px 7px;border-radius:3px;background:#F0FDF4;color:#166634}}
.audit-action.login{{background:#EFF6FF;color:#1D4ED8}}
.audit-action.delete,.audit-action.disable{{background:#FFF0F0;color:#ED1C24}}
.audit-action.edit{{background:#FFFBEB;color:#92400E}}
</style></head><body>
<div class="no-print"><button onclick="window.print()" style="padding:6px 16px;background:#1d4ed8;color:#fff;border:none;border-radius:6px;cursor:pointer;margin-right:8px">🖨 Print / Save PDF</button><button onclick="window.close()" style="padding:6px 16px;background:#f1f5f9;color:#0f1a2e;border:1px solid #e2e8f0;border-radius:6px;cursor:pointer">Close</button></div>
<h1>{title}</h1><div class="sub">{subtitle}</div>
{body}
</body></html>"""

# ── ORG TREE HTML/PDF ──────────────────────────────────────────────────────
@app.route('/api/export/org_tree_html')
def export_org_tree_html():
    from datetime import datetime as _dt
    fy  = request.args.get('fy','')
    loc = request.args.get('loc','')
    db  = get_db()

    emps = {e['id']: dict(e) for e in db.execute("SELECT * FROM employees ORDER BY level,name")}
    mps_map = {}
    for row in db.execute("SELECT mp_id,emp_id FROM mp_owners"):
        mps_map.setdefault(row['emp_id'],[]).append(row['mp_id'])
    cps_map = {}
    for row in db.execute("SELECT cp_id,emp_id FROM cp_owners"):
        cps_map.setdefault(row['emp_id'],[]).append(row['cp_id'])

    # Compliance by employee
    perf_q = "SELECT emp_code,COUNT(*) tot,SUM(CASE WHEN status='C' THEN 1 ELSE 0 END) comp FROM perf WHERE 1=1"
    p_params = []
    if fy: perf_q += " AND fy=?"; p_params.append(fy)
    perf_q += " GROUP BY emp_code"
    perf_by_code = {r['emp_code']: r for r in db.execute(perf_q, p_params)}

    rows = []
    for lvl in [1,2,3]:
        for e in [v for v in emps.values() if v.get('level')==lvl]:
            pr = perf_by_code.get(e.get('emp_code',''), {})
            tot = pr.get('tot',0); comp = pr.get('comp',0)
            pct = f"{round(comp/tot*100,1)}%" if tot else "—"
            mgr = emps.get(e.get('manager_id',''),{}).get('name','—')
            rows.append([
                f"L{lvl}", e.get('emp_code',''), e.get('name',''), e.get('role','') or e.get('dept',''),
                mgr, len(mps_map.get(e['id'],[])), len(cps_map.get(e['id'],[])),
                f"{comp}/{tot}" if tot else "—", pct
            ])

    cols = ['Level','Emp Code','Name','Role/Dept','Manager','MPs','CPs','Comp/Total','Score%']
    thead = ''.join(f'<th>{h}</th>' for h in cols)
    tbody = ''
    cur_lvl = None
    for row in rows:
        if row[0] != cur_lvl:
            cur_lvl = row[0]
            lbl = {'L1':'Head of Department','L2':'Team Lead / Manager','L3':'Operations Staff'}.get(cur_lvl,'')
            tbody += f'<tr class="role-row"><td colspan="9" style="padding:10px;font-size:12px;color:#1e40af">▸ {cur_lvl} — {lbl}</td></tr>'
        score = row[8]
        color = '#16a34a' if score not in ('—','') and float(score.replace('%','') or 0)>=95 else ('#d97706' if score not in ('—','') and float(score.replace('%','') or 0)>=80 else '#dc2626') if score not in ('—','') else '#6b7a99'
        tbody += f'<tr><td>{row[0]}</td><td class="ref">{row[1]}</td><td><b>{row[2]}</b></td><td style="color:#475569">{row[3]}</td><td style="color:#6b7a99">{row[4]}</td><td style="text-align:center">{row[5]}</td><td style="text-align:center">{row[6]}</td><td style="text-align:center">{row[7]}</td><td style="font-weight:700;color:{color}">{score}</td></tr>'

    body = f'<table><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table>'
    fy_label = f"FY {fy}" if fy else "All FY"
    subtitle = f"{fy_label} · {loc or 'All Locations'} · Generated {_dt.now().strftime('%Y-%m-%d %H:%M')}"
    return _html_page("Organisation Tree — Compliance Summary", body, subtitle), 200, {'Content-Type':'text/html;charset=utf-8'}


# ── SINGLE EMPLOYEE MPCP EXCEL ─────────────────────────────────────────────
@app.route('/api/export/employee_mpcp_excel/<eid>')
def export_employee_mpcp_excel(eid):
    from datetime import datetime as _dt
    db  = get_db()
    emp = db.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    if not emp: return jsonify({'error':'Not found'}), 404
    emp = dict(emp)
    mps = db.execute("SELECT m.* FROM mps m JOIN mp_owners o ON o.mp_id=m.id WHERE o.emp_id=? ORDER BY m.ref", (eid,)).fetchall()
    cols = ['Role/Objective','S.N.','Managing Point','UoM','MP-Ref','Target','Freq','CP S.N.','Checking Point','CP-Ref','CP Target','CP Freq']
    rows = []
    for mp in mps:
        mp = dict(mp)
        cps = db.execute("SELECT * FROM cps WHERE mp_id=? ORDER BY ref", (mp['id'],)).fetchall()
        if not cps:
            rows.append([emp.get('role',''), mp['ref'], mp['title'], '', mp['ref'], mp.get('target',''), mp.get('freq',''), '', '', '', '', ''])
        for j,cp in enumerate(cps):
            cp = dict(cp)
            rows.append([
                emp.get('role','') if j==0 else '',
                mp['ref'] if j==0 else '',
                mp['title'] if j==0 else '',
                '', mp['ref'] if j==0 else '',
                mp.get('target','') if j==0 else '',
                mp.get('freq','') if j==0 else '',
                cp['ref'], cp['title'], cp['ref'],
                cp.get('target',''), cp.get('freq','')
            ])
    data = _xl_sheet(cols, rows)
    fname = f"MPCP_{emp.get('emp_code','EMP')}_{emp['name'].replace(' ','_')}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return _xl_response(data, fname)


# ── FULL TEAM MPCP BOOK (Excel — one row per CP, hierarchical) ─────────────
@app.route('/api/export/team_mpcp_excel')
def export_team_mpcp_excel():
    from datetime import datetime as _dt
    fy  = request.args.get('fy','')
    db  = get_db()
    emps = db.execute("SELECT * FROM employees ORDER BY level,name").fetchall()
    cols = ['Level','Emp Code','Employee','Dept','Role/Obj','MP Ref','Managing Point','MP Target','MP Freq','CP Ref','Checking Point','CP Target','CP Freq','FY','Compliance%']
    rows = []
    for emp in emps:
        emp = dict(emp)
        mps = db.execute("SELECT m.* FROM mps m JOIN mp_owners o ON o.mp_id=m.id WHERE o.emp_id=? ORDER BY m.ref",(emp['id'],)).fetchall()
        if not mps:
            rows.append([f"L{emp.get('level',3)}", emp.get('emp_code',''), emp['name'], emp.get('dept',''), emp.get('role',''), '','','','','','','','','',''])
            continue
        for mp in mps:
            mp = dict(mp)
            cps = db.execute("SELECT * FROM cps WHERE mp_id=? ORDER BY ref",(mp['id'],)).fetchall()
            perf_q = "SELECT COUNT(*) tot,SUM(CASE WHEN status='C' THEN 1 ELSE 0 END) comp FROM perf WHERE mp_ref=? AND emp_code=?"
            p_params = [mp['ref'], emp.get('emp_code','')]
            if fy: perf_q += " AND fy=?"; p_params.append(fy)
            pr = dict(db.execute(perf_q, p_params).fetchone() or {})
            tot = pr.get('tot',0) or 0; comp = pr.get('comp',0) or 0
            pct = f"{round(comp/tot*100,1)}%" if tot else ""
            if not cps:
                rows.append([f"L{emp.get('level',3)}", emp.get('emp_code',''), emp['name'], emp.get('dept',''), emp.get('role',''), mp['ref'], mp['title'], mp.get('target',''), mp.get('freq',''), '','','','', fy, pct])
                continue
            for j,cp in enumerate(cps):
                cp = dict(cp)
                rows.append([
                    f"L{emp.get('level',3)}" if j==0 else '',
                    emp.get('emp_code','') if j==0 else '',
                    emp['name'] if j==0 else '',
                    emp.get('dept','') if j==0 else '',
                    emp.get('role','') if j==0 else '',
                    mp['ref'] if j==0 else '',
                    mp['title'] if j==0 else '',
                    mp.get('target','') if j==0 else '',
                    mp.get('freq','') if j==0 else '',
                    cp['ref'], cp['title'], cp.get('target',''), cp.get('freq',''),
                    fy if j==0 else '', pct if j==0 else ''
                ])
    data = _xl_sheet(cols, rows)
    fname = f"Team_MPCP_Book_{fy or 'All'}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return _xl_response(data, fname)


# ── SECTOR SUMMARY EXCEL ───────────────────────────────────────────────────
@app.route('/api/export/sector_summary_excel')
def export_sector_summary_excel():
    from datetime import datetime as _dt
    fy  = request.args.get('fy','')
    db  = get_db()
    sectors = db.execute("SELECT * FROM sectors ORDER BY name").fetchall()
    cols = ['Sector','Employee','Level','Emp Code','MPs','CPs','Compliance%']
    rows = []
    for sec in sectors:
        sec = dict(sec)
        emps = db.execute("""SELECT e.* FROM employees e
            JOIN emp_sectors es ON es.emp_id=e.id WHERE es.sector_id=? ORDER BY e.level,e.name""", (sec['id'],)).fetchall()
        if not emps:
            rows.append([sec['name'],'—','','','','','']); continue
        for j,emp in enumerate(emps):
            emp = dict(emp)
            mp_cnt = db.execute("SELECT COUNT(*) c FROM mp_owners WHERE emp_id=?",(emp['id'],)).fetchone()['c']
            cp_cnt = db.execute("SELECT COUNT(*) c FROM cp_owners WHERE emp_id=?",(emp['id'],)).fetchone()['c']
            pq = "SELECT COUNT(*) tot,SUM(CASE WHEN status='C' THEN 1 ELSE 0 END) comp FROM perf WHERE emp_code=?"
            pp = [emp.get('emp_code','')]
            if fy: pq += " AND fy=?"; pp.append(fy)
            pr = dict(db.execute(pq,pp).fetchone() or {})
            tot = pr.get('tot',0) or 0; comp = pr.get('comp',0) or 0
            rows.append([sec['name'] if j==0 else '', emp['name'], f"L{emp.get('level',3)}", emp.get('emp_code',''), mp_cnt, cp_cnt, f"{round(comp/tot*100,1)}%" if tot else "—"])
    data = _xl_sheet(cols, rows)
    fname = f"Sector_Summary_{fy or 'All'}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return _xl_response(data, fname)


@app.route('/api/export/employee_mpcp_html/<eid>')
def export_employee_mpcp_html(eid):
    import datetime as _dt
    db  = get_db()
    emp = db.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    if not emp: return jsonify({'error':'Not found'}), 404
    emp = dict(emp)
    mps = db.execute("SELECT m.* FROM mps m JOIN mp_owners o ON o.mp_id=m.id WHERE o.emp_id=? ORDER BY m.ref", (eid,)).fetchall()

    rows_html = ''
    for mp in mps:
        mp = dict(mp)
        cps = db.execute("SELECT * FROM cps WHERE mp_id=? ORDER BY ref", (mp['id'],)).fetchall()
        cp_rows = ''.join(
            f'<tr><td style="padding:3px 6px;border:1px solid #e5e7eb;font-size:9px;color:#1d4ed8;font-family:monospace;white-space:nowrap">{dict(cp)["ref"]}</td>'
            f'<td style="padding:3px 6px;border:1px solid #e5e7eb;font-size:9px;color:#374151">{dict(cp)["title"]}</td>'
            f'<td style="padding:3px 6px;border:1px solid #e5e7eb;font-size:9px;text-align:center;white-space:nowrap">{dict(cp).get("target","—")}</td>'
            f'<td style="padding:3px 6px;border:1px solid #e5e7eb;font-size:9px;text-align:center">{dict(cp).get("freq","Monthly")}</td></tr>'
            for cp in cps)
        rows_html += (
            f'<tr style="background:#eef2ff;page-break-inside:avoid">'
            f'<td colspan="2" style="padding:4px 6px;border:1px solid #c7d2fe;font-weight:700;font-size:9.5px">'
            f'<span style="background:#0a1628;color:#fff;padding:1px 6px;border-radius:4px;font-size:8px;margin-right:6px;font-family:monospace">{mp["ref"]}</span>{mp["title"]}</td>'
            f'<td style="padding:4px 6px;border:1px solid #c7d2fe;font-size:9px;text-align:center;white-space:nowrap">{mp.get("target","—")}</td>'
            f'<td style="padding:4px 6px;border:1px solid #c7d2fe;font-size:9px;text-align:center">{mp.get("freq","Monthly")}</td></tr>{cp_rows}')

    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,sans-serif;color:#0f1a2e;font-size:10px;padding:14px;width:190mm}}
.header{{display:flex;align-items:center;gap:10px;margin-bottom:10px;padding-bottom:8px;border-bottom:2px solid #0a1628}}
.avatar{{width:36px;height:36px;border-radius:50%;background:#0a1628;color:#fff;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px;flex-shrink:0}}
h1{{font-size:14px;font-weight:700;color:#0a1628;margin-bottom:2px}}
.meta{{font-size:9px;color:#6b7a99}}
.badge{{display:inline-block;padding:1px 6px;border-radius:4px;font-size:8px;font-weight:700;margin-right:3px}}
.b-gray{{background:#f1f5f9;color:#475569}}
.b-blue{{background:#eff6ff;color:#1d4ed8}}
table{{width:100%;border-collapse:collapse;margin-top:8px;font-size:9px}}
thead th{{background:#0a1628;color:#fff;padding:5px 6px;text-align:left;font-size:9px;font-weight:700}}
thead th:nth-child(3),thead th:nth-child(4){{width:90px;text-align:center}}
.footer{{margin-top:8px;font-size:8px;color:#9ca3af;text-align:right;border-top:1px solid #e5e7eb;padding-top:4px}}
.no-print{{margin-bottom:8px}}
@page{{size:A4 portrait;margin:10mm}}
@media print{{
  .no-print{{display:none!important}}
  body{{padding:0;width:190mm}}
  *{{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
  tr{{page-break-inside:avoid}}
}}
.tab-bar{{display:flex;gap:4px;margin-bottom:20px;border-bottom:2px solid #EEEEEE;padding-bottom:0}}
.tab-btn{{font-family:'Montserrat',sans-serif;font-size:11px;font-weight:700;padding:8px 18px;border:none;background:none;cursor:pointer;color:#777;border-bottom:3px solid transparent;margin-bottom:-2px;text-transform:uppercase;letter-spacing:.3px}}
.tab-btn.active{{color:#ED1C24;border-bottom-color:#ED1C24}}
.tab-btn:hover{{color:#ED1C24}}
.tab-panel{{display:none}}.tab-panel.active{{display:block}}
.audit-table td{{font-size:11px;padding:8px 12px}}
.audit-action{{font-family:'Montserrat',sans-serif;font-weight:700;font-size:10px;padding:2px 7px;border-radius:3px;background:#F0FDF4;color:#166534}}
.audit-action.login{{background:#EFF6FF;color:#1D4ED8}}
.audit-action.delete,.audit-action.disable{{background:#FFF0F0;color:#ED1C24}}
.audit-action.edit{{background:#FFFBEB;color:#92400E}}
</style></head><body>
<div class="no-print">
  <button onclick="window.print()" style="padding:5px 14px;background:#1d4ed8;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:11px">&#128438; Print / Save PDF (A4)</button>
</div>
<div class="header">
  <div class="avatar">{"".join(w[0] for w in emp["name"].split()[:2]).upper()}</div>
  <div style="flex:1">
    <h1>{emp["name"]}</h1>
    <div class="meta" style="margin-bottom:4px">{emp.get("role","") or emp.get("dept","")}</div>
    <div>
      <span class="badge b-gray">{emp.get("emp_code","—")}</span>
      <span class="badge b-gray">Level {emp.get("level","")}</span>
      <span class="badge b-blue">{len(mps)} MPs &middot; {sum(1 for mp in mps for _ in db.execute("SELECT id FROM cps WHERE mp_id=?", (dict(mp)["id"],)).fetchall())} CPs</span>
    </div>
  </div>
  <div style="text-align:right;font-size:8px;color:#9ca3af">MPCP Framework<br>Sipradi Trading Pvt. Ltd.<br>{_dt.datetime.now().strftime("%d %b %Y")}</div>
</div>
<table>
  <thead><tr><th>Ref</th><th>Managing Point / Checking Point</th><th style="text-align:center">Target / SLA</th><th style="text-align:center">Frequency</th></tr></thead>
  <tbody>{rows_html or '<tr><td colspan="4" style="padding:16px;text-align:center;color:#9ca3af;font-style:italic">No MPs assigned</td></tr>'}</tbody>
</table>
<div class="footer">Confidential · MPCP System · Sipradi Trading Pvt. Ltd. · Generated {_dt.datetime.now().strftime("%d %b %Y %H:%M")}</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}





# ══════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════════

CHANGE_PW_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Change Password &mdash; MPCP</title>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@500;600;700;800&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Montserrat',sans-serif;background:#F4F4F4;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center}
.wrapper{width:400px;max-width:95vw}
.brand{text-align:center;margin-bottom:24px}
.brand h1{font-size:16px;font-weight:800;color:#1A1A1A;margin-bottom:4px}
.brand p{font-size:11px;color:#999}
.card{background:#fff;border-radius:4px;border:1px solid #DDDDDD;padding:28px;box-shadow:0 2px 8px rgba(0,0,0,.08)}
.fg{margin-bottom:16px}
.fg label{display:block;font-size:10px;font-weight:700;color:#666;margin-bottom:6px;text-transform:uppercase;letter-spacing:.5px}
.fg input{width:100%;padding:10px 12px;border:1.5px solid #DDDDDD;border-radius:3px;font-size:13px;font-family:'Montserrat',sans-serif}
.fg input:focus{border-color:#ED1C24;outline:none}
.btn{width:100%;padding:12px;background:#ED1C24;color:#fff;border:none;border-radius:3px;font-family:'Montserrat',sans-serif;font-size:12px;font-weight:700;cursor:pointer;text-transform:uppercase;letter-spacing:.5px;margin-top:4px}
.btn:hover{background:#C1121F}
.err{background:#FFF0F0;border-left:4px solid #ED1C24;color:#C1121F;padding:10px 14px;border-radius:3px;font-size:12px;margin-bottom:16px}
.ok{background:#F0FDF4;border-left:4px solid #16A34A;color:#166534;padding:10px 14px;border-radius:3px;font-size:12px;margin-bottom:16px}
.foot{text-align:center;margin-top:20px;font-size:10px;color:#999}
.back{display:block;text-align:center;margin-top:14px;font-size:11px;color:#ED1C24;text-decoration:none;font-weight:600}
</style></head><body>
<div class="wrapper">
  <div class="brand">
    <h1>&#128274; Change Password</h1>
    <p>MPCP Management System</p>
  </div>
  <div class="card">
    {% if error %}<div class="err">&#9888;&nbsp; {{ error }}</div>{% endif %}
    {% if success %}<div class="ok">&#10003;&nbsp; {{ success }}</div>{% endif %}
    {% if not success %}
    <form method="POST" action="/change_password">
      <div class="fg">
        <label>Username</label>
        <input type="text" name="username" placeholder="Your username" required>
      </div>
      <div class="fg">
        <label>Current Password</label>
        <input type="password" name="current_password" placeholder="Current password" required>
      </div>
      <div class="fg">
        <label>New Password</label>
        <input type="password" name="new_password" placeholder="Min 6 characters" required>
      </div>
      <div class="fg">
        <label>Confirm New Password</label>
        <input type="password" name="confirm_password" placeholder="Repeat new password" required>
      </div>
      <button class="btn" type="submit">Change Password &rarr;</button>
    </form>
    {% endif %}
  </div>
  <a href="/login" class="back">&#8592; Back to Login</a>
  <div class="foot">&copy; Govinda Upadhyay &mdash; MPCP Management V 3.0</div>
</div>
</body></html>"""

@app.route('/change_password', methods=['GET','POST'])
def change_password_page():
    error = None
    success = None
    if request.method == 'POST':
        username = request.form.get('username','').strip().lower()
        cur_pw   = request.form.get('current_password','')
        new_pw   = request.form.get('new_password','')
        conf_pw  = request.form.get('confirm_password','')
        if not username or not cur_pw or not new_pw or not conf_pw:
            error = 'All fields are required.'
        elif len(new_pw) < 6:
            error = 'New password must be at least 6 characters.'
        elif new_pw != conf_pw:
            error = 'New passwords do not match.'
        else:
            db = get_master_conn()
            user = db.execute("SELECT * FROM users WHERE username=? AND active=1",(username,)).fetchone()
            if not user or not verify_password(cur_pw, user['password_hash']):
                error = 'Invalid username or current password.'
            else:
                db.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (hash_password(new_pw), user['id']))
                db.commit()
                log_audit('PASSWORD_CHANGE','user',user['id'],'Self-service password change')
                success = 'Password changed successfully! You can now login with your new password.'
            db.close()
    return render_template_string(CHANGE_PW_HTML, error=error, success=success)


LOGIN_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MPCP Management System — Sipradi Trading</title>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;800&family=Roboto:wght@400;500&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Roboto',sans-serif;background:#F4F4F4;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center}
.top-bar{display:none}

.wrapper{width:420px;margin-top:0}
.brand{text-align:center;margin-bottom:28px}
.brand-logo{width:64px;height:64px;background:#ED1C24;border-radius:8px;display:flex;align-items:center;justify-content:center;margin:0 auto 14px;font-size:28px}
.brand h1{font-family:'Montserrat',sans-serif;font-size:22px;font-weight:800;color:#2E2E2E;letter-spacing:-.5px}
.brand p{font-size:12px;color:#777;margin-top:4px;font-weight:500}
.card{background:#fff;border-radius:6px;padding:36px;box-shadow:0 4px 20px rgba(0,0,0,.10);border-top:4px solid #ED1C24}
.fg{margin-bottom:18px}
label{display:block;font-family:'Montserrat',sans-serif;font-size:10px;font-weight:700;color:#777;margin-bottom:6px;text-transform:uppercase;letter-spacing:.7px}
input{width:100%;padding:11px 14px;border:1.5px solid #DDDDDD;border-radius:4px;font-size:14px;font-family:'Roboto',sans-serif;color:#2E2E2E;outline:none;transition:.2s}
input:focus{border-color:#ED1C24;box-shadow:0 0 0 3px rgba(215,25,32,.08)}
.btn{width:100%;padding:13px;background:#ED1C24;color:#fff;border:none;border-radius:4px;font-size:13px;font-family:'Montserrat',sans-serif;font-weight:700;cursor:pointer;margin-top:4px;transition:.2s;letter-spacing:.5px;text-transform:uppercase}
.btn:hover{background:#C1121F}
.err{background:#FFF5F5;border:1px solid #FECACA;color:#ED1C24;padding:10px 14px;border-radius:4px;font-size:12px;margin-bottom:16px;font-weight:500}
.foot{text-align:center;margin-top:20px;font-size:10px;color:#aaa;font-family:'Montserrat',sans-serif;letter-spacing:.3px}
.divider{border:none;border-top:1px solid #DDDDDD;margin:16px 0}
.tab-bar{display:flex;gap:4px;margin-bottom:20px;border-bottom:2px solid #EEEEEE;padding-bottom:0}
.tab-btn{font-family:'Montserrat',sans-serif;font-size:11px;font-weight:700;padding:8px 18px;border:none;background:none;cursor:pointer;color:#777;border-bottom:3px solid transparent;margin-bottom:-2px;text-transform:uppercase;letter-spacing:.3px}
.tab-btn.active{color:#ED1C24;border-bottom-color:#ED1C24}
.tab-btn:hover{color:#ED1C24}
.tab-panel{display:none}.tab-panel.active{display:block}
.audit-table td{font-size:11px;padding:8px 12px}
.audit-action{font-family:'Montserrat',sans-serif;font-weight:700;font-size:10px;padding:2px 7px;border-radius:3px;background:#F0FDF4;color:#166534}
.audit-action.login{background:#EFF6FF;color:#1D4ED8}
.audit-action.delete,.audit-action.disable{background:#FFF0F0;color:#ED1C24}
.audit-action.edit{background:#FFFBEB;color:#92400E}
</style></head><body>

<div class="wrapper">
  <div class="brand">
    <div class="brand-logo">&#128200;</div>
    <h1>MPCP Management System</h1>
    <p>Manpower Compliance & Performance Control</p>
  </div>
  <div class="card">
    {% if error %}<div class="err">&#9888;&nbsp; {{ error }}</div>{% endif %}
    <form method="POST" action="/login">
      <div class="fg">
        <label>Username</label>
        <input type="text" name="username" placeholder="Enter your username" autofocus required value="{{ prefill_username }}">
      </div>
      <div class="fg">
        <label>Password</label>
        <input type="password" name="password" placeholder="Enter your password" required>
      </div>
      <button class="btn" type="submit">Sign In &rarr;</button>
    </form>
    <div style="text-align:center;margin-top:12px">
      <a href="/change_password" style="font-size:11px;color:#ED1C24;text-decoration:none;font-family:'Montserrat',sans-serif;font-weight:600">🔒 Change Password</a>
    </div>
  </div>
  <div class="foot">&copy; Govinda Upadhyay &mdash; MPCP Management V 3.0</div>
</div>
</body></html>"""
@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        db = get_master_conn()
        user = db.execute(
            "SELECT u.*,d.name dept_name FROM users u "
            "LEFT JOIN departments d ON d.code=u.dept_code "
            "WHERE u.username=? AND u.active=1", (username,)
        ).fetchone()
        db.close()
        if user and verify_password(password, user['password_hash']):
            session.permanent = True
            session['mpcp_user'] = {
                'id':        user['id'],
                'username':  user['username'],
                'full_name': user['full_name'],
                'role':      user['role'],
                'dept_code': user['dept_code'],
                'dept_name': dict(user).get('dept_name') or 'All Departments',
                'emp_code':  user['emp_code'] or '',
                'active_dept': user['dept_code']
            }
            log_audit('LOGIN','user',user['id'],'Login: '+user['username']+' role='+user['role'])
            return redirect('/')
        error = 'Invalid username or password. Please try again.'
    import datetime as _dt
    return render_template_string(LOGIN_HTML,
        error=error,
        prefill_username='',
        year=_dt.datetime.now().year)

@app.route('/logout')
def logout():
    log_audit("LOGOUT","user","","User logged out")
    session.clear()
    return redirect('/login')

@app.route('/api/auth/me')
def auth_me():
    u = current_user()
    if not u: return jsonify({'error':'Not authenticated'}),401
    return jsonify({k:v for k,v in u.items() if k!='password_hash'})

@app.route('/api/auth/switch_dept', methods=['POST'])
def switch_dept():
    err = require_role('master_admin')
    if err: return err
    dept = (request.json or {}).get('dept_code','')
    if dept:
        db = get_master_conn()
        exists = db.execute("SELECT id FROM departments WHERE code=?", (dept,)).fetchone()
        db.close()
        if not exists: return json_error('Department not found')
    u = session['mpcp_user'].copy()
    u['active_dept'] = dept or None
    session['mpcp_user'] = u
    return jsonify({'ok':True,'active_dept':dept})

# ── DEPARTMENT MANAGEMENT (master_admin only) ──────────────────────────────
def current_user():
    return session.get('mpcp_user')

def require_role(*roles):
    u = current_user()
    if not u: return jsonify({'error':'Not authenticated'}),401
    if roles and u.get('role') not in roles:
        return jsonify({'error':'Insufficient permissions'}),403
    return None

def perf_emp_filter(query, params):
    u = current_user()
    if u and u.get('role') == 'user' and u.get('emp_code'):
        query += ' AND emp_code=?'
        params.append(u['emp_code'])
    return query, params

def get_dept_db_path(dept_code):
    return os.path.join(DATA_DIR, f'{dept_code}.db')

def get_master_conn():
    conn = sqlite3.connect(MASTER_DB)
    conn.row_factory = sqlite3.Row
    return conn

def _init_dept_db(path):
    conn = sqlite3.connect(path)
    conn.executescript(SCHEMA)
    conn.row_factory = sqlite3.Row
    _ensure_mpcp_status_columns(conn)
    conn.commit(); conn.close()


@app.route('/api/departments', methods=['GET','POST'])
def departments_api():
    err = require_role('master_admin')
    if err: return err
    db = get_master_conn()
    if request.method == 'GET':
        rows = R(db.execute("SELECT * FROM departments ORDER BY name").fetchall())
        # Attach user count per dept
        for r in rows:
            r['user_count'] = db.execute(
                "SELECT COUNT(*) c FROM users WHERE dept_code=? AND active=1",
                (r['code'],)).fetchone()['c']
        db.close(); return jsonify(rows)
    d = request.json or {}
    if not d.get('code') or not d.get('name'):
        return json_error('code and name required')
    did = uid()
    log_audit("DEPT_CREATE","department","","Created department")
    db.execute("INSERT INTO departments VALUES(?,?,?,?,?)",
        (did, d['code'].lower().replace(' ','_'), d['name'], 1,
         datetime.datetime.now().isoformat()))
    # Auto-create the dept DB
    _init_dept_db(get_dept_db_path(d['code'].lower().replace(' ','_')))
    db.commit(); db.close()
    return jsonify({'id':did})

@app.route('/api/departments/<did>', methods=['PUT','DELETE'])
def department_api(did):
    err = require_role('master_admin')
    if err: return err
    db = get_master_conn()
    if request.method == 'DELETE':
        row = db.execute("SELECT code FROM departments WHERE id=?", (did,)).fetchone()
        if not row:
            db.close(); return jsonify({"ok":False, "error":"Not found"}), 404
        code = row["code"]
        db.execute("DELETE FROM departments WHERE id=?", (did,))
        db.execute("DELETE FROM users WHERE dept_code=?", (code,))
        db.commit(); db.close()
        # Remove the dept DB file
        dept_db = os.path.join(DATA_DIR, code + '.db')
        if os.path.exists(dept_db):
            os.remove(dept_db)
        log_audit("DEPT_DELETE", "department", did, f"Deleted department {code}")
        return jsonify({"ok":True})
    d = request.json or {}
    log_audit("DEPT_UPDATE","department",did,"Updated department")
    db.execute("UPDATE departments SET name=?,active=? WHERE id=?",
        (d.get('name'), 1 if d.get('active',True) else 0, did))
    db.commit(); db.close(); return jsonify({'ok':True})

# ── USER MANAGEMENT (master_admin + dept_admin) ────────────────────────────
@app.route('/api/users', methods=['GET','POST'])
def users_api():
    err = require_role('master_admin','dept_admin')
    if err: return err
    u = current_user()
    db = get_master_conn()
    if request.method == 'GET':
        if u['role'] == 'master_admin':
            rows = R(db.execute(
                "SELECT u.id,u.username,u.full_name,u.role,u.dept_code,"
                "u.emp_code,u.active,u.created_at,d.name dept_name "
                "FROM users u LEFT JOIN departments d ON d.code=u.dept_code "
                "ORDER BY u.dept_code,u.full_name").fetchall())
        else:
            rows = R(db.execute(
                "SELECT u.id,u.username,u.full_name,u.role,u.dept_code,"
                "u.emp_code,u.active,u.created_at,d.name dept_name "
                "FROM users u LEFT JOIN departments d ON d.code=u.dept_code "
                "WHERE u.dept_code=? ORDER BY u.full_name",
                (u['dept_code'],)).fetchall())
        db.close(); return jsonify(rows)

    d = request.json or {}
    missing = validate_required(d,'username','password','full_name','role')
    if missing: return json_error(f"Missing: {', '.join(missing)}")
    # dept_admin can only create users in own dept
    dept_code = d.get('dept_code') or u.get('dept_code')
    if u['role']=='dept_admin': dept_code = u['dept_code']
    if d['role'] not in ('master_admin','dept_admin','moderator','user'):
        return json_error('Invalid role')
    new_id = uid()
    try:
        db.execute("INSERT INTO users VALUES(?,?,?,?,?,?,?,?,?)",
            (new_id, d['username'].strip().lower(),
             hash_password(d['password']),
             d['full_name'], d['role'],
             dept_code, d.get('emp_code',''), 1,
             datetime.datetime.now().isoformat()))
        log_audit('USER_CREATE','user',new_id,'API created: '+d['username'])
        db.commit()
    except sqlite3.IntegrityError:
        return json_error('Username already exists', 409)
    finally: db.close()
    return jsonify({'id':new_id})

@app.route('/api/users/<uid2>', methods=['PUT','DELETE'])
def user_api(uid2):
    err = require_role('master_admin','dept_admin')
    if err: return err
    u = current_user()
    db = get_master_conn()
    target = db.execute("SELECT * FROM users WHERE id=?", (uid2,)).fetchone()
    if not target: db.close(); return json_error('User not found',404)
    # dept_admin can only manage own dept users
    if u['role']=='dept_admin' and dict(target).get('dept_code')!=u['dept_code']:
        db.close(); return json_error('Access denied',403)
    if request.method == 'DELETE':
        db.execute("UPDATE users SET active=0 WHERE id=?", (uid2,))
        db.commit(); db.close(); return jsonify({'ok':True})
    d = request.json or {}
    pw_hash = dict(target)['password_hash']
    if d.get('password'): pw_hash = hash_password(d['password'])
    db.execute(
        "UPDATE users SET full_name=?,role=?,dept_code=?,emp_code=?,active=?,password_hash=? WHERE id=?",
        (d.get('full_name', target['full_name']),
         d.get('role', target['role']),
         d.get('dept_code', target['dept_code']),
         d.get('emp_code', target['emp_code']),
         1 if d.get('active',True) else 0,
         pw_hash, uid2))
    db.commit(); db.close()
    return jsonify({'ok':True})

@app.route('/api/master/summary')
def master_summary():
    err = require_role('master_admin')
    if err: return err
    db_m = get_master_conn()
    depts = R(db_m.execute("SELECT * FROM departments WHERE active=1").fetchall())
    db_m.close()
    summary = []
    for dept in depts:
        path = get_dept_db_path(dept['code'])
        if not os.path.exists(path):
            summary.append({**dept,'employees':0,'mps':0,'cps':0,'compliance':None})
            continue
        d = sqlite3.connect(path); d.row_factory=sqlite3.Row
        emps = d.execute("SELECT COUNT(*) c FROM employees").fetchone()['c']
        mps  = d.execute("SELECT COUNT(*) c FROM mps").fetchone()['c']
        cps  = d.execute("SELECT COUNT(*) c FROM cps").fetchone()['c']
        perf = d.execute(
            "SELECT COUNT(*) tot,SUM(CASE WHEN status='C' THEN 1 ELSE 0 END) comp FROM perf"
        ).fetchone()
        pct = round(perf['comp']/perf['tot']*100,1) if perf['tot'] else None
        d.close()
        summary.append({**dept,'employees':emps,'mps':mps,'cps':cps,'compliance':pct})
    return jsonify(summary)


# ── MASTERS DEFAULT SEED ───────────────────────────────────────────────────

# ── MASTERS API ─────────────────────────────────────────────────────────────
@app.route('/api/masters', methods=['GET'])
def masters_api_get():
    db = get_master_conn()
    rows = R(db.execute("SELECT * FROM masters WHERE active=1 ORDER BY category,sort_order,value").fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/masters/<category>', methods=['GET'])
def masters_by_category(category):
    if not current_user(): return jsonify({'error':'Not authenticated'}), 401
    db = get_master_conn()
    rows = R(db.execute("SELECT * FROM masters WHERE category=? AND active=1 ORDER BY sort_order,value",(category,)).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/masters', methods=['POST'])
def masters_api_create():
    err = require_role('master_admin')
    if err: return err
    d = request.json or {}
    if not d.get('category') or not d.get('value'):
        return json_error('category and value required')
    db = get_master_conn()
    mid = uid()
    try:
        db.execute("INSERT INTO masters(id,category,value,label,sort_order) VALUES(?,?,?,?,?)",
                  (mid,d['category'],d['value'],d.get('label',d['value']),d.get('sort_order',0)))
        db.commit()
    except Exception as e:
        db.close()
        return json_error('Value already exists in this category')
    db.close()
    log_audit('MASTER_CREATE','masters',mid,f"{d['category']}: {d['value']}")
    return jsonify({'ok':True,'id':mid})

@app.route('/api/masters/<mid>', methods=['PUT'])
def masters_api_update(mid):
    err = require_role('master_admin')
    if err: return err
    d = request.json or {}
    db = get_master_conn()
    db.execute("UPDATE masters SET value=?,label=?,sort_order=? WHERE id=?",
              (d.get('value'),d.get('label',''),d.get('sort_order',0),mid))
    db.commit(); db.close()
    log_audit('MASTER_UPDATE','masters',mid,f"Updated: {d.get('value')}")
    return jsonify({'ok':True})

@app.route('/api/masters/<mid>', methods=['DELETE'])
def masters_api_delete(mid):
    err = require_role('master_admin')
    if err: return err
    db = get_master_conn()
    db.execute("DELETE FROM masters WHERE id=?", (mid,))
    db.commit(); db.close()
    log_audit('MASTER_DELETE','masters',mid,'Deleted master item')
    return jsonify({'ok':True})

# ── ADMIN PANEL ────────────────────────────────────────────────────────────
ADMIN_HTML = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MPCP Admin Panel</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
:root{
  --red:#FF2D3A;--red-dark:#D91F2A;--red-light:rgba(255,45,58,.1);
  --blue:#2D9EFF;--blue-dark:#1A7FD4;--blue-light:rgba(45,158,255,.1);
  --green:#10DF8A;--amber:#FFB020;--violet:#8B5CF6;--pink:#FF4DC4;
  --bg:#0A0E1A;--bg-2:#0F1422;--bg-3:#141828;
  --surface:#1A1F30;--surface-2:#1F253A;--surface-3:#252B40;
  --border:rgba(255,255,255,.08);--border-light:rgba(255,255,255,.05);
  --border-glow:rgba(45,158,255,.25);
  --text:#F0F4FF;--text-2:#B8C4E0;--muted:#6B7A9F;
  --radius:8px;
  --shadow:0 2px 8px rgba(0,0,0,.4);
  --shadow-md:0 4px 24px rgba(0,0,0,.5);
  --font:'Inter',sans-serif;--font-mono:'JetBrains Mono',monospace;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--font);background:var(--bg);color:var(--text);font-size:13px;min-height:100vh;
  background-image:linear-gradient(rgba(45,158,255,.03) 1px,transparent 1px),linear-gradient(90deg,rgba(45,158,255,.03) 1px,transparent 1px);
  background-size:40px 40px;-webkit-font-smoothing:antialiased}
.topbar{background:linear-gradient(180deg,rgba(15,18,32,.98) 0%,rgba(10,14,26,.98) 100%);border-bottom:1px solid var(--border);padding:0 24px;height:54px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:200;backdrop-filter:blur(20px);box-shadow:0 4px 32px rgba(0,0,0,.6),0 1px 0 rgba(45,158,255,.15)}
.topbar-logo{display:flex;align-items:center;gap:10px}
.topbar-logo .icon{width:32px;height:32px;background:linear-gradient(135deg,var(--red),#FF8C42);border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:16px}
.topbar-logo .name{font-weight:800;font-size:15px;letter-spacing:-.3px;color:#fff}
.topbar-logo .name em{background:linear-gradient(135deg,#FF2D3A,#FF8C42);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;font-style:normal}
.topbar-logo .sub{font-size:10px;color:var(--muted);font-family:var(--font-mono);letter-spacing:.5px}
.topbar-actions{display:flex;gap:8px;align-items:center}
.tbtn{font-size:11px;font-weight:600;padding:6px 14px;border-radius:6px;border:1px solid var(--border);background:rgba(255,255,255,.04);color:var(--text-2);cursor:pointer;text-decoration:none;transition:.2s;display:inline-flex;align-items:center;gap:6px;font-family:var(--font)}
.tbtn:hover{border-color:var(--border-glow);color:var(--blue);background:var(--blue-light)}
.container{max-width:1280px;margin:28px auto;padding:0 24px}
.tab-bar{display:flex;gap:2px;margin-bottom:24px;background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:4px}
.tab-btn{flex:1;font-size:11px;font-weight:600;padding:9px 16px;border:none;background:none;cursor:pointer;color:var(--muted);border-radius:7px;transition:.2s;font-family:var(--font);display:flex;align-items:center;justify-content:center;gap:6px}
.tab-btn:hover{color:var(--text-2);background:rgba(255,255,255,.04)}
.tab-btn.active{background:linear-gradient(135deg,rgba(255,45,58,.15),rgba(255,140,66,.08));color:var(--red);border:1px solid rgba(255,45,58,.2)}
.tab-panel{display:none}.tab-panel.active{display:block}
.card{background:var(--surface);border:1px solid var(--border);border-radius:10px;margin-bottom:20px;overflow:hidden}
.card-head{padding:14px 20px;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center;background:linear-gradient(180deg,var(--surface-2),var(--surface))}
.card-head h2{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--text-2);display:flex;align-items:center;gap:8px}
.card-head h2 span{color:var(--muted);font-weight:400;text-transform:none;letter-spacing:0;font-size:11px}
table{width:100%;border-collapse:collapse}
thead th{padding:10px 14px;text-align:left;font-size:10px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;border-bottom:1px solid var(--border);background:var(--bg-3)}
tbody td{padding:11px 14px;border-bottom:1px solid var(--border-light);font-size:12px;vertical-align:middle;color:var(--text-2)}
tbody tr:hover{background:rgba(45,158,255,.03)}
tbody tr:last-child td{border-bottom:none}
.badge{display:inline-flex;align-items:center;padding:3px 8px;border-radius:20px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px;font-family:var(--font-mono)}
.b-master_admin{background:rgba(255,45,58,.12);color:#FF6B6B;border:1px solid rgba(255,45,58,.3)}
.b-dept_admin{background:rgba(45,158,255,.12);color:var(--blue);border:1px solid rgba(45,158,255,.3)}
.b-moderator{background:rgba(255,176,32,.12);color:var(--amber);border:1px solid rgba(255,176,32,.3)}
.b-user{background:rgba(16,223,138,.12);color:var(--green);border:1px solid rgba(16,223,138,.3)}
.status-dot{width:7px;height:7px;border-radius:50%;display:inline-block;margin-right:5px}
.status-active .status-dot{background:var(--green);box-shadow:0 0 6px var(--green)}
.status-inactive .status-dot{background:var(--muted)}
.btn{font-family:var(--font);padding:7px 14px;border:none;border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;transition:.15s;display:inline-flex;align-items:center;gap:5px}
.btn-primary{background:linear-gradient(135deg,var(--red),var(--red-dark));color:#fff;box-shadow:0 2px 8px rgba(255,45,58,.3)}
.btn-primary:hover{box-shadow:0 4px 16px rgba(255,45,58,.4);transform:translateY(-1px)}
.btn-ghost{background:rgba(255,255,255,.04);color:var(--text-2);border:1px solid var(--border)}
.btn-ghost:hover{border-color:var(--border-glow);color:var(--blue)}
.btn-danger{background:rgba(255,45,58,.1);color:#FF6B6B;border:1px solid rgba(255,45,58,.2)}
.btn-danger:hover{background:var(--red);color:#fff}
.btn-warn{background:rgba(255,176,32,.1);color:var(--amber);border:1px solid rgba(255,176,32,.2)}
.btn-success{background:rgba(16,223,138,.1);color:var(--green);border:1px solid rgba(16,223,138,.2)}
.btn-success:hover{background:rgba(16,223,138,.2)}
.btn-sm{padding:4px 10px;font-size:10px}
.form-row{display:flex;gap:12px;flex-wrap:wrap;padding:18px 20px;border-bottom:1px solid var(--border);align-items:flex-end}
.form-row .fg{flex:1;min-width:130px}
.fg label,.mfg label{display:block;font-size:10px;font-weight:700;color:var(--muted);margin-bottom:6px;text-transform:uppercase;letter-spacing:.5px}
.fg input,.fg select,.mfg input,.mfg select{width:100%;padding:9px 12px;border:1.5px solid var(--border);border-radius:6px;font-size:12px;font-family:var(--font);color:var(--text);background:var(--bg-3);outline:none;transition:.2s;-webkit-appearance:none}
.fg input:focus,.fg select:focus,.mfg input:focus,.mfg select:focus{border-color:var(--border-glow);box-shadow:0 0 0 3px rgba(45,158,255,.08)}
.fg select option,.mfg select option{background:var(--bg-2);color:var(--text)}
.search-box{padding:7px 12px 7px 32px;border:1.5px solid var(--border);border-radius:6px;font-size:12px;width:220px;font-family:var(--font);outline:none;background:var(--bg-3);color:var(--text);transition:.2s;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='%236B7A9F' stroke-width='2'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cpath d='M21 21l-4.35-4.35'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:10px center}
.search-box:focus{border-color:var(--border-glow)}
.pw-input-wrap{position:relative}
.pw-input-wrap input{padding-right:36px}
.pw-toggle{position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;color:var(--muted);font-size:14px;padding:2px;line-height:1}
.msg-ok,.msg-err{padding:12px 16px;border-radius:8px;font-size:12px;margin-bottom:20px;font-weight:500;display:flex;align-items:center;gap:10px}
.msg-ok{background:rgba(16,223,138,.08);border:1px solid rgba(16,223,138,.25);color:var(--green)}
.msg-err{background:rgba(255,45,58,.08);border:1px solid rgba(255,45,58,.25);color:#FF6B6B}
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:300;align-items:center;justify-content:center;backdrop-filter:blur(4px)}
.modal-overlay.open{display:flex}
.modal{background:var(--surface);border:1px solid var(--border-glow);border-radius:12px;width:520px;max-width:95vw;max-height:90vh;overflow:hidden;box-shadow:0 24px 80px rgba(0,0,0,.7),0 0 0 1px rgba(45,158,255,.1)}
.modal-head{padding:16px 20px;display:flex;justify-content:space-between;align-items:center;background:linear-gradient(135deg,rgba(255,45,58,.12),rgba(255,140,66,.06));border-bottom:1px solid var(--border)}
.modal-head h3{font-size:13px;font-weight:700;color:var(--text);display:flex;align-items:center;gap:8px}
.modal-body{padding:24px 20px;overflow-y:auto;max-height:60vh}
.modal-foot{padding:14px 20px;border-top:1px solid var(--border);display:flex;gap:8px;justify-content:flex-end;background:var(--bg-3)}
.mfg{margin-bottom:16px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.xbtn{background:none;border:none;cursor:pointer;color:var(--muted);font-size:18px;line-height:1;padding:2px;transition:.2s}
.xbtn:hover{color:var(--text)}
.masters-cats{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px}
.masters-cats .cat-btn{font-size:11px;font-weight:600;padding:7px 16px;border-radius:20px;border:1px solid var(--border);background:rgba(255,255,255,.03);color:var(--muted);cursor:pointer;transition:.2s;font-family:var(--font)}
.masters-cats .cat-btn:hover{color:var(--text-2);border-color:var(--border-glow)}
.masters-cats .cat-btn.active{background:var(--blue-light);color:var(--blue);border-color:rgba(45,158,255,.4)}
.masters-empty{text-align:center;padding:40px 20px;color:var(--muted);font-size:12px}
.stat-pills{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px}
.stat-pill{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;flex:1;min-width:130px}
.stat-pill .val{font-size:24px;font-weight:800;font-family:var(--font-mono)}
.stat-pill .lbl{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-top:3px}
.dept-pill{font-family:var(--font-mono);font-size:10px;font-weight:700;padding:3px 10px;border-radius:20px;background:var(--blue-light);color:var(--blue);border:1px solid rgba(45,158,255,.3)}
.pw-strength{height:3px;border-radius:2px;margin-top:5px;transition:.3s;width:0;background:var(--red)}
@media(max-width:700px){.grid2{grid-template-columns:1fr}.form-row{flex-direction:column}.topbar-logo .sub{display:none}}
</style></head><body>
<div class="topbar">
  <div class="topbar-logo">
    <div class="icon">&#9881;</div>
    <div>
      <div class="name"><em>MPCP</em> Admin Panel</div>
      <div class="sub">USER MANAGEMENT &amp; SYSTEM MASTERS</div>
    </div>
  </div>
  <div class="topbar-actions">
    {% if current_user_role == 'master_admin' %}<span class="badge b-master_admin">Master Admin</span>{% elif current_user_role == 'dept_admin' %}<span class="badge b-dept_admin">Dept Admin &mdash; {{ current_dept_name }}</span>{% endif %}
    <a href="/" class="tbtn">&#8592; Back to App</a>
    <a href="/logout" class="tbtn">Sign Out &#128682;</a>
  </div>
</div>
<div class="container">
  {% if msg %}<div class="{{ 'msg-ok' if msg_type=='ok' else 'msg-err' }}"><span>{{ '✓' if msg_type=='ok' else '⚠' }}</span>&nbsp;{{ msg }}</div>{% endif %}
  <div class="tab-bar">
    <button class="tab-btn active" onclick="switchAdminTab('users',this)">&#128100; Users</button>
    <button class="tab-btn" onclick="switchAdminTab('masters',this)">&#9881; Masters</button>
    <button class="tab-btn" onclick="switchAdminTab('audit',this)">&#128203; Audit Log</button>
  </div>

  <!-- USERS -->
  <div id="admin-tab-users" class="tab-panel active">
    <div class="stat-pills">
      <div class="stat-pill"><div class="val" style="color:var(--blue)">{{ users|length }}</div><div class="lbl">Total Users</div></div>
      <div class="stat-pill"><div class="val" style="color:var(--green)">{{ users|selectattr('active')|list|length }}</div><div class="lbl">Active</div></div>
      <div class="stat-pill"><div class="val" style="color:var(--violet)">{{ departments|length }}</div><div class="lbl">Departments</div></div>
      <div class="stat-pill"><div class="val" style="color:var(--red)">{{ users|selectattr('role','in',['master_admin','dept_admin'])|list|length }}</div><div class="lbl">Admins</div></div>
    </div>
    <div class="card">
      <div class="card-head"><h2>&#43; Create New User</h2></div>
      <form method="POST" action="/admin/users/create">
        <div class="form-row">
          <div class="fg"><label>Full Name *</label><input name="full_name" placeholder="Full name" required autocomplete="off"></div>
          <div class="fg"><label>Username *</label><input name="username" placeholder="username" required autocomplete="off"></div>
          <div class="fg"><label>Password *</label>
            <div class="pw-input-wrap">
              <input name="password" type="password" id="create-pw" placeholder="Min 6 chars" required oninput="checkPwStrength('create-pw','create-pw-str')">
              <button type="button" class="pw-toggle" onclick="togglePw('create-pw',this)">&#128065;</button>
            </div>
            <div class="pw-strength" id="create-pw-str"></div>
          </div>
          <div class="fg"><label>Role *</label>
            <select name="role">
              <option value="user">User</option>
              <option value="moderator">Moderator</option>
              {% if current_user_role in ('master_admin','dept_admin') %}<option value="dept_admin">Dept Admin</option>{% endif %}
              {% if current_user_role == 'master_admin' %}<option value="master_admin">Master Admin</option>{% endif %}
            </select>
          </div>
          <div class="fg"><label>Department</label>
            {% if current_user_role == 'master_admin' %}
            <select name="dept_code"><option value="">&#8212; None (Master) &#8212;</option>{% for d in departments %}<option value="{{ d.code }}">{{ d.name }}</option>{% endfor %}</select>
            {% else %}<input type="text" value="{{ current_dept_name }}" disabled style="opacity:.5"><input type="hidden" name="dept_code" value="{{ current_dept_code }}">{% endif %}
          </div>
          <div class="fg"><label>Emp Code</label><input name="emp_code" placeholder="EMP000XXX"></div>
          <div class="fg" style="min-width:auto;align-self:flex-end"><button class="btn btn-primary" type="submit">&#43; Create</button></div>
        </div>
      </form>
    </div>
    <div class="card">
      <div class="card-head">
        <h2>&#128100; All Users <span>({{ users|length }})</span></h2>
        <div style="display:flex;gap:8px;align-items:center">
          <input class="search-box" placeholder="Filter users..." oninput="filterUsers(this.value)" id="user-search">
          <select id="role-filter" onchange="filterUsers(document.getElementById('user-search').value)" style="padding:7px 10px;border:1.5px solid var(--border);border-radius:6px;font-size:11px;background:var(--bg-3);color:var(--text-2);font-family:var(--font);cursor:pointer;outline:none">
            <option value="">All Roles</option>
            <option value="master_admin">Master Admin</option>
            <option value="dept_admin">Dept Admin</option>
            <option value="moderator">Moderator</option>
            <option value="user">User</option>
          </select>
        </div>
      </div>
      <div style="overflow-x:auto"><table id="users-table">
        <thead><tr>
          <th>Name / Username</th><th>Department</th><th>Role</th><th>Emp Code</th><th>Status</th><th>Reset Password</th><th style="text-align:center">Actions</th>
        </tr></thead>
        <tbody>{% for u in users %}
        <tr data-role="{{ u.role }}" data-name="{{ u.full_name|lower }} {{ u.username|lower }}" data-dept="{{ (u.dept_name or '')|lower }}">
          <td>
            <div style="display:flex;align-items:center;gap:10px">
              <div style="width:30px;height:30px;border-radius:8px;background:linear-gradient(135deg,{% if u.role=='master_admin' %}#FF2D3A,#FF8C42{% elif u.role=='dept_admin' %}#2D9EFF,#1A7FD4{% else %}#8B5CF6,#5B21B6{% endif %});display:flex;align-items:center;justify-content:center;font-weight:800;font-size:12px;color:#fff;flex-shrink:0">{{ u.full_name[0]|upper }}</div>
              <div><div style="font-weight:700;font-size:12px;color:var(--text)">{{ u.full_name }}</div><div style="color:var(--muted);font-size:10px;font-family:'JetBrains Mono',monospace">@{{ u.username }}</div></div>
            </div>
          </td>
          <td>{% if u.dept_name %}<span class="dept-pill">{{ u.dept_name }}</span>{% else %}<span style="color:var(--muted)">&#8212;</span>{% endif %}</td>
          <td><span class="badge b-{{ u.role }}">{{ u.role.replace('_',' ') }}</span></td>
          <td style="font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--muted)">{{ u.emp_code or '&#8212;' }}</td>
          <td><span class="status-{{ 'active' if u.active else 'inactive' }}" style="display:flex;align-items:center;font-size:11px;font-weight:600;color:{{ '#10DF8A' if u.active else '#6B7A9F' }}"><span class="status-dot"></span>{{ 'Active' if u.active else 'Inactive' }}</span></td>
          <td>
            <form method="POST" action="/admin/users/{{ u.id }}/reset" style="display:flex;gap:6px;align-items:center">
              <div class="pw-input-wrap"><input name="new_password" type="password" placeholder="New password" style="padding:6px 32px 6px 10px;border:1.5px solid var(--border);border-radius:6px;font-size:11px;width:130px;background:var(--bg-3);color:var(--text);outline:none;transition:.2s" onfocus="this.style.borderColor='var(--border-glow)'" onblur="this.style.borderColor='var(--border)'"><button type="button" class="pw-toggle" onclick="togglePwEl(this.previousElementSibling,this)">&#128065;</button></div>
              <button class="btn btn-warn btn-sm" type="submit">Reset</button>
            </form>
          </td>
          <td>
            <div style="display:flex;gap:5px;justify-content:center;flex-wrap:wrap">
              <button class="btn btn-ghost btn-sm" onclick="openEdit('{{ u.id }}','{{ u.username }}','{{ u.full_name }}','{{ u.role }}','{{ u.dept_code or '' }}','{{ u.emp_code or '' }}')">&#9998; Edit</button>
              <form method="POST" action="/admin/users/{{ u.id }}/toggle" style="display:inline"><button class="btn btn-sm {{ 'btn-danger' if u.active else 'btn-success' }}" type="submit">{{ 'Disable' if u.active else 'Enable' }}</button></form>
              {% if current_user_role == 'master_admin' and u.username != 'admin' %}<button class="btn btn-danger btn-sm" onclick="confirmDelete('{{ u.id }}','{{ u.username }}')">&#128465;</button>{% endif %}
            </div>
          </td>
        </tr>{% endfor %}</tbody>
      </table></div>
    </div>
  </div>

  <!-- MASTERS -->
  <div id="admin-tab-masters" class="tab-panel">
    <div class="card">
      <div class="card-head"><h2>&#9881; System Masters <span>Manage all dropdown values used across the app</span></h2><button class="btn btn-primary btn-sm" onclick="openAddMaster()">&#43; Add Item</button></div>
      <div style="padding:18px 20px">
        <div class="masters-cats">
          <button class="cat-btn active" onclick="loadMasters('frequency',this)">&#9200; Frequencies</button>
          <button class="cat-btn" onclick="loadMasters('unit',this)">&#128207; Units / UoM</button>
          <button class="cat-btn" onclick="loadMasters('emp_level',this)">&#127970; Employee Levels</button>
          <button class="cat-btn" onclick="loadMasters('cp_source',this)">&#128203; CP Sources</button>
          <button class="cat-btn" onclick="loadMasters('location_type',this)">&#128205; Location Types</button>
        </div>
        <div id="masters-content"><p class="masters-empty">&#9881;<br>Select a category above.</p></div>
      </div>
    </div>
  </div>

  <!-- AUDIT -->
  <div id="admin-tab-audit" class="tab-panel">
    <div class="card">
      <div class="card-head">
        <h2>&#128203; Audit Log</h2>
        <div style="display:flex;gap:8px;align-items:center">
          <input class="search-box" placeholder="Filter activity..." oninput="filterAudit(this.value)">
          <button class="btn btn-ghost btn-sm" onclick="loadAudit()">&#8635; Refresh</button>
        </div>
      </div>
      <div style="overflow-x:auto"><table id="audit-table">
        <thead><tr><th style="width:130px">Time</th><th style="width:36px"></th><th>Activity</th><th style="text-align:right">IP</th></tr></thead>
        <tbody id="audit-tbody"><tr><td colspan="4" style="text-align:center;padding:30px;color:var(--muted)">Click Refresh to load audit records</td></tr></tbody>
      </table></div>
    </div>
  </div>
</div>

<!-- EDIT MODAL -->
<div class="modal-overlay" id="edit-modal">
  <div class="modal">
    <div class="modal-head"><h3>&#9998; Edit User</h3><button class="xbtn" onclick="closeEdit()">&#10005;</button></div>
    <form method="POST" id="edit-form">
      <div class="modal-body">
        <div class="grid2">
          <div class="mfg"><label>Full Name *</label><input name="name" id="e-name" required autocomplete="off"></div>
          <div class="mfg"><label>Username</label><input name="username" id="e-username" style="opacity:.5;cursor:not-allowed" readonly></div>
        </div>
        <div class="grid2">
          <div class="mfg"><label>Role *</label><select name="role" id="e-role"><option value="user">User</option><option value="moderator">Moderator</option>{% if current_user_role in ('master_admin','dept_admin') %}<option value="dept_admin">Dept Admin</option>{% endif %}{% if current_user_role == 'master_admin' %}<option value="master_admin">Master Admin</option>{% endif %}</select></div>
          <div class="mfg"><label>Department</label>{% if current_user_role == 'master_admin' %}<select name="dept_code" id="e-dept"><option value="">&#8212; None (Master) &#8212;</option>{% for d in departments %}<option value="{{ d.code }}">{{ d.name }}</option>{% endfor %}</select>{% else %}<input type="text" value="{{ current_dept_name }}" disabled style="opacity:.5"><input type="hidden" name="dept_code" value="{{ current_dept_code }}">{% endif %}</div>
        </div>
        <div class="mfg"><label>Emp Code</label><input name="emp_code" id="e-emp" placeholder="EMP000XXX"></div>
      </div>
      <div class="modal-foot"><button type="button" class="btn btn-ghost" onclick="closeEdit()">Cancel</button><button type="submit" class="btn btn-primary">Save Changes</button></div>
    </form>
  </div>
</div>

<!-- MASTERS MODAL -->
<div class="modal-overlay" id="master-modal">
  <div class="modal">
    <div class="modal-head"><h3 id="master-modal-title">&#43; Add Master Item</h3><button class="xbtn" onclick="closeMasterModal()">&#10005;</button></div>
    <div class="modal-body">
      <input type="hidden" id="master-id">
      <div class="mfg"><label>Category *</label><select id="master-cat"><option value="frequency">Frequencies</option><option value="unit">Units / UoM</option><option value="emp_level">Employee Levels</option><option value="cp_source">CP Sources</option><option value="location_type">Location Types</option></select></div>
      <div class="mfg"><label>Value * <span style="font-size:10px;color:var(--muted);font-weight:400">(stored internally, e.g. "Weekly")</span></label><input type="text" id="master-value" placeholder="e.g. Weekly" autocomplete="off"></div>
      <div class="mfg"><label>Display Label <span style="font-size:10px;color:var(--muted);font-weight:400">(shown in dropdowns — defaults to value)</span></label><input type="text" id="master-label" placeholder="e.g. Every Week"></div>
      <div class="mfg"><label>Sort Order</label><input type="number" id="master-sort" value="0" min="0"></div>
    </div>
    <div class="modal-foot"><button class="btn btn-ghost" onclick="closeMasterModal()">Cancel</button><button class="btn btn-primary" onclick="saveMaster()">Save</button></div>
  </div>
</div>

<!-- DELETE CONFIRM -->
<div class="modal-overlay" id="del-modal">
  <div class="modal" style="width:380px">
    <div class="modal-head" style="background:rgba(255,45,58,.12)"><h3>&#128465; Confirm Delete</h3><button class="xbtn" onclick="closeDelModal()">&#10005;</button></div>
    <div class="modal-body">
      <p style="font-size:13px;color:var(--text-2);margin-bottom:8px">Delete user <strong id="del-username" style="color:var(--red)"></strong>?</p>
      <p style="font-size:11px;color:var(--muted)">This cannot be undone. All assignments linked to this user will be removed.</p>
    </div>
    <div class="modal-foot"><button class="btn btn-ghost" onclick="closeDelModal()">Cancel</button><form method="POST" id="del-form" style="display:inline"><button class="btn" type="submit" style="background:var(--red);color:#fff;font-weight:700">Delete</button></form></div>
  </div>
</div>

<script>
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
function togglePw(id,btn){var el=document.getElementById(id);el.type=el.type==='password'?'text':'password';btn.textContent=el.type==='password'?'👁':'🙈'}
function togglePwEl(el,btn){el.type=el.type==='password'?'text':'password';btn.textContent=el.type==='password'?'👁':'🙈'}
function checkPwStrength(id,barId){
  var pw=document.getElementById(id)?.value||'';var bar=document.getElementById(barId);if(!bar)return;
  var s=0;if(pw.length>=6)s++;if(pw.length>=10)s++;if(/[A-Z]/.test(pw))s++;if(/[0-9]/.test(pw))s++;if(/[^A-Za-z0-9]/.test(pw))s++;
  bar.style.width=(Math.min(s/4,1)*100)+'%';bar.style.background=s<=1?'var(--red)':s<=3?'var(--amber)':'var(--green)';
}
function switchAdminTab(name,btn){
  document.querySelectorAll('.tab-panel').forEach(function(p){p.classList.remove('active')});
  document.querySelectorAll('.tab-btn').forEach(function(b){b.classList.remove('active')});
  document.getElementById('admin-tab-'+name).classList.add('active');btn.classList.add('active');
  if(name==='audit')loadAudit();
  if(name==='masters')loadMasters(_mastersCat,document.querySelector('.cat-btn.active'));
}
function filterUsers(q){
  var rf=document.getElementById('role-filter')?.value||'';q=(q||'').toLowerCase();
  document.querySelectorAll('#users-table tbody tr').forEach(function(r){
    r.style.display=((!q||(r.dataset.name||'').includes(q)||(r.dataset.dept||'').includes(q))&&(!rf||r.dataset.role===rf))?'':'none';
  });
}
function openEdit(id,username,name,role,dept,emp){
  document.getElementById('edit-form').action='/admin/users/'+id+'/edit';
  document.getElementById('e-name').value=name;document.getElementById('e-username').value=username;
  document.getElementById('e-role').value=role;var de=document.getElementById('e-dept');if(de)de.value=dept;
  document.getElementById('e-emp').value=emp;document.getElementById('edit-modal').classList.add('open');
  document.getElementById('e-name').focus();
}
function closeEdit(){document.getElementById('edit-modal').classList.remove('open')}
function confirmDelete(id,username){
  document.getElementById('del-username').textContent='@'+username;
  document.getElementById('del-form').action='/admin/users/'+id+'/delete';
  document.getElementById('del-modal').classList.add('open');
}
function closeDelModal(){document.getElementById('del-modal').classList.remove('open')}
var _mastersCat='frequency';
var catLabels={frequency:'Frequencies',unit:'Units / UoM',emp_level:'Employee Levels',cp_source:'CP Sources',location_type:'Location Types'};
var catIcons={frequency:'⏰',unit:'📐',emp_level:'🏗',cp_source:'📋',location_type:'📍'};
function loadMasters(cat,btn){
  _mastersCat=cat;
  document.querySelectorAll('.cat-btn').forEach(function(b){b.classList.remove('active')});
  if(btn)btn.classList.add('active');
  document.getElementById('master-cat').value=cat;
  var el=document.getElementById('masters-content');
  el.innerHTML='<div style="text-align:center;padding:32px;color:var(--muted);font-size:12px">Loading...</div>';
  fetch('/api/masters/'+cat,{credentials:'include'}).then(function(r){return r.json()}).then(function(rows){
    var html='<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">'
      +'<div style="font-size:12px;font-weight:600;color:var(--text-2)">'+(catIcons[cat]||'')+' '+catLabels[cat]
      +'<span style="color:var(--muted);font-weight:400;margin-left:8px">('+rows.length+' items)</span></div>'
      +'<button class="btn btn-primary btn-sm" onclick="openAddMaster()">&#43; Add Item</button></div>';
    if(!rows.length){
      html+='<div class="masters-empty" style="border:1px dashed var(--border);border-radius:8px;padding:32px">'
        +'<div style="font-size:28px;margin-bottom:8px">'+(catIcons[cat]||'⚙')+'</div>'
        +'No items yet. Add one to populate this dropdown in the app.</div>';
    } else {
      html+='<table><thead><tr><th>Value</th><th>Display Label</th><th>Sort</th><th>Actions</th></tr></thead><tbody>';
      rows.forEach(function(r){
        html+='<tr>'
          +'<td><code style="font-family:var(--font-mono);font-size:11px;color:var(--blue);background:var(--blue-light);padding:2px 8px;border-radius:4px">'+esc(r.value)+'</code></td>'
          +'<td style="color:var(--text)">'+esc(r.label||r.value)+'</td>'
          +'<td style="color:var(--muted);font-size:11px;font-family:var(--font-mono)">'+r.sort_order+'</td>'
          +'<td><div style="display:flex;gap:6px">'
          +'<button class="btn btn-ghost btn-sm" data-mid="'+esc(r.id)+'" data-mval="'+esc(r.value)+'" data-mlbl="'+esc(r.label||r.value)+'" data-msort="'+r.sort_order+'" onclick="openEditMasterFromBtn(this)">&#9998; Edit</button>'
          +'<button class="btn btn-danger btn-sm" data-mid="'+esc(r.id)+'" data-mval="'+esc(r.value)+'" onclick="deleteMasterFromBtn(this)">&#128465;</button>'
          +'</div></td></tr>';
      });
      html+='</tbody></table>';
    }
    el.innerHTML=html;
  }).catch(function(e){el.innerHTML='<p style="color:var(--red);padding:12px;font-size:12px">Error: '+esc(e.message)+'</p>'});
}
function openAddMaster(){
  document.getElementById('master-modal-title').textContent='+ Add Master Item';
  document.getElementById('master-id').value='';document.getElementById('master-cat').value=_mastersCat;
  document.getElementById('master-value').value='';document.getElementById('master-label').value='';
  document.getElementById('master-sort').value='0';document.getElementById('master-modal').classList.add('open');
  document.getElementById('master-value').focus();
}
function openEditMasterFromBtn(b){openEditMaster({id:b.dataset.mid,value:b.dataset.mval,label:b.dataset.mlbl,sort_order:b.dataset.msort})}
function deleteMasterFromBtn(b){deleteMaster(b.dataset.mid,b.dataset.mval)}
function openEditMaster(r){
  document.getElementById('master-modal-title').textContent='✏ Edit Item';
  document.getElementById('master-id').value=r.id;document.getElementById('master-cat').value=_mastersCat;
  document.getElementById('master-value').value=r.value;document.getElementById('master-label').value=r.label||r.value;
  document.getElementById('master-sort').value=r.sort_order||0;document.getElementById('master-modal').classList.add('open');
  document.getElementById('master-value').focus();
}
function closeMasterModal(){document.getElementById('master-modal').classList.remove('open')}
function saveMaster(){
  var mid=document.getElementById('master-id').value;var cat=document.getElementById('master-cat').value;
  var val=document.getElementById('master-value').value.trim();var lbl=document.getElementById('master-label').value.trim();
  var srt=parseInt(document.getElementById('master-sort').value)||0;
  if(!val){alert('Value is required');return}
  var url=mid?'/api/masters/'+mid:'/api/masters';var method=mid?'PUT':'POST';
  var body=mid?{value:val,label:lbl,sort_order:srt}:{category:cat,value:val,label:lbl,sort_order:srt};
  fetch(url,{method:method,credentials:'include',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)})
    .then(function(r){return r.json()}).then(function(d){
      if(d.ok||d.id){closeMasterModal();loadMasters(_mastersCat,null)}else alert(d.error||'Error saving');
    }).catch(function(e){alert('Error: '+e.message)});
}
function deleteMaster(id,val){
  if(!confirm('Delete "'+val+'"? This cannot be undone.'))return;
  fetch('/api/masters/'+id,{method:'DELETE',credentials:'include'}).then(function(r){return r.json()}).then(function(d){
    if(d.ok)loadMasters(_mastersCat,null);else alert(d.error||'Error');
  });
}
function filterAudit(q){
  q=q.toLowerCase();document.querySelectorAll('#audit-table tbody tr').forEach(function(r){r.style.display=r.textContent.toLowerCase().includes(q)?'':'none'});
}
function loadAudit(){
  var tb=document.getElementById('audit-tbody');
  tb.innerHTML="<tr><td colspan='4' style='text-align:center;padding:28px;color:var(--muted)'>Loading...</td></tr>";
  fetch('/api/audit_log').then(function(r){return r.json()}).then(function(rows){
    if(!rows.length){tb.innerHTML="<tr><td colspan='4' style='text-align:center;padding:28px;color:var(--muted)'>No audit records yet.</td></tr>";return}
    var icons={LOGIN:'🔑',LOGOUT:'🚪',USER_CREATE:'👤',USER_EDIT:'✏️',USER_ENABLE:'✅',USER_DISABLE:'🚫',PASSWORD_RESET:'🗝️',PASSWORD_CHANGE:'🔒',EMP_CREATE:'👷',EMP_UPDATE:'✏️',EMP_DELETE:'🗑️',MP_SAVE:'📋',MP_DELETE:'🗑️',CP_SAVE:'📌',CP_DELETE:'🗑️',PERF_IMPORT:'📊',DEPT_CREATE:'🏢',LOC_SAVE:'📍',MPCP_PUBLISH:'🚀',MPCP_STATUS:'🔄',MASTER_CREATE:'⚙️',MASTER_UPDATE:'✏️'};
    var bgMap={LOGIN:'rgba(45,158,255,.04)',USER_CREATE:'rgba(16,223,138,.04)',USER_DISABLE:'rgba(255,45,58,.04)',USER_ENABLE:'rgba(16,223,138,.04)',PASSWORD_RESET:'rgba(255,176,32,.04)',EMP_DELETE:'rgba(255,45,58,.04)',MPCP_PUBLISH:'rgba(139,92,246,.04)'};
    var tcMap={LOGIN:'var(--blue)',USER_CREATE:'var(--green)',USER_DISABLE:'var(--red)',USER_ENABLE:'var(--green)',USER_EDIT:'var(--amber)',PASSWORD_RESET:'var(--amber)',EMP_CREATE:'var(--green)',EMP_DELETE:'var(--red)',MP_SAVE:'var(--green)',MP_DELETE:'var(--red)',CP_SAVE:'var(--green)',CP_DELETE:'var(--red)',PERF_IMPORT:'var(--violet)',MPCP_PUBLISH:'var(--violet)',MASTER_CREATE:'var(--blue)',MASTER_UPDATE:'var(--amber)'};
    function desc(r){
      var a=r.action,who=r.actor_name||'System',d=r.detail||'';
      var m={LOGIN:who+' signed in'+(d.includes('master_admin')?' as Master Admin':d.includes('dept_admin')?' as Dept Admin':''),
        LOGOUT:who+' signed out',PASSWORD_CHANGE:'Password changed by '+who,PASSWORD_RESET:who+' reset a password',
        USER_CREATE:who+' created user: '+d.replace('Admin created:','').replace('API created:','').trim().split(' ')[0],
        USER_ENABLE:who+' enabled a user account',USER_DISABLE:who+' disabled a user account',USER_EDIT:who+' edited a user profile',
        EMP_CREATE:who+' added an employee record',EMP_UPDATE:who+' updated an employee',EMP_DELETE:who+' deleted an employee',
        MP_SAVE:who+' saved a Managing Point',MP_DELETE:who+' deleted an MP',CP_SAVE:who+' saved a Checking Point',CP_DELETE:who+' deleted a CP',
        PERF_IMPORT:who+' imported '+d.split(' ')[0]+' perf records',DEPT_CREATE:who+' created a department',LOC_SAVE:who+' saved a location',
        MPCP_PUBLISH:who+' published MPCP: '+d,MPCP_STATUS:who+' changed MPCP status',MASTER_CREATE:who+' added master: '+d,MASTER_UPDATE:who+' updated master: '+d};
      return m[a]||who+': '+a.replace(/_/g,' ').toLowerCase();
    }
    tb.innerHTML=rows.map(function(r){
      var ts=(r.ts||'').replace('T',' ').slice(0,16);
      return '<tr style="background:'+(bgMap[r.action]||'transparent')+';border-bottom:1px solid var(--border-light)">'
        +'<td style="padding:10px 14px;color:var(--muted);font-family:var(--font-mono);font-size:10px;white-space:nowrap">'+esc(ts)+'</td>'
        +'<td style="padding:10px 8px;text-align:center;font-size:14px">'+(icons[r.action]||'📝')+'</td>'
        +'<td style="padding:10px 14px;font-size:12px;font-weight:600;color:'+(tcMap[r.action]||'var(--text-2)')+'">'+esc(desc(r))+'</td>'
        +'<td style="padding:10px 14px;text-align:right;color:var(--muted);font-size:10px;font-family:var(--font-mono)">'+esc(r.ip||'')+'</td></tr>';
    }).join('');
  }).catch(function(e){tb.innerHTML="<tr><td colspan='4' style='color:var(--red);padding:12px'>Error: "+esc(e.message)+"</td></tr>"});
}
['edit-modal','master-modal','del-modal'].forEach(function(id){
  var el=document.getElementById(id);if(el)el.addEventListener('click',function(e){if(e.target===this)this.classList.remove('open')});
});
document.addEventListener('keydown',function(e){if(e.key==='Escape'){closeEdit();closeMasterModal();closeDelModal()}});
</script></body></html>"""

def admin_users_data(db, current_role, current_dept):
    if current_role == 'master_admin':
        users = R(db.execute(
            "SELECT u.*,d.name dept_name FROM users u LEFT JOIN departments d ON d.code=u.dept_code ORDER BY u.dept_code,u.full_name"
        ).fetchall())
    else:
        users = R(db.execute(
            "SELECT u.*,d.name dept_name FROM users u LEFT JOIN departments d ON d.code=u.dept_code WHERE u.dept_code=? ORDER BY u.full_name",
            (current_dept,)
        ).fetchall())
    depts = R(db.execute("SELECT code,name FROM departments WHERE active=1 ORDER BY name").fetchall())
    return users, depts


@app.route('/api/change_password', methods=['POST'])
def change_password():
    u = current_user()
    if not u: return json_error('Not authenticated', 401)
    d = request.json or {}
    cur = d.get('current','')
    new_pw = d.get('new_password','')
    if not cur or not new_pw:
        return json_error('All fields required')
    if len(new_pw) < 6:
        return json_error('Password must be at least 6 characters')
    db = get_master_conn()
    user = db.execute("SELECT * FROM users WHERE id=?", (u['id'],)).fetchone()
    if not user or not verify_password(cur, user['password_hash']):
        db.close()
        return json_error('Current password is incorrect', 403)
    db.execute("UPDATE users SET password_hash=? WHERE id=?",
               (hash_password(new_pw), u['id']))
    db.commit()
    log_audit('PASSWORD_CHANGE', 'user', u['id'], 'Self-service password change')
    db.close()
    return jsonify({'ok': True})

@app.route('/api/audit_log')
def get_audit_log():
    err = require_role('master_admin')
    if err: return err
    db = get_master_conn()
    rows = db.execute(
        "SELECT * FROM audit_log ORDER BY ts DESC LIMIT 200"
    ).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route('/admin')
@app.route('/admin/users')
def admin_panel():
    err = require_role('master_admin','dept_admin')
    if err: return err
    u = current_user()
    db = get_master_conn()
    users, depts = admin_users_data(db, u['role'], u.get('dept_code'))
    db.close()
    return render_template_string(ADMIN_HTML, users=users, departments=depts,
        current_user_role=u['role'],
        current_dept_code=u.get('dept_code',''),
        current_dept_name=next((d['name'] for d in depts if d['code']==u.get('dept_code')), 'All Departments'),
        msg=None, msg_type=None)

@app.route('/admin/users/create', methods=['POST'])
def admin_create_user():
    err = require_role('master_admin','dept_admin')
    if err: return err
    u = current_user()
    d = request.form
    pw = d.get('password','')
    if not d.get('full_name') or not d.get('username') or not pw:
        return _admin_msg('Full name, username and password required', 'err')
    if len(pw) < 6:
        return _admin_msg('Password must be at least 6 characters', 'err')
    role = d.get('role','user')
    if u['role'] == 'dept_admin':
        role = role if role in ('user','moderator','dept_admin') else 'user'
        dept_code = u['dept_code']  # dept_admin locked to own dept
    else:
        # master_admin can assign any dept (including None for master-level users)
        dept_code = d.get('dept_code','').strip() or None
    db = get_master_conn()
    try:
        new_id = uid()
        db.execute("INSERT INTO users VALUES(?,?,?,?,?,?,?,?,?)",
            (new_id, d['username'].strip().lower(), hash_password(pw),
             d['full_name'], role, dept_code, d.get('emp_code',''), 1,
             datetime.datetime.now().isoformat()))
        db.commit()
        log_audit('USER_CREATE','user',new_id,'Admin created: '+d['username'])
        return _admin_msg(f"User @{d['username']} created successfully", 'ok')
    except sqlite3.IntegrityError:
        return _admin_msg(f"Username @{d['username']} already exists", 'err')
    finally: db.close()

@app.route('/admin/users/<uid2>/delete', methods=['POST'])
def admin_delete_user(uid2):
    err = require_role('master_admin')
    if err: return jsonify({'ok': False, 'error': 'Access denied'}), 403
    u = current_user()
    if u['id'] == uid2:
        return jsonify({'ok': False, 'error': 'Cannot delete your own account'})
    db = get_master_conn()
    try:
        user = db.execute('SELECT username FROM users WHERE id=?', (uid2,)).fetchone()
        if not user:
            return jsonify({'ok': False, 'error': 'User not found'})
        if user['username'] == 'admin':
            return jsonify({'ok': False, 'error': 'Cannot delete master admin account'})
        db.execute('DELETE FROM users WHERE id=?', (uid2,))
        log_audit('USER_DELETE', 'user', uid2, f'Admin deleted user {user["username"]}')
        db.commit()
        return jsonify({'ok': True, 'msg': f'User {user["username"]} deleted'})
    except Exception as ex:
        return jsonify({'ok': False, 'error': str(ex)})
    finally:
        db.close()

@app.route('/admin/users/<uid2>/reset', methods=['POST'])
def admin_reset_pw(uid2):
    err = require_role('master_admin','dept_admin')
    if err: return err
    pw = request.form.get('new_password','')
    if len(pw) < 6:
        return _admin_msg('Password must be at least 6 characters', 'err')
    db = get_master_conn()
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(pw), uid2))
    log_audit('PASSWORD_RESET', 'user', uid2, 'Admin password reset')
    db.commit(); db.close()
    return _admin_msg('Password reset successfully', 'ok')


@app.route('/admin/users/<uid2>/edit', methods=['POST'])
def admin_edit_user(uid2):
    err = require_role('master_admin','dept_admin')
    if err: return err
    name     = request.form.get('name','').strip()
    username = request.form.get('username','').strip()
    role     = request.form.get('role','').strip()
    dept     = request.form.get('dept','').strip()
    emp      = request.form.get('emp_code','').strip()
    if not name or not username or not role:
        return _admin_msg('Name, username and role are required', 'err')
    db = get_master_conn()
    try:
        existing = db.execute('SELECT id FROM users WHERE username=? AND id!=?', (username, uid2)).fetchone()
        if existing:
            return _admin_msg(f'Username @{username} already exists', 'err')
        db.execute('UPDATE users SET full_name=?, username=?, role=?, dept_code=?, emp_code=? WHERE id=?',
                   (name, username, role, dept, emp, uid2))
        log_audit('USER_EDIT', 'user', uid2, f'Admin edited user {username}')
        db.commit()
        return _admin_msg('User updated successfully', 'ok')
    except Exception as ex:
        return _admin_msg(str(ex), 'err')
    finally:
        db.close()
@app.route('/admin/users/<uid2>/toggle', methods=['POST'])
def admin_toggle_user(uid2):
    err = require_role('master_admin','dept_admin')
    if err: return err
    u = current_user()
    db = get_master_conn()
    target = db.execute("SELECT * FROM users WHERE id=?", (uid2,)).fetchone()
    if not target: db.close(); return _admin_msg('User not found', 'err')
    if u['role']=='dept_admin' and dict(target).get('dept_code')!=u['dept_code']:
        db.close(); return _admin_msg('Access denied', 'err')
    new_active = 0 if dict(target)['active'] else 1
    db.execute("UPDATE users SET active=? WHERE id=?", (new_active, uid2))
    log_audit('USER_'+(' ENABLE' if new_active else 'DISABLE').strip(), 'user', uid2, f'active={new_active}')
    db.commit(); db.close()
    status = 'enabled' if new_active else 'disabled'
    return _admin_msg(f"User {status} successfully", 'ok')

def _admin_msg(msg, msg_type):
    u = current_user()
    db = get_master_conn()
    users, depts = admin_users_data(db, u['role'], u.get('dept_code'))
    db.close()
    return render_template_string(ADMIN_HTML, users=users, departments=depts,
        current_user_role=u['role'],
        current_dept_code=u.get('dept_code',''),
        current_dept_name=next((d['name'] for d in depts if d['code']==u.get('dept_code')), 'All Departments'),
        msg=msg, msg_type=msg_type)

if __name__ == '__main__':
    app.run(debug=False, host="0.0.0.0", port=7860)